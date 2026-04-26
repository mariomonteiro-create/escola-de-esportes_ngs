"""
SGE MasterPro Web — Backend Flask v2
Sistema de Gestão Esportiva — Colégio 7 de Setembro
Novidades v2:
  - Níveis de permissão (admin / editor / viewer)
  - Gestão de usuários pelo admin
  - Módulo Escolinhas (alunos + importação xlsx)
  - Staff por escolinha (professores e estagiários)
  - Dashboard com métricas gerais e escolinhas
  - Excel em todas as abas
  - waitress como servidor WSGI
  - Banco persistente em /data no Render
"""

from flask import Flask, request, jsonify, session, send_file, render_template
from flask_cors import CORS
from datetime import datetime, timedelta
import sqlite3, os, io, base64, tempfile, functools

try:
    from fpdf import FPDF
    PDF_OK = True
except ImportError:
    PDF_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    XLS_OK = True
except ImportError:
    XLS_OK = False

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "c7s_sge_secret_2026")

# ─────────────────────────────────────────────────────────────
# PATHS — banco persistente no Render, local no Windows/Linux dev
# ─────────────────────────────────────────────────────────────
IS_RENDER = os.environ.get("RENDER") == "true"
BASE_DIR  = os.path.dirname(os.path.abspath(__file__))

if IS_RENDER:
    DATA_DIR = "/data"                    # disco persistente montado no Render
    os.makedirs(DATA_DIR, exist_ok=True)  # cria /data se ainda não existir
else:
    DATA_DIR = BASE_DIR                   # pasta do projeto em dev local

DB_PATH = os.path.join(DATA_DIR, "SGE_MasterPro_V12.db")

app.config["SESSION_COOKIE_SECURE"]   = IS_RENDER
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_HTTPONLY"] = True

CORS(app, supports_credentials=True,
     origins=["http://localhost:5000","http://127.0.0.1:5000","https://*.onrender.com"])

NOME_ESCOLA  = "COLEGIO 7 DE SETEMBRO - NGS"
NOME_SISTEMA = "Sistema de Gestao Esportiva - Coordenacao de Cursos Livres"

# ─────────────────────────────────────────────────────────────
# DB
# ─────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db(); c = conn.cursor()

    # ── tabela usuarios com nivel de permissão ──
    c.execute("""CREATE TABLE IF NOT EXISTS usuarios
                 (user TEXT PRIMARY KEY, senha TEXT,
                  nivel TEXT DEFAULT 'editor',
                  nome_completo TEXT DEFAULT '')""")
    for col_sql in [
        "ALTER TABLE usuarios ADD COLUMN nivel TEXT DEFAULT 'editor'",
        "ALTER TABLE usuarios ADD COLUMN nome_completo TEXT DEFAULT ''",
    ]:
        try: c.execute(col_sql)
        except: pass
    # usuários padrão
    c.execute("INSERT OR IGNORE INTO usuarios (user,senha,nivel,nome_completo) VALUES ('admin','c7s2026','admin','Administrador')")
    c.execute("INSERT OR IGNORE INTO usuarios (user,senha,nivel,nome_completo) VALUES ('italo','esporte','editor','Italo')")

    # ── atletas ──
    c.execute("""CREATE TABLE IF NOT EXISTS atletas
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT, mod TEXT, cat TEXT, naipe TEXT)""")
    # ── tecnicos ──
    c.execute("""CREATE TABLE IF NOT EXISTS tecnicos
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  nome TEXT, mod TEXT, cat TEXT, naipe TEXT, cargo TEXT DEFAULT 'Tecnico')""")
    # ── jogos ──
    c.execute("""CREATE TABLE IF NOT EXISTS jogos
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  data TEXT, hora TEXT, mod TEXT, cat TEXT, adv TEXT, comp TEXT,
                  tipo TEXT, local TEXT, saida TEXT,
                  retorno TEXT DEFAULT '', placar_c7s TEXT DEFAULT '', placar_adv TEXT DEFAULT '')""")
    for col in ["ALTER TABLE jogos ADD COLUMN retorno TEXT DEFAULT ''",
                "ALTER TABLE jogos ADD COLUMN placar_c7s TEXT DEFAULT ''",
                "ALTER TABLE jogos ADD COLUMN placar_adv TEXT DEFAULT ''"]:
        try: c.execute(col)
        except: pass
    # ── convocacoes ──
    c.execute("""CREATE TABLE IF NOT EXISTS convocacoes
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  jogo_id INTEGER, atleta_id INTEGER, status TEXT DEFAULT 'Convocado',
                  UNIQUE(jogo_id, atleta_id))""")
    # ── listas dominio ──
    c.execute("CREATE TABLE IF NOT EXISTS modalidades (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT UNIQUE)")
    for m in ["Futsal","Volei","Basquete","Handebol"]:
        c.execute("INSERT OR IGNORE INTO modalidades (nome) VALUES (?)",(m,))
    c.execute("CREATE TABLE IF NOT EXISTS categorias (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT UNIQUE)")
    for cat in ["Sub-12","Sub-14","Sub-15","Sub-17","Livre"]:
        c.execute("INSERT OR IGNORE INTO categorias (nome) VALUES (?)",(cat,))
    c.execute("CREATE TABLE IF NOT EXISTS naipes (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT UNIQUE)")
    for n in ["Masculino","Feminino"]:
        c.execute("INSERT OR IGNORE INTO naipes (nome) VALUES (?)",(n,))

    # ── estoque ──
    c.execute("PRAGMA table_info(estoque)")
    cols_est = [r[1] for r in c.fetchall()]
    if "tamanho" in cols_est or "quantidade" in cols_est:
        c.execute("ALTER TABLE estoque RENAME TO estoque_legado")
        c.execute("CREATE TABLE estoque (id INTEGER PRIMARY KEY AUTOINCREMENT, tipo TEXT, nome TEXT, estoque_minimo INTEGER DEFAULT 5)")
        c.execute("INSERT OR IGNORE INTO estoque (tipo,nome,estoque_minimo) SELECT DISTINCT tipo,nome,5 FROM estoque_legado")
        conn.commit()
    c.execute("PRAGMA table_info(estoque)")
    cols_est = [r[1] for r in c.fetchall()]
    if not cols_est:
        c.execute("CREATE TABLE IF NOT EXISTS estoque (id INTEGER PRIMARY KEY AUTOINCREMENT, tipo TEXT, nome TEXT, estoque_minimo INTEGER DEFAULT 5)")
    if cols_est and "estoque_minimo" not in cols_est:
        try: c.execute("ALTER TABLE estoque ADD COLUMN estoque_minimo INTEGER DEFAULT 5")
        except: pass
    c.execute("""CREATE TABLE IF NOT EXISTS estoque_variantes
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  estoque_id INTEGER, tamanho TEXT DEFAULT '', numero TEXT DEFAULT '', quantidade INTEGER DEFAULT 0)""")
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='estoque_legado'")
    if c.fetchone():
        try:
            c.execute("""INSERT OR IGNORE INTO estoque_variantes (estoque_id,tamanho,numero,quantidade)
                         SELECT e.id,l.tamanho,l.numero,l.quantidade FROM estoque_legado l
                         JOIN estoque e ON e.tipo=l.tipo AND e.nome=l.nome WHERE l.tamanho!='' OR l.quantidade>0""")
            c.execute("DROP TABLE estoque_legado")
            conn.commit()
        except: pass

    # ── config ──
    c.execute("CREATE TABLE IF NOT EXISTS config (chave TEXT PRIMARY KEY, valor TEXT)")
    c.execute("INSERT OR IGNORE INTO config VALUES ('assinatura_direcao','Atenciosamente,\nCoordenacao de Cursos Livres\nCOLEGIO 7 DE SETEMBRO - NGS')")
    c.execute("INSERT OR IGNORE INTO config VALUES ('assinatura_trans','Atenciosamente,\nCoordenacao de Cursos Livres\nCOLEGIO 7 DE SETEMBRO - NGS')")
    c.execute("INSERT OR IGNORE INTO config VALUES ('logo_base64','')")

    # ── atividades ──
    c.execute("""CREATE TABLE IF NOT EXISTS atividades
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  tipo_atividade TEXT, nome_atividade TEXT, dia_semana TEXT,
                  horario_inicio TEXT, horario_fim TEXT, local TEXT,
                  professor TEXT, estagiario TEXT DEFAULT '', observacoes TEXT DEFAULT '')""")
    for t in ["atv_nomes","atv_profs","atv_estags","atv_quadras","atv_horarios"]:
        c.execute(f"CREATE TABLE IF NOT EXISTS {t} (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT UNIQUE)")
    for n in ["Futsal Escolinha","Volei Escolinha","Basquete Escolinha","Futsal Selecao","Volei Selecao","Ed. Fisica Turma A"]:
        c.execute("INSERT OR IGNORE INTO atv_nomes (nome) VALUES (?)",(n,))
    for q in ["Quadra Coberta","Quadra Externa","Ginasio","Piscina","Campo de Futebol"]:
        c.execute("INSERT OR IGNORE INTO atv_quadras (nome) VALUES (?)",(q,))
    for h in ["07:00","07:30","08:00","08:30","09:00","10:00","11:00","13:00","13:30","14:00","15:00","16:00","17:00","18:00","19:00"]:
        c.execute("INSERT OR IGNORE INTO atv_horarios (nome) VALUES (?)",(h,))

    # ── escolinhas ──
    c.execute("""CREATE TABLE IF NOT EXISTS escolinhas
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  nome TEXT UNIQUE NOT NULL)""")
    c.execute("""CREATE TABLE IF NOT EXISTS escolinha_alunos
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  escolinha_id INTEGER NOT NULL,
                  nome TEXT NOT NULL,
                  turma TEXT DEFAULT '',
                  responsavel TEXT DEFAULT '',
                  telefone TEXT DEFAULT '',
                  data_nasc TEXT DEFAULT '',
                  observacoes TEXT DEFAULT '',
                  FOREIGN KEY(escolinha_id) REFERENCES escolinhas(id))""")
    for col_sql in [
        "ALTER TABLE escolinha_alunos ADD COLUMN turma TEXT DEFAULT ''",
        "ALTER TABLE escolinha_alunos ADD COLUMN responsavel TEXT DEFAULT ''",
        "ALTER TABLE escolinha_alunos ADD COLUMN telefone TEXT DEFAULT ''",
        "ALTER TABLE escolinha_alunos ADD COLUMN data_nasc TEXT DEFAULT ''",
        "ALTER TABLE escolinha_alunos ADD COLUMN observacoes TEXT DEFAULT ''",
    ]:
        try: c.execute(col_sql)
        except: pass

    # ── staff das escolinhas (professores e estagiários por turma) ──
    c.execute("""CREATE TABLE IF NOT EXISTS escolinha_staff
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  escolinha_id INTEGER NOT NULL,
                  nome        TEXT NOT NULL,
                  cargo       TEXT DEFAULT 'Professor',
                  turma       TEXT DEFAULT '',
                  horario     TEXT DEFAULT '',
                  observacoes TEXT DEFAULT '',
                  FOREIGN KEY(escolinha_id) REFERENCES escolinhas(id))""")
    for col_sql in [
        "ALTER TABLE escolinha_staff ADD COLUMN turma TEXT DEFAULT ''",
        "ALTER TABLE escolinha_staff ADD COLUMN horario TEXT DEFAULT ''",
        "ALTER TABLE escolinha_staff ADD COLUMN observacoes TEXT DEFAULT ''",
    ]:
        try: c.execute(col_sql)
        except: pass

    conn.commit(); conn.close()

# ─────────────────────────────────────────────────────────────
# AUTH / PERMISSÕES
# ─────────────────────────────────────────────────────────────
def get_nivel():
    """Retorna o nível do usuário logado ou '' se não autenticado."""
    user = session.get("user")
    if not user: return ""
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT nivel FROM usuarios WHERE user=?", (user,))
    row = cur.fetchone(); conn.close()
    return row["nivel"] if row else "viewer"

def login_required(f):
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user"):
            return jsonify({"error":"Nao autenticado"}), 401
        return f(*args, **kwargs)
    return wrapper

def admin_required(f):
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user"):
            return jsonify({"error":"Nao autenticado"}), 401
        if get_nivel() != "admin":
            return jsonify({"error":"Permissao insuficiente"}), 403
        return f(*args, **kwargs)
    return wrapper

def editor_required(f):
    """Permite admin e editor; bloqueia viewer."""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user"):
            return jsonify({"error":"Nao autenticado"}), 401
        if get_nivel() == "viewer":
            return jsonify({"error":"Permissao insuficiente - apenas leitura"}), 403
        return f(*args, **kwargs)
    return wrapper

# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def parse_data(s):
    try: return datetime.strptime(s, "%d/%m/%Y").date()
    except: return datetime.max.date()

def safe(s):
    if s is None: return ""
    return str(s).encode("latin-1", errors="replace").decode("latin-1")

def get_logo_tempfile():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT valor FROM config WHERE chave='logo_base64'")
        row = cur.fetchone(); conn.close()
        if not row or not row["valor"]: return None
        data = base64.b64decode(row["valor"])
        ext  = ".jpg" if data[:2] == b'\xff\xd8' else ".png"
        tmp  = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
        tmp.write(data); tmp.close(); return tmp.name
    except: return None

def cleanup(path):
    try:
        if path and os.path.exists(path): os.unlink(path)
    except: pass

def calc_hora_saida(hora_jogo, ant):
    if ant in ["1h antes","2h antes"]:
        try:
            h = datetime.strptime(hora_jogo,"%H:%M") - timedelta(hours=1 if ant=="1h antes" else 2)
            return h.strftime("%H:%M")
        except: pass
    return hora_jogo

def get_jogos_futuros():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT id,data,hora,mod,cat,adv,local,tipo,comp,saida,retorno,placar_c7s,placar_adv FROM jogos")
    rows = cur.fetchall(); conn.close()
    futuros = []; hoje = datetime.now().date()
    for r in rows:
        try:
            dt = datetime.strptime(f"{r['data']} {r['hora']}", "%d/%m/%Y %H:%M")
            if dt.date() >= hoje:
                futuros.append({"id":r["id"],"data":r["data"],"hora":r["hora"],
                    "mod":r["mod"],"cat":r["cat"],"adv":r["adv"],"local":r["local"],
                    "tipo":r["tipo"],"comp":r["comp"],"saida_ant":r["saida"],"retorno":r["retorno"],
                    "placar":f"{r['placar_c7s']} x {r['placar_adv']}" if r["placar_c7s"] else "","dt_obj":dt})
        except: pass
    futuros.sort(key=lambda x: x["dt_obj"]); return futuros

# ─────────────────────────────────────────────────────────────
# PDF HELPER
# ─────────────────────────────────────────────────────────────
def make_pdf(logo_path=None):
    class PDF_Elite(FPDF):
        def header(self):
            tx = 52 if logo_path else 10
            if logo_path:
                try: self.image(logo_path, 10, 6, 38)
                except: pass
            self.set_xy(tx, 9)
            self.set_font("Arial","B",16); self.set_text_color(34,28,137)
            self.cell(200-tx, 9, safe(NOME_ESCOLA), 0, 2, "L")
            self.set_x(tx); self.set_font("Arial","",9); self.set_text_color(100,100,100)
            self.cell(200-tx, 5, safe(NOME_SISTEMA), 0, 1, "L")
            self.set_draw_color(242,179,26); self.set_line_width(0.8)
            self.line(10,38,200,38); self.ln(12)
        def add_watermark(self):
            if logo_path:
                try: self.image(logo_path, x=45, y=100, w=120)
                except: pass
        def footer(self):
            self.set_y(-15); self.set_font("Arial","I",8); self.set_text_color(150,150,150)
            self.cell(0,10,f'Doc. Oficial C7S - {datetime.now().strftime("%d/%m/%Y %H:%M")} - Pag {self.page_no()}/{{nb}}',0,0,"C")
    pdf = PDF_Elite(); pdf.alias_nb_pages(); return pdf

def pdf_to_response(pdf, filename):
    buf = io.BytesIO()
    try: pdf.output(buf)
    except TypeError:
        raw = pdf.output(dest="S")
        buf.write(raw.encode("latin-1") if isinstance(raw, str) else raw)
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=filename)

def xls_header(ws, titulo, subtitulo, ncols):
    """Cabeçalho padrão para planilhas Excel."""
    azul = PatternFill("solid", fgColor="221C89")
    aln_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
    col_letra = chr(64 + ncols)
    ws.merge_cells(f"A1:{col_letra}1"); ws["A1"] = NOME_ESCOLA
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="221C89"); ws["A1"].alignment = aln_c
    ws.merge_cells(f"A2:{col_letra}2"); ws["A2"] = titulo
    ws["A2"].font = Font(name="Arial", bold=True, size=11, color="4A4A6A"); ws["A2"].alignment = aln_c
    ws.merge_cells(f"A3:{col_letra}3"); ws["A3"] = subtitulo
    ws["A3"].font = Font(name="Arial", italic=True, size=9, color="888888"); ws["A3"].alignment = aln_c
    ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 18; ws.append([])
    return azul, aln_c

def xls_add_headers(ws, headers, azul, aln_c):
    borda = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    ws.append(headers); hr = ws.max_row
    for i in range(1, len(headers)+1):
        cel = ws.cell(hr, i)
        cel.fill = azul; cel.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cel.alignment = aln_c; cel.border = borda
    ws.row_dimensions[hr].height = 24
    return borda

def xls_add_row(ws, vals, borda, aln_c, idx):
    cinza  = PatternFill("solid", fgColor="EEF0F5")
    branco = PatternFill("solid", fgColor="FFFFFF")
    ws.append(vals); r = ws.max_row
    fill = cinza if idx % 2 == 0 else branco
    for ci in range(1, len(vals)+1):
        cel = ws.cell(r, ci)
        cel.fill = fill; cel.font = Font(name="Arial", size=10)
        cel.alignment = aln_c; cel.border = borda
    ws.row_dimensions[r].height = 20

# ─────────────────────────────────────────────────────────────
# AUTH ROUTES
# ─────────────────────────────────────────────────────────────
@app.route("/api/auth/login", methods=["POST"])
def login():
    d = request.json or {}
    user = d.get("user","").strip(); senha = d.get("senha","").strip()
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT user,nivel FROM usuarios WHERE user=? AND senha=?",(user,senha))
    row = cur.fetchone(); conn.close()
    if row:
        session["user"] = user
        return jsonify({"ok":True,"user":user,"nivel":row["nivel"]})
    return jsonify({"error":"Usuario ou senha incorretos"}), 401

@app.route("/api/auth/logout", methods=["POST"])
def logout():
    session.clear(); return jsonify({"ok":True})

@app.route("/api/auth/me")
def me():
    user = session.get("user")
    if not user: return jsonify({"user":None,"nivel":""})
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT nivel,nome_completo FROM usuarios WHERE user=?", (user,))
    row = cur.fetchone(); conn.close()
    return jsonify({"user":user,"nivel":row["nivel"] if row else "","nome":row["nome_completo"] if row else ""})

# ─────────────────────────────────────────────────────────────
# GESTÃO DE USUÁRIOS (somente admin)
# ─────────────────────────────────────────────────────────────
@app.route("/api/usuarios", methods=["GET"])
@admin_required
def get_usuarios():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT user,nivel,nome_completo FROM usuarios ORDER BY user")
    rows = [dict(r) for r in cur.fetchall()]; conn.close()
    return jsonify(rows)

@app.route("/api/usuarios", methods=["POST"])
@admin_required
def add_usuario():
    d = request.json or {}
    user  = d.get("user","").strip()
    senha = d.get("senha","").strip()
    nivel = d.get("nivel","editor")
    nome  = d.get("nome_completo","").strip()
    if not user or not senha: return jsonify({"error":"Usuario e senha sao obrigatorios"}), 400
    if nivel not in ("admin","editor","viewer"): return jsonify({"error":"Nivel invalido"}), 400
    conn = get_db(); cur = conn.cursor()
    try:
        cur.execute("INSERT INTO usuarios (user,senha,nivel,nome_completo) VALUES (?,?,?,?)",(user,senha,nivel,nome))
        conn.commit()
    except: conn.close(); return jsonify({"error":"Usuario ja existe"}), 409
    conn.close(); return jsonify({"ok":True}), 201

@app.route("/api/usuarios/<uid>", methods=["PUT"])
@admin_required
def update_usuario(uid):
    d = request.json or {}
    nivel = d.get("nivel","editor")
    nome  = d.get("nome_completo","").strip()
    if nivel not in ("admin","editor","viewer"): return jsonify({"error":"Nivel invalido"}), 400
    conn = get_db(); cur = conn.cursor()
    fields = ["nivel=?","nome_completo=?"]; vals = [nivel, nome]
    if d.get("senha","").strip():
        fields.append("senha=?"); vals.append(d["senha"].strip())
    vals.append(uid)
    cur.execute(f"UPDATE usuarios SET {','.join(fields)} WHERE user=?", vals)
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/usuarios/<uid>", methods=["DELETE"])
@admin_required
def delete_usuario(uid):
    if uid == session.get("user"): return jsonify({"error":"Nao pode excluir a si mesmo"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM usuarios WHERE user=?",(uid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# ─────────────────────────────────────────────────────────────
# LOGO
# ─────────────────────────────────────────────────────────────
@app.route("/api/logo", methods=["GET"])
def get_logo():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT valor FROM config WHERE chave='logo_base64'")
    row = cur.fetchone(); conn.close()
    return jsonify({"logo": row["valor"] if row else ""})

@app.route("/api/logo", methods=["POST"])
@editor_required
def upload_logo():
    if "file" not in request.files: return jsonify({"error":"Nenhum arquivo"}), 400
    f = request.files["file"]; data = f.read()
    if len(data) > 2*1024*1024: return jsonify({"error":"Arquivo muito grande (max 2 MB)"}), 400
    b64 = base64.b64encode(data).decode()
    conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT OR REPLACE INTO config (chave,valor) VALUES ('logo_base64',?)",(b64,))
    conn.commit(); conn.close()
    return jsonify({"ok":True,"logo":b64})

@app.route("/api/logo", methods=["DELETE"])
@editor_required
def delete_logo():
    conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT OR REPLACE INTO config (chave,valor) VALUES ('logo_base64','')")
    conn.commit(); conn.close(); return jsonify({"ok":True})

# ─────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────
@app.route("/api/dashboard")
@login_required
def dashboard():
    conn = get_db(); cur = conn.cursor()

    # ── atletas ──
    cur.execute("SELECT COUNT(*) as t FROM atletas")
    total_atletas = cur.fetchone()["t"]

    cur.execute("SELECT mod, COUNT(*) as total FROM atletas GROUP BY mod ORDER BY total DESC")
    atletas_por_mod = [dict(r) for r in cur.fetchall()]

    cur.execute("SELECT cat, COUNT(*) as total FROM atletas GROUP BY cat ORDER BY total DESC")
    atletas_por_cat = [dict(r) for r in cur.fetchall()]

    # ── jogos ──
    cur.execute("SELECT COUNT(*) as t FROM jogos")
    total_jogos = cur.fetchone()["t"]

    cur.execute("SELECT COUNT(*) as t FROM jogos WHERE placar_c7s != ''")
    jogos_realizados = cur.fetchone()["t"]

    # ── resultados gerais ──
    cur.execute("SELECT placar_c7s, placar_adv FROM jogos WHERE placar_c7s != ''")
    vit = emp = der = 0
    for r in cur.fetchall():
        try:
            pc, pa = int(r["placar_c7s"]), int(r["placar_adv"])
            if pc > pa:    vit += 1
            elif pc == pa: emp += 1
            else:          der += 1
        except: pass

    # ── estoque crítico ──
    cur.execute("""SELECT e.nome, e.tipo, e.estoque_minimo,
                          COALESCE(SUM(v.quantidade), 0) as total
                   FROM estoque e
                   LEFT JOIN estoque_variantes v ON v.estoque_id = e.id
                   GROUP BY e.id
                   HAVING total = 0 OR (total * 1.0 / MAX(e.estoque_minimo, 1)) <= 0.5
                   ORDER BY total ASC
                   LIMIT 8""")
    estoque_critico = [dict(r) for r in cur.fetchall()]

    # ── escolinhas com total de alunos e staff ──
    cur.execute("""SELECT e.id, e.nome,
                          COUNT(DISTINCT a.id) as total_alunos,
                          COUNT(DISTINCT s.id) as total_staff
                   FROM escolinhas e
                   LEFT JOIN escolinha_alunos a ON a.escolinha_id = e.id
                   LEFT JOIN escolinha_staff  s ON s.escolinha_id = e.id
                   GROUP BY e.id
                   ORDER BY e.nome""")
    escolinhas = [dict(r) for r in cur.fetchall()]
    total_alunos = sum(e["total_alunos"] for e in escolinhas)

    # ── técnicos (comissão técnica das seleções) ──
    cur.execute("SELECT COUNT(*) as t FROM tecnicos")
    total_tecnicos = cur.fetchone()["t"]

    cur.execute("SELECT cargo, COUNT(*) as t FROM tecnicos GROUP BY cargo ORDER BY t DESC")
    tec_por_cargo = [dict(r) for r in cur.fetchall()]

    # ── staff das escolinhas por cargo ──
    cur.execute("""SELECT cargo, COUNT(*) as t
                   FROM escolinha_staff GROUP BY cargo ORDER BY t DESC""")
    staff_cargos = {r["cargo"]: r["t"] for r in cur.fetchall()}
    total_professores   = staff_cargos.get("Professor",   0)
    total_estagiarios   = staff_cargos.get("Estagiario",  0)
    total_coordenadores = staff_cargos.get("Coordenador", 0)
    total_auxiliares    = staff_cargos.get("Auxiliar",    0)
    total_staff_esc     = sum(staff_cargos.values())

    conn.close()
    return jsonify({
        "atletas": {
            "total":   total_atletas,
            "por_mod": atletas_por_mod,
            "por_cat": atletas_por_cat,
        },
        "jogos": {
            "total":      total_jogos,
            "realizados": jogos_realizados,
            "vitorias":   vit,
            "empates":    emp,
            "derrotas":   der,
        },
        "estoque_critico": estoque_critico,
        "escolinhas": {
            "lista":        escolinhas,
            "total_alunos": total_alunos,
        },
        "equipe": {
            "tecnicos":        total_tecnicos,
            "tec_por_cargo":   tec_por_cargo,
            "professores":     total_professores,
            "estagiarios":     total_estagiarios,
            "coordenadores":   total_coordenadores,
            "auxiliares":      total_auxiliares,
            "total_staff_esc": total_staff_esc,
        },
    })

# ─────────────────────────────────────────────────────────────
# ATLETAS
# ─────────────────────────────────────────────────────────────
@app.route("/api/atletas", methods=["GET"])
@login_required
def get_atletas():
    conn = get_db(); cur = conn.cursor()
    q = "SELECT id,nome,mod,cat,naipe FROM atletas"; p = []; conds = []
    for k in ["mod","cat","naipe"]:
        v = request.args.get(k)
        if v: conds.append(f"{k}=?"); p.append(v)
    if conds: q += " WHERE " + " AND ".join(conds)
    cur.execute(q + " ORDER BY nome ASC", p)
    rows = [dict(r) for r in cur.fetchall()]; conn.close(); return jsonify(rows)

@app.route("/api/atletas", methods=["POST"])
@editor_required
def add_atleta():
    d = request.json or {}
    nome = d.get("nome","").strip()
    if not nome: return jsonify({"error":"Nome obrigatorio"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT INTO atletas (nome,mod,cat,naipe) VALUES (?,?,?,?)",
                (nome, d.get("mod","Futsal"), d.get("cat","Sub-15"), d.get("naipe","Masculino")))
    conn.commit(); new_id = cur.lastrowid; conn.close()
    return jsonify({"id":new_id}), 201

@app.route("/api/atletas/<int:aid>", methods=["PUT"])
@editor_required
def update_atleta(aid):
    d = request.json or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE atletas SET nome=?,mod=?,cat=?,naipe=? WHERE id=?",
                (d.get("nome"),d.get("mod"),d.get("cat"),d.get("naipe"),aid))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/atletas/<int:aid>", methods=["DELETE"])
@editor_required
def delete_atleta(aid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM atletas WHERE id=?", (aid,))
    cur.execute("DELETE FROM convocacoes WHERE atleta_id=?", (aid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/excel/atletas")
@login_required
def excel_atletas():
    if not XLS_OK: return "Instale openpyxl", 500
    mod   = request.args.get("mod",""); cat = request.args.get("cat",""); naipe = request.args.get("naipe","")
    conn = get_db(); cur = conn.cursor()
    q = "SELECT nome,mod,cat,naipe FROM atletas"; p = []; conds = []
    for k,v in [("mod",mod),("cat",cat),("naipe",naipe)]:
        if v: conds.append(f"{k}=?"); p.append(v)
    if conds: q += " WHERE "+" AND ".join(conds)
    cur.execute(q+" ORDER BY nome ASC", p)
    rows = cur.fetchall(); conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Atletas"
    sub = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}" + (f" | Filtros: {mod} {cat} {naipe}".strip() if mod or cat or naipe else "")
    azul, aln_c = xls_header(ws, "LISTA DE ATLETAS", sub, 4)
    borda = xls_add_headers(ws, ["Nome","Modalidade","Categoria","Naipe"], azul, aln_c)
    for i,r in enumerate(rows,1): xls_add_row(ws, [r["nome"],r["mod"],r["cat"],r["naipe"]], borda, aln_c, i)
    for col,w in zip("ABCD",[40,15,12,12]): ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"Atletas_{datetime.now().strftime('%Y%m%d')}.xlsx")

# ─────────────────────────────────────────────────────────────
# TECNICOS
# ─────────────────────────────────────────────────────────────
@app.route("/api/tecnicos", methods=["GET"])
@login_required
def get_tecnicos():
    conn = get_db(); cur = conn.cursor()
    q = "SELECT id,nome,mod,cat,naipe,cargo FROM tecnicos"; p = []; conds = []
    for k in ["mod","cat","naipe"]:
        v = request.args.get(k)
        if v: conds.append(f"{k}=?"); p.append(v)
    if conds: q += " WHERE " + " AND ".join(conds)
    cur.execute(q + " ORDER BY nome ASC", p)
    rows = [dict(r) for r in cur.fetchall()]; conn.close(); return jsonify(rows)

@app.route("/api/tecnicos", methods=["POST"])
@editor_required
def add_tecnico():
    d = request.json or {}
    nome = d.get("nome","").strip()
    if not nome: return jsonify({"error":"Nome obrigatorio"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT INTO tecnicos (nome,mod,cat,naipe,cargo) VALUES (?,?,?,?,?)",
                (nome,d.get("mod","Futsal"),d.get("cat","Sub-15"),d.get("naipe","Masculino"),d.get("cargo","Tecnico")))
    conn.commit(); new_id = cur.lastrowid; conn.close()
    return jsonify({"id":new_id}), 201

@app.route("/api/tecnicos/<int:tid>", methods=["PUT"])
@editor_required
def update_tecnico(tid):
    d = request.json or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE tecnicos SET nome=?,mod=?,cat=?,naipe=?,cargo=? WHERE id=?",
                (d.get("nome"),d.get("mod"),d.get("cat"),d.get("naipe"),d.get("cargo"),tid))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/tecnicos/<int:tid>", methods=["DELETE"])
@editor_required
def delete_tecnico(tid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM tecnicos WHERE id=?",(tid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/excel/tecnicos")
@login_required
def excel_tecnicos():
    if not XLS_OK: return "Instale openpyxl", 500
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT nome,cargo,mod,cat,naipe FROM tecnicos ORDER BY nome")
    rows = cur.fetchall(); conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Comissao Tecnica"
    azul, aln_c = xls_header(ws,"COMISSÃO TÉCNICA",f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",5)
    borda = xls_add_headers(ws,["Nome","Cargo","Modalidade","Categoria","Naipe"],azul,aln_c)
    for i,r in enumerate(rows,1): xls_add_row(ws,[r["nome"],r["cargo"],r["mod"],r["cat"],r["naipe"]],borda,aln_c,i)
    for col,w in zip("ABCDE",[38,22,15,12,12]): ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"ComissaoTecnica_{datetime.now().strftime('%Y%m%d')}.xlsx")

# ─────────────────────────────────────────────────────────────
# JOGOS
# ─────────────────────────────────────────────────────────────
@app.route("/api/jogos", methods=["GET"])
@login_required
def get_jogos():
    conn = get_db(); cur = conn.cursor()
    q = "SELECT id,data,hora,mod,cat,adv,comp,tipo,local,saida,retorno,placar_c7s,placar_adv FROM jogos"
    p = []; conds = []
    if request.args.get("todos","0") != "1":
        for k in ["mod","cat"]:
            v = request.args.get(k)
            if v: conds.append(f"{k}=?"); p.append(v)
    if conds: q += " WHERE " + " AND ".join(conds)
    cur.execute(q, p)
    rows = []
    for r in cur.fetchall():
        d = dict(r)
        d["placar"] = f"{d['placar_c7s']} x {d['placar_adv']}" if d["placar_c7s"] else ""
        d["sort_key"] = str(parse_data(d["data"])); rows.append(d)
    conn.close(); rows.sort(key=lambda x: x["sort_key"]); return jsonify(rows)

@app.route("/api/jogos", methods=["POST"])
@editor_required
def add_jogo():
    d = request.json or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT INTO jogos (data,hora,mod,cat,adv,comp,tipo,local,saida,retorno) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (d.get("data"),d.get("hora"),d.get("mod"),d.get("cat"),d.get("adv"),
                 d.get("comp"),d.get("tipo","CASA"),d.get("local"),d.get("saida","No horario"),d.get("retorno","")))
    conn.commit(); new_id = cur.lastrowid; conn.close()
    return jsonify({"id":new_id}), 201

@app.route("/api/jogos/<int:jid>", methods=["PUT"])
@editor_required
def update_jogo(jid):
    d = request.json or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE jogos SET data=?,hora=?,mod=?,cat=?,adv=?,comp=?,tipo=?,local=?,saida=?,retorno=? WHERE id=?",
                (d.get("data"),d.get("hora"),d.get("mod"),d.get("cat"),d.get("adv"),
                 d.get("comp"),d.get("tipo"),d.get("local"),d.get("saida"),d.get("retorno"),jid))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/jogos/<int:jid>", methods=["DELETE"])
@editor_required
def delete_jogo(jid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM jogos WHERE id=?",(jid,))
    cur.execute("DELETE FROM convocacoes WHERE jogo_id=?",(jid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/jogos/<int:jid>/placar", methods=["PATCH"])
@editor_required
def set_placar(jid):
    d = request.json or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE jogos SET placar_c7s=?,placar_adv=? WHERE id=?",
                (d.get("placar_c7s",""),d.get("placar_adv",""),jid))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/excel/jogos")
@login_required
def excel_jogos():
    if not XLS_OK: return "Instale openpyxl", 500
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT data,hora,mod,cat,adv,comp,tipo,local,saida,retorno,placar_c7s,placar_adv FROM jogos")
    rows = sorted(cur.fetchall(), key=lambda r: parse_data(str(r[0]))); conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Agenda"
    azul, aln_c = xls_header(ws,"AGENDA DE JOGOS",f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",10)
    borda = xls_add_headers(ws,["Data","Hora","Mod","Cat","Adversário","Competição","Mando","Local","Retorno","Placar"],azul,aln_c)
    for i,r in enumerate(rows,1):
        placar = f"{r['placar_c7s']}x{r['placar_adv']}" if r["placar_c7s"] else "-"
        xls_add_row(ws,[r["data"],r["hora"],r["mod"],r["cat"],r["adv"],r["comp"] or "-",
                        r["tipo"],r["local"] or "-",r["retorno"] or "-",placar],borda,aln_c,i)
    for col,w in zip("ABCDEFGHIJ",[12,8,12,10,25,25,8,28,12,10]): ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"Agenda_{datetime.now().strftime('%Y%m%d')}.xlsx")

# ─────────────────────────────────────────────────────────────
# CONVOCACOES
# ─────────────────────────────────────────────────────────────
@app.route("/api/convocacoes/<int:jogo_id>", methods=["GET"])
@login_required
def get_convocacao(jogo_id):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT mod,cat,naipe FROM jogos WHERE id=?",(jogo_id,))
    jogo = cur.fetchone()
    if not jogo: conn.close(); return jsonify([])
    cur.execute("SELECT id,nome FROM atletas WHERE mod=? AND cat=? AND naipe=? ORDER BY nome ASC",
                (jogo["mod"],jogo["cat"],jogo["naipe"]))
    atletas = cur.fetchall(); result = []
    for a in atletas:
        cur.execute("SELECT status FROM convocacoes WHERE jogo_id=? AND atleta_id=?",(jogo_id,a["id"]))
        row = cur.fetchone()
        if not row:
            cur.execute("INSERT OR IGNORE INTO convocacoes (jogo_id,atleta_id,status) VALUES (?,?,?)",(jogo_id,a["id"],"Convocado"))
            conn.commit(); status = "Convocado"
        else: status = row["status"]
        result.append({"id":a["id"],"nome":a["nome"],"status":status})
    conn.close(); return jsonify(result)

@app.route("/api/convocacoes/<int:jogo_id>/<int:atleta_id>", methods=["PATCH"])
@editor_required
def update_convocacao(jogo_id, atleta_id):
    d = request.json or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE convocacoes SET status=? WHERE jogo_id=? AND atleta_id=?",
                (d.get("status","Convocado"),jogo_id,atleta_id))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/excel/convocacao/<int:jogo_id>")
@login_required
def excel_convocacao(jogo_id):
    if not XLS_OK: return "Instale openpyxl", 500
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT data,hora,adv,local,mod,cat,naipe FROM jogos WHERE id=?",(jogo_id,))
    jogo = cur.fetchone()
    if not jogo: return "Jogo nao encontrado", 404
    cur.execute("SELECT a.nome,cv.status FROM atletas a JOIN convocacoes cv ON a.id=cv.atleta_id WHERE cv.jogo_id=? ORDER BY a.nome ASC",(jogo_id,))
    lista = cur.fetchall(); conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Convocacao"
    titulo = f"CONVOCAÇÃO — C7S x {jogo['adv']} — {jogo['data']}"
    sub = f"{jogo['mod']} {jogo['cat']} {jogo['naipe']} | Local: {jogo['local']} | Hora: {jogo['hora']}"
    azul, aln_c = xls_header(ws, titulo, sub, 3)
    borda = xls_add_headers(ws,["#","Nome do Atleta","Status"],azul,aln_c)
    for i,(nome,status) in enumerate(lista,1): xls_add_row(ws,[i,nome,status],borda,aln_c,i)
    for col,w in zip("ABC",[5,40,15]): ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"Convocacao_{jogo['mod']}_{jogo['data'].replace('/','_')}.xlsx")

# ─────────────────────────────────────────────────────────────
# ESTATISTICAS
# ─────────────────────────────────────────────────────────────
@app.route("/api/estatisticas", methods=["GET"])
@login_required
def get_estatisticas():
    mod = request.args.get("mod",""); cat = request.args.get("cat","")
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT data,adv,tipo,placar_c7s,placar_adv FROM jogos WHERE mod=? AND cat=? AND placar_c7s!='' ORDER BY data ASC",(mod,cat))
    rows = cur.fetchall(); conn.close()
    vit=emp=der=gp=gc=0; jogos=[]
    for r in rows:
        try:
            pi,ai = int(r["placar_c7s"]),int(r["placar_adv"])
            gp+=pi; gc+=ai
            if pi>ai: res="Vitoria"; vit+=1
            elif pi==ai: res="Empate"; emp+=1
            else: res="Derrota"; der+=1
            jogos.append({"data":r["data"],"adv":r["adv"],"tipo":r["tipo"],"placar":f"{pi} x {ai}","resultado":res})
        except: pass
    total = vit+emp+der
    return jsonify({"total":total,"vitorias":vit,"empates":emp,"derrotas":der,
                    "gols_pro":gp,"gols_contra":gc,"saldo":gp-gc,
                    "aproveitamento":round((vit/total)*100,1) if total else 0,"jogos":jogos})

@app.route("/api/excel/estatisticas")
@login_required
def excel_estatisticas():
    if not XLS_OK: return "Instale openpyxl", 500
    mod = request.args.get("mod",""); cat = request.args.get("cat","")
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT data,adv,tipo,placar_c7s,placar_adv FROM jogos WHERE mod=? AND cat=? AND placar_c7s!='' ORDER BY data ASC",(mod,cat))
    rows = cur.fetchall(); conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Estatisticas"
    azul, aln_c = xls_header(ws,f"ESTATÍSTICAS — {mod} {cat}",f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",5)
    borda = xls_add_headers(ws,["Data","Adversário","Mando","Placar","Resultado"],azul,aln_c)
    for i,r in enumerate(rows,1):
        try:
            pi,ai = int(r["placar_c7s"]),int(r["placar_adv"])
            res = "Vitoria" if pi>ai else ("Empate" if pi==ai else "Derrota")
            xls_add_row(ws,[r["data"],r["adv"],r["tipo"],f"{pi}x{ai}",res],borda,aln_c,i)
        except: pass
    for col,w in zip("ABCDE",[12,28,8,10,12]): ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"Estatisticas_{mod}_{cat}_{datetime.now().strftime('%Y%m%d')}.xlsx")

# ─────────────────────────────────────────────────────────────
# ESTOQUE
# ─────────────────────────────────────────────────────────────
@app.route("/api/estoque", methods=["GET"])
@login_required
def get_estoque():
    tipo = request.args.get("tipo","Escolinha"); termo = request.args.get("q","").lower()
    conn = get_db(); cur = conn.cursor()
    cur.execute("""SELECT e.id,e.nome,e.estoque_minimo,COALESCE(SUM(v.quantidade),0) as total
                   FROM estoque e LEFT JOIN estoque_variantes v ON v.estoque_id=e.id
                   WHERE e.tipo=? GROUP BY e.id ORDER BY e.nome ASC""",(tipo,))
    rows = []
    for r in cur.fetchall():
        if termo and termo not in r["nome"].lower(): continue
        total = r["total"]; minimo = r["estoque_minimo"]
        if total==0: status="zero"
        elif total/max(minimo,1)>1.0: status="ok"
        elif total/max(minimo,1)>0.5: status="baixo"
        else: status="critico"
        rows.append({"id":r["id"],"nome":r["nome"],"minimo":minimo,"total":total,"status":status})
    conn.close(); return jsonify(rows)

@app.route("/api/estoque", methods=["POST"])
@editor_required
def add_uniforme():
    d = request.json or {}
    nome = d.get("nome","").strip()
    if not nome: return jsonify({"error":"Nome obrigatorio"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT INTO estoque (tipo,nome,estoque_minimo) VALUES (?,?,?)",
                (d.get("tipo","Escolinha"),nome,int(d.get("minimo",5))))
    conn.commit(); new_id = cur.lastrowid; conn.close()
    return jsonify({"id":new_id}), 201

@app.route("/api/estoque/<int:uid>", methods=["PUT"])
@editor_required
def update_uniforme(uid):
    d = request.json or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE estoque SET nome=?,estoque_minimo=? WHERE id=?",
                (d.get("nome"),int(d.get("minimo",5)),uid))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/estoque/<int:uid>", methods=["DELETE"])
@editor_required
def delete_uniforme(uid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM estoque WHERE id=?",(uid,))
    cur.execute("DELETE FROM estoque_variantes WHERE estoque_id=?",(uid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/estoque/<int:uid>/variantes", methods=["GET"])
@login_required
def get_variantes(uid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT estoque_minimo FROM estoque WHERE id=?",(uid,))
    est = cur.fetchone(); minimo = est["estoque_minimo"] if est else 5
    cur.execute("SELECT id,tamanho,numero,quantidade FROM estoque_variantes WHERE estoque_id=? ORDER BY tamanho,numero",(uid,))
    rows = []
    for r in cur.fetchall():
        qty = r["quantidade"]
        if qty==0: st="ZERADO"
        elif qty/max(minimo,1)>1.0: st="OK"
        elif qty/max(minimo,1)>0.5: st="BAIXO"
        else: st="CRITICO"
        rows.append({"id":r["id"],"tamanho":r["tamanho"],"numero":r["numero"],"quantidade":qty,"status":st})
    conn.close(); return jsonify(rows)

@app.route("/api/estoque/<int:uid>/variantes", methods=["POST"])
@editor_required
def add_variante(uid):
    d = request.json or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT INTO estoque_variantes (estoque_id,tamanho,numero,quantidade) VALUES (?,?,?,?)",
                (uid,d.get("tamanho","").upper(),d.get("numero",""),int(d.get("quantidade",0))))
    conn.commit(); new_id = cur.lastrowid; conn.close()
    return jsonify({"id":new_id}), 201

@app.route("/api/estoque/variantes/<int:vid>", methods=["PUT"])
@editor_required
def update_variante(vid):
    d = request.json or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE estoque_variantes SET tamanho=?,numero=?,quantidade=? WHERE id=?",
                (d.get("tamanho","").upper(),d.get("numero",""),int(d.get("quantidade",0)),vid))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/estoque/variantes/<int:vid>/ajuste", methods=["PATCH"])
@editor_required
def ajuste_variante(vid):
    d = request.json or {}
    tipo = d.get("tipo","entrada"); qtd = int(d.get("quantidade",0))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT quantidade FROM estoque_variantes WHERE id=?",(vid,))
    row = cur.fetchone()
    if not row: conn.close(); return jsonify({"error":"Nao encontrado"}), 404
    nova = row["quantidade"]+qtd if tipo=="entrada" else row["quantidade"]-qtd
    if nova < 0: conn.close(); return jsonify({"error":f"Estoque insuficiente! Disponivel: {row['quantidade']}"}), 400
    cur.execute("UPDATE estoque_variantes SET quantidade=? WHERE id=?",(nova,vid))
    conn.commit(); conn.close(); return jsonify({"ok":True,"nova_quantidade":nova})

@app.route("/api/estoque/variantes/<int:vid>", methods=["DELETE"])
@editor_required
def delete_variante(vid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM estoque_variantes WHERE id=?",(vid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/excel/estoque")
@login_required
def excel_estoque():
    if not XLS_OK: return "Instale openpyxl", 500
    tipo = request.args.get("tipo","Escolinha")
    conn = get_db(); cur = conn.cursor()
    cur.execute("""SELECT e.nome,e.estoque_minimo,v.tamanho,v.numero,v.quantidade
                   FROM estoque e LEFT JOIN estoque_variantes v ON v.estoque_id=e.id
                   WHERE e.tipo=? ORDER BY e.nome,v.tamanho,v.numero""",(tipo,))
    rows = cur.fetchall(); conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = f"Estoque {tipo}"
    azul, aln_c = xls_header(ws,f"ESTOQUE — {tipo.upper()}",f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",5)
    borda = xls_add_headers(ws,["Item","Mínimo","Tamanho","Número","Quantidade"],azul,aln_c)
    for i,r in enumerate(rows,1):
        xls_add_row(ws,[r["nome"],r["estoque_minimo"],r["tamanho"] or "-",r["numero"] or "-",r["quantidade"] or 0],borda,aln_c,i)
    for col,w in zip("ABCDE",[35,10,12,10,12]): ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"Estoque_{tipo}_{datetime.now().strftime('%Y%m%d')}.xlsx")

# ─────────────────────────────────────────────────────────────
# ATIVIDADES
# ─────────────────────────────────────────────────────────────
@app.route("/api/atividades", methods=["GET"])
@login_required
def get_atividades():
    conn = get_db(); cur = conn.cursor()
    cur.execute("""SELECT id,tipo_atividade,nome_atividade,dia_semana,horario_inicio,
                          horario_fim,local,professor,estagiario,observacoes FROM atividades ORDER BY
                          CASE dia_semana WHEN 'Segunda-feira' THEN 1 WHEN 'Terca-feira' THEN 2
                          WHEN 'Quarta-feira' THEN 3 WHEN 'Quinta-feira' THEN 4
                          WHEN 'Sexta-feira' THEN 5 WHEN 'Sabado' THEN 6 ELSE 7 END, horario_inicio""")
    rows = [dict(r) for r in cur.fetchall()]; conn.close(); return jsonify(rows)

@app.route("/api/atividades", methods=["POST"])
@editor_required
def add_atividade():
    d = request.json or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute("""INSERT INTO atividades
                   (tipo_atividade,nome_atividade,dia_semana,horario_inicio,horario_fim,local,professor,estagiario,observacoes)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (d.get("tipo_atividade"),d.get("nome_atividade"),d.get("dia_semana"),
                 d.get("horario_inicio"),d.get("horario_fim"),d.get("local"),
                 d.get("professor"),d.get("estagiario",""),d.get("observacoes","")))
    conn.commit(); new_id = cur.lastrowid; conn.close(); return jsonify({"id":new_id}), 201

@app.route("/api/atividades/<int:aid>", methods=["PUT"])
@editor_required
def update_atividade(aid):
    d = request.json or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute("""UPDATE atividades SET tipo_atividade=?,nome_atividade=?,dia_semana=?,
                   horario_inicio=?,horario_fim=?,local=?,professor=?,estagiario=?,observacoes=? WHERE id=?""",
                (d.get("tipo_atividade"),d.get("nome_atividade"),d.get("dia_semana"),
                 d.get("horario_inicio"),d.get("horario_fim"),d.get("local"),
                 d.get("professor"),d.get("estagiario",""),d.get("observacoes",""),aid))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/atividades/<int:aid>", methods=["DELETE"])
@editor_required
def delete_atividade(aid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM atividades WHERE id=?",(aid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/excel/atividades")
@login_required
def excel_atividades():
    if not XLS_OK: return "Instale openpyxl", 500
    conn = get_db(); cur = conn.cursor()
    cur.execute("""SELECT tipo_atividade,nome_atividade,dia_semana,horario_inicio,horario_fim,
                          local,professor,estagiario,observacoes FROM atividades ORDER BY
                          CASE dia_semana WHEN 'Segunda-feira' THEN 1 WHEN 'Terca-feira' THEN 2
                          WHEN 'Quarta-feira' THEN 3 WHEN 'Quinta-feira' THEN 4
                          WHEN 'Sexta-feira' THEN 5 WHEN 'Sabado' THEN 6 ELSE 7 END, horario_inicio""")
    rows = cur.fetchall(); conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Grade Atividades"
    azul, aln_c = xls_header(ws,"GRADE DE ATIVIDADES",f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",9)
    borda = xls_add_headers(ws,["Tipo","Atividade","Dia","Início","Fim","Local","Professor","Estagiário","Obs"],azul,aln_c)
    for i,r in enumerate(rows,1):
        xls_add_row(ws,[r["tipo_atividade"],r["nome_atividade"],r["dia_semana"],
                        r["horario_inicio"],r["horario_fim"],r["local"] or "-",
                        r["professor"] or "-",r["estagiario"] or "-",r["observacoes"] or "-"],borda,aln_c,i)
    for col,w in zip("ABCDEFGHI",[14,22,14,8,8,20,22,18,20]): ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"Atividades_{datetime.now().strftime('%Y%m%d')}.xlsx")

# ─────────────────────────────────────────────────────────────
# ESCOLINHAS
# ─────────────────────────────────────────────────────────────
@app.route("/api/escolinhas", methods=["GET"])
@login_required
def get_escolinhas():
    conn = get_db(); cur = conn.cursor()
    cur.execute("""SELECT e.id,e.nome,COUNT(a.id) as total_alunos
                   FROM escolinhas e LEFT JOIN escolinha_alunos a ON a.escolinha_id=e.id
                   GROUP BY e.id ORDER BY e.nome""")
    rows = [dict(r) for r in cur.fetchall()]; conn.close(); return jsonify(rows)

@app.route("/api/escolinhas", methods=["POST"])
@editor_required
def add_escolinha():
    nome = (request.json or {}).get("nome","").strip()
    if not nome: return jsonify({"error":"Nome obrigatorio"}), 400
    conn = get_db(); cur = conn.cursor()
    try:
        cur.execute("INSERT INTO escolinhas (nome) VALUES (?)",(nome,))
        conn.commit(); new_id = cur.lastrowid
    except: conn.close(); return jsonify({"error":"Ja existe"}), 409
    conn.close(); return jsonify({"id":new_id}), 201

@app.route("/api/escolinhas/<int:eid>", methods=["PUT"])
@editor_required
def update_escolinha(eid):
    nome = (request.json or {}).get("nome","").strip()
    if not nome: return jsonify({"error":"Nome obrigatorio"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE escolinhas SET nome=? WHERE id=?",(nome,eid))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/escolinhas/<int:eid>", methods=["DELETE"])
@editor_required
def delete_escolinha(eid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM escolinha_alunos WHERE escolinha_id=?",(eid,))
    cur.execute("DELETE FROM escolinha_staff  WHERE escolinha_id=?",(eid,))
    cur.execute("DELETE FROM escolinhas WHERE id=?",(eid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/escolinhas/<int:eid>/alunos", methods=["GET"])
@login_required
def get_alunos_escolinha(eid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("""SELECT id,nome,turma,responsavel,telefone,data_nasc,observacoes
                   FROM escolinha_alunos WHERE escolinha_id=? ORDER BY nome ASC""",(eid,))
    rows = [dict(r) for r in cur.fetchall()]; conn.close(); return jsonify(rows)

@app.route("/api/escolinhas/<int:eid>/alunos", methods=["POST"])
@editor_required
def add_aluno_escolinha(eid):
    d = request.json or {}
    nome = d.get("nome","").strip()
    if not nome: return jsonify({"error":"Nome obrigatorio"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("""INSERT INTO escolinha_alunos (escolinha_id,nome,turma,responsavel,telefone,data_nasc,observacoes)
                   VALUES (?,?,?,?,?,?,?)""",
                (eid,nome,d.get("turma",""),d.get("responsavel",""),
                 d.get("telefone",""),d.get("data_nasc",""),d.get("observacoes","")))
    conn.commit(); new_id = cur.lastrowid; conn.close()
    return jsonify({"id":new_id}), 201

@app.route("/api/escolinhas/alunos/<int:aid>", methods=["PUT"])
@editor_required
def update_aluno_escolinha(aid):
    d = request.json or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute("""UPDATE escolinha_alunos SET nome=?,turma=?,responsavel=?,telefone=?,data_nasc=?,observacoes=?
                   WHERE id=?""",
                (d.get("nome",""),d.get("turma",""),d.get("responsavel",""),
                 d.get("telefone",""),d.get("data_nasc",""),d.get("observacoes",""),aid))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/escolinhas/alunos/<int:aid>", methods=["DELETE"])
@editor_required
def delete_aluno_escolinha(aid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM escolinha_alunos WHERE id=?",(aid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/escolinhas/<int:eid>/importar", methods=["POST"])
@editor_required
def importar_alunos(eid):
    """Importa alunos via planilha xlsx/xls.
    Colunas esperadas (qualquer ordem): nome, turma, responsavel, telefone, data_nasc, observacoes
    Linha 1 = cabeçalho."""
    if not XLS_OK: return jsonify({"error":"openpyxl nao instalado"}), 500
    if "file" not in request.files: return jsonify({"error":"Nenhum arquivo"}), 400
    f = request.files["file"]
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1,max_row=1))]
        col_map = {}
        aliases = {
            "nome":["nome","name","aluno","estudante"],
            "turma":["turma","grupo","classe","class","nivel"],
            "responsavel":["responsavel","responsável","pai/mae","pai","mae","guardian"],
            "telefone":["telefone","fone","cel","celular","phone","whatsapp"],
            "data_nasc":["data_nasc","data nascimento","nascimento","datanasc","nasc","birthday","dt_nasc"],
            "observacoes":["observacoes","observações","obs","notes","nota"],
        }
        for field, alts in aliases.items():
            for alt in alts:
                if alt in headers:
                    col_map[field] = headers.index(alt); break
        if "nome" not in col_map:
            return jsonify({"error":"Coluna 'nome' nao encontrada na planilha"}), 400
        conn = get_db(); cur = conn.cursor()
        inseridos = 0; ignorados = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            def cel(field): return str(row[col_map[field]] or "").strip() if field in col_map else ""
            nome = cel("nome")
            if not nome: ignorados += 1; continue
            cur.execute("""INSERT INTO escolinha_alunos (escolinha_id,nome,turma,responsavel,telefone,data_nasc,observacoes)
                           VALUES (?,?,?,?,?,?,?)""",
                        (eid,nome,cel("turma"),cel("responsavel"),cel("telefone"),cel("data_nasc"),cel("observacoes")))
            inseridos += 1
        conn.commit(); conn.close()
        return jsonify({"ok":True,"inseridos":inseridos,"ignorados":ignorados})
    except Exception as e:
        return jsonify({"error":f"Erro ao processar planilha: {str(e)}"}), 400

@app.route("/api/excel/escolinha/<int:eid>")
@login_required
def excel_escolinha(eid):
    if not XLS_OK: return "Instale openpyxl", 500
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT nome FROM escolinhas WHERE id=?",(eid,))
    esc = cur.fetchone()
    if not esc: return "Escolinha nao encontrada", 404
    cur.execute("""SELECT nome,turma,responsavel,telefone,data_nasc,observacoes
                   FROM escolinha_alunos WHERE escolinha_id=? ORDER BY nome ASC""",(eid,))
    rows = cur.fetchall(); conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Alunos"
    azul, aln_c = xls_header(ws,f"ESCOLINHA — {esc['nome'].upper()}",f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total: {len(rows)} alunos",6)
    borda = xls_add_headers(ws,["Nome","Turma","Responsável","Telefone","Dt. Nasc.","Observações"],azul,aln_c)
    for i,r in enumerate(rows,1):
        xls_add_row(ws,[r["nome"],r["turma"] or "-",r["responsavel"] or "-",
                        r["telefone"] or "-",r["data_nasc"] or "-",r["observacoes"] or "-"],borda,aln_c,i)
    for col,w in zip("ABCDEF",[38,14,28,16,12,25]): ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"Escolinha_{esc['nome'].replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx")

@app.route("/api/excel/escolinha/<int:eid>/modelo")
@login_required
def modelo_importacao(eid):
    """Gera planilha modelo para importação de alunos."""
    if not XLS_OK: return "Instale openpyxl", 500
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Alunos"
    azul = PatternFill("solid", fgColor="221C89")
    aln_c = Alignment(horizontal="center", vertical="center")
    borda = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    headers = ["nome","turma","responsavel","telefone","data_nasc","observacoes"]
    ws.append(headers)
    for i,h in enumerate(headers,1):
        cel = ws.cell(1,i)
        cel.fill = azul; cel.font = Font(name="Arial",bold=True,size=10,color="FFFFFF")
        cel.alignment = aln_c; cel.border = borda
    ws.append(["João da Silva","Sub-10","Maria da Silva","(85) 99999-9999","15/03/2015",""])
    for col,w in zip("ABCDEF",[38,12,28,18,12,25]): ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="Modelo_Importacao_Alunos.xlsx")

# ─────────────────────────────────────────────────────────────
# ESCOLINHA — STAFF (professores e estagiários por turma)
# ─────────────────────────────────────────────────────────────
CARGOS_VALIDOS = ("Professor", "Estagiario", "Coordenador", "Auxiliar")

@app.route("/api/escolinhas/<int:eid>/staff", methods=["GET"])
@login_required
def get_staff_escolinha(eid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("""SELECT id, nome, cargo, turma, horario, observacoes
                   FROM escolinha_staff
                   WHERE escolinha_id = ?
                   ORDER BY cargo, nome ASC""", (eid,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close(); return jsonify(rows)

@app.route("/api/escolinhas/<int:eid>/staff", methods=["POST"])
@editor_required
def add_staff_escolinha(eid):
    d = request.json or {}
    nome  = d.get("nome", "").strip()
    cargo = d.get("cargo", "Professor").strip()
    if not nome:
        return jsonify({"error": "Nome obrigatorio"}), 400
    if cargo not in CARGOS_VALIDOS:
        return jsonify({"error": f"Cargo invalido. Use: {', '.join(CARGOS_VALIDOS)}"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("""INSERT INTO escolinha_staff
                   (escolinha_id, nome, cargo, turma, horario, observacoes)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (eid, nome, cargo,
                 d.get("turma", ""),
                 d.get("horario", ""),
                 d.get("observacoes", "")))
    conn.commit(); new_id = cur.lastrowid; conn.close()
    return jsonify({"id": new_id}), 201

@app.route("/api/escolinhas/staff/<int:sid>", methods=["PUT"])
@editor_required
def update_staff_escolinha(sid):
    d = request.json or {}
    cargo = d.get("cargo", "Professor").strip()
    if cargo not in CARGOS_VALIDOS:
        return jsonify({"error": f"Cargo invalido. Use: {', '.join(CARGOS_VALIDOS)}"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("""UPDATE escolinha_staff
                   SET nome=?, cargo=?, turma=?, horario=?, observacoes=?
                   WHERE id=?""",
                (d.get("nome", "").strip(), cargo,
                 d.get("turma", ""),
                 d.get("horario", ""),
                 d.get("observacoes", ""),
                 sid))
    conn.commit(); conn.close(); return jsonify({"ok": True})

@app.route("/api/escolinhas/staff/<int:sid>", methods=["DELETE"])
@editor_required
def delete_staff_escolinha(sid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM escolinha_staff WHERE id=?", (sid,))
    conn.commit(); conn.close(); return jsonify({"ok": True})

@app.route("/api/excel/escolinha/<int:eid>/staff")
@login_required
def excel_staff_escolinha(eid):
    if not XLS_OK: return "Instale openpyxl", 500
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT nome FROM escolinhas WHERE id=?", (eid,))
    esc = cur.fetchone()
    if not esc: return "Escolinha nao encontrada", 404
    cur.execute("""SELECT nome, cargo, turma, horario, observacoes
                   FROM escolinha_staff
                   WHERE escolinha_id=?
                   ORDER BY cargo, nome ASC""", (eid,))
    rows = cur.fetchall(); conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Staff"
    azul, aln_c = xls_header(
        ws,
        f"STAFF — {esc['nome'].upper()}",
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total: {len(rows)} pessoas",
        5
    )
    borda = xls_add_headers(ws, ["Nome","Cargo","Turma/Grupo","Horário","Observações"], azul, aln_c)
    for i, r in enumerate(rows, 1):
        xls_add_row(ws,
                    [r["nome"], r["cargo"],
                     r["turma"] or "Todas",
                     r["horario"] or "-",
                     r["observacoes"] or "-"],
                    borda, aln_c, i)
    for col, w in zip("ABCDE", [38,14,16,14,28]):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Staff_{esc['nome'].replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )

# ─────────────────────────────────────────────────────────────
# COMUNICACAO
# ─────────────────────────────────────────────────────────────
@app.route("/api/comunicacao/direcao")
@login_required
def comunicacao_direcao():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT valor FROM config WHERE chave='assinatura_direcao'")
    row = cur.fetchone(); assinatura = row["valor"] if row else ""; conn.close()
    jogos = get_jogos_futuros(); total = len(jogos)
    tem_casa = any(j["tipo"]=="CASA" for j in jogos); tem_fora = any(j["tipo"]=="FORA" for j in jogos)
    if tem_casa and tem_fora: desc = "dentro e fora dos nossos dominios"
    elif tem_casa: desc = "dentro dos nossos dominios"
    elif tem_fora: desc = "fora dos nossos dominios"
    else: desc = "previstos"
    txt = f"Prezados,\n\nInformamos que teremos {total} jogo{'s' if total!=1 else ''} essa semana, {desc}.\n\n"
    if not jogos: txt += "Nao ha jogos futuros cadastrados no momento.\n\n"
    else:
        for secao, tipo_f in [("JOGOS EM CASA","CASA"),("JOGOS FORA DE CASA","FORA")]:
            bloco = [j for j in jogos if j["tipo"]==tipo_f]
            if not bloco: continue
            txt += f"━━━  {secao}  ━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            for j in bloco:
                hs = calc_hora_saida(j["hora"], j["saida_ant"])
                ret = j["retorno"] if j["retorno"] and j["retorno"] not in ("A definir","") else "A confirmar"
                txt += f"  Data          : {j['data']}\n  Horario       : {j['hora']}h\n"
                txt += f"  Modalidade    : {j['mod']} {j['cat']}\n  Competicao    : {j['comp']}\n"
                txt += f"  Confronto     : C7S x {j['adv']}\n  Local         : {j['local']}\n"
                if tipo_f=="FORA": txt += f"  Saida (C7S)   : {hs}h\n"
                txt += f"  Retorno       : {ret}\n\n"
    txt += assinatura; return jsonify({"texto":txt})

@app.route("/api/comunicacao/transporte")
@login_required
def comunicacao_transporte():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT valor FROM config WHERE chave='assinatura_trans'")
    row = cur.fetchone(); assinatura = row["valor"] if row else ""; conn.close()
    jogos = [j for j in get_jogos_futuros() if j["tipo"]=="FORA"]
    txt = "Prezados,\n\nGostaramos de solicitar os transportes da semana:\n\n"
    if not jogos: txt += "Nao ha jogos fora de casa previstos.\n\n"
    else:
        for i, j in enumerate(jogos, 1):
            hs = calc_hora_saida(j["hora"], j["saida_ant"])
            ret = j["retorno"] if j["retorno"] and j["retorno"] not in ("A definir","") else "A confirmar"
            txt += f"{'─'*44}\nVIAGEM {i}\n\n"
            txt += f"  Jogo              : C7S x {j['adv']}\n  Modalidade        : {j['mod']} {j['cat']}\n"
            txt += f"  Competicao        : {j['comp']}\n  Data              : {j['data']}\n"
            txt += f"  Saida (C7S)       : {hs}h\n  Destino           : {j['local']}\n"
            txt += f"  Horario do Jogo   : {j['hora']}h\n  Retorno           : {ret}\n\n"
    txt += f"{'─'*44}\n\n{assinatura}"; return jsonify({"texto":txt})

# ─────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────
@app.route("/api/config", methods=["GET"])
@login_required
def get_config():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT chave,valor FROM config")
    cfg = {r["chave"]:r["valor"] for r in cur.fetchall()}
    cfg.pop("logo_base64", None); conn.close(); return jsonify(cfg)

@app.route("/api/config", methods=["PUT"])
@editor_required
def update_config():
    d = request.json or {}
    conn = get_db(); cur = conn.cursor()
    for chave, valor in d.items():
        if chave == "logo_base64": continue
        cur.execute("INSERT OR REPLACE INTO config (chave,valor) VALUES (?,?)",(chave,valor))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# ─────────────────────────────────────────────────────────────
# LISTAS
# ─────────────────────────────────────────────────────────────
TABELAS_VALIDAS = {"modalidades","categorias","naipes",
                   "atv_nomes","atv_profs","atv_estags","atv_quadras","atv_horarios"}

@app.route("/api/listas/<tabela>", methods=["GET"])
@login_required
def get_lista(tabela):
    if tabela not in TABELAS_VALIDAS: return jsonify({"error":"Tabela invalida"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute(f"SELECT id,nome FROM {tabela} ORDER BY nome")
    rows = [dict(r) for r in cur.fetchall()]; conn.close(); return jsonify(rows)

@app.route("/api/listas/<tabela>", methods=["POST"])
@editor_required
def add_lista(tabela):
    if tabela not in TABELAS_VALIDAS: return jsonify({"error":"Tabela invalida"}), 400
    nome = (request.json or {}).get("nome","").strip()
    if not nome: return jsonify({"error":"Nome obrigatorio"}), 400
    conn = get_db(); cur = conn.cursor()
    try:
        cur.execute(f"INSERT INTO {tabela} (nome) VALUES (?)",(nome,))
        conn.commit(); new_id = cur.lastrowid
    except: conn.close(); return jsonify({"error":"Ja existe"}), 409
    conn.close(); return jsonify({"id":new_id,"nome":nome}), 201

@app.route("/api/listas/<tabela>/<int:lid>", methods=["DELETE"])
@editor_required
def del_lista(tabela, lid):
    if tabela not in TABELAS_VALIDAS: return jsonify({"error":"Tabela invalida"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute(f"DELETE FROM {tabela} WHERE id=?",(lid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/listas/<tabela>/<int:lid>", methods=["PUT"])
@editor_required
def rename_lista(tabela, lid):
    if tabela not in TABELAS_VALIDAS: return jsonify({"error":"Tabela invalida"}), 400
    nome = (request.json or {}).get("nome","").strip()
    if not nome: return jsonify({"error":"Nome obrigatorio"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute(f"UPDATE {tabela} SET nome=? WHERE id=?",(nome,lid))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# ─────────────────────────────────────────────────────────────
# PDF — LISTA VISITANTE / DISPENSA EF
# ─────────────────────────────────────────────────────────────
@app.route("/api/pdf/lista")
@login_required
def pdf_lista():
    if not PDF_OK: return "ERRO: Instale fpdf2", 500
    mod=request.args.get("mod",""); cat=request.args.get("cat",""); naipe=request.args.get("naipe",""); tipo=request.args.get("tipo","visitante")
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT nome FROM atletas WHERE mod=? AND cat=? AND naipe=? ORDER BY nome ASC",(mod,cat,naipe))
    atletas = cur.fetchall()
    cur.execute("SELECT nome,cargo FROM tecnicos WHERE mod=? AND cat=? AND naipe=? ORDER BY nome ASC",(mod,cat,naipe))
    tecs = cur.fetchall(); conn.close()
    logo = get_logo_tempfile()
    pdf = make_pdf(logo); pdf.add_page(); pdf.add_watermark()
    t_doc = "LISTA OFICIAL VISITANTE" if tipo=="visitante" else "DISPENSA - EDUCACAO FISICA"
    pdf.set_font("Arial","B",14); pdf.set_text_color(34,28,137)
    pdf.cell(190,10,safe(t_doc),ln=True,align="C")
    pdf.set_font("Arial","B",11); pdf.set_text_color(74,74,106)
    pdf.cell(190,8,safe(f"Selecao: {mod.upper()}  .  Categoria: {cat}  .  Naipe: {naipe}"),ln=True,align="C"); pdf.ln(8)
    if tecs:
        pdf.set_font("Arial","B",10); pdf.set_fill_color(242,179,26); pdf.set_text_color(34,28,137)
        pdf.cell(190,8,safe("COMISSAO TECNICA"),1,1,"C",True)
        pdf.set_font("Arial","",10); pdf.set_text_color(0,0,0)
        for row in tecs: pdf.cell(130,9,safe(f"  {row[0]}"),1,0,"L"); pdf.cell(60,9,safe(f"  {row[1]}"),1,1,"L")
        pdf.ln(4)
    pdf.set_font("Arial","B",10); pdf.set_fill_color(34,28,137); pdf.set_text_color(255,255,255)
    pdf.cell(20,10,safe("No"),1,0,"C",True); pdf.cell(170,10,safe("NOME COMPLETO DO ATLETA"),1,1,"L",True)
    pdf.set_font("Arial","",10); pdf.set_text_color(0,0,0); z = False
    for i,(nome,) in enumerate(atletas,1):
        pdf.set_fill_color(244,247,246) if z else pdf.set_fill_color(255,255,255)
        pdf.cell(20,10,str(i),1,0,"C",True); pdf.cell(170,10,safe(f"  {nome}"),1,1,"L",True); z = not z
    fn = f"{'Visitante' if tipo=='visitante' else 'DispensaEF'}_{mod}_{cat}.pdf"
    cleanup(logo); return pdf_to_response(pdf, fn)

# ─────────────────────────────────────────────────────────────
# PDF — AGENDA
# ─────────────────────────────────────────────────────────────
@app.route("/api/pdf/agenda")
@login_required
def pdf_agenda():
    if not PDF_OK: return "ERRO: Instale fpdf2", 500
    mod=request.args.get("mod",""); cat=request.args.get("cat",""); todos=request.args.get("todos","0")=="1"
    conn = get_db(); cur = conn.cursor()
    if todos: cur.execute("SELECT data,hora,adv,comp,tipo,local,saida,retorno,placar_c7s,placar_adv,mod,cat FROM jogos")
    else: cur.execute("SELECT data,hora,adv,comp,tipo,local,saida,retorno,placar_c7s,placar_adv,mod,cat FROM jogos WHERE mod=? AND cat=?",(mod,cat))
    jogos = sorted(cur.fetchall(), key=lambda r: parse_data(str(r[0]))); conn.close()
    logo = get_logo_tempfile(); pdf = make_pdf(logo); pdf.add_page(); pdf.add_watermark()
    pdf.set_font("Arial","B",14); pdf.set_text_color(34,28,137)
    pdf.cell(190,10,safe("AGENDA DE JOGOS E COMPETICOES"),ln=True,align="C")
    pdf.set_font("Arial","B",11); pdf.set_text_color(74,74,106)
    sub = "TODAS AS MODALIDADES" if todos else f"Selecao: {mod.upper()}  .  Categoria: {cat}"
    pdf.cell(190,8,safe(sub),ln=True,align="C"); pdf.ln(6)
    pdf.set_font("Arial","B",9); pdf.set_fill_color(34,28,137); pdf.set_text_color(255,255,255)
    for h,w in [("Data",24),("Hora",16),("Mod",18),("Cat",16),("Adversario",36),("Competicao",36),("Mando",16),("Local",26),("Retorno",14),("Placar",18)]:
        pdf.cell(w,10,safe(h),1,0,"C",True)
    pdf.ln(); pdf.set_font("Arial","",8); pdf.set_text_color(0,0,0); z = False
    for r in jogos:
        data,hora,adv,comp,tipo,local,saida,retorno,pc,pa,mod2,cat2 = r
        placar = f"{pc}x{pa}" if pc else "-"
        pdf.set_fill_color(244,247,246) if z else pdf.set_fill_color(255,255,255)
        for val,w in [(data,24),(hora,16),(mod2,18),(cat2,16),(adv,36),(comp,36),(tipo,16),(local,26),(retorno or "-",14),(placar,18)]:
            pdf.cell(w,9,safe(str(val))[:16],1,0,"C",True)
        pdf.ln(); z = not z
    fn = f"Agenda_{'TODAS' if todos else f'{mod}_{cat}'}.pdf"
    cleanup(logo); return pdf_to_response(pdf, fn)

# ─────────────────────────────────────────────────────────────
# PDF — CONVOCACAO
# ─────────────────────────────────────────────────────────────
@app.route("/api/pdf/convocacao/<int:jogo_id>")
@login_required
def pdf_convocacao(jogo_id):
    if not PDF_OK: return "ERRO: Instale fpdf2", 500
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT data,hora,adv,local,mod,cat,naipe FROM jogos WHERE id=?",(jogo_id,))
    jogo = cur.fetchone()
    if not jogo: conn.close(); return "Jogo nao encontrado", 404
    mod=jogo["mod"]; cat=jogo["cat"]; naipe=jogo["naipe"]
    cur.execute("SELECT a.nome,cv.status FROM atletas a JOIN convocacoes cv ON a.id=cv.atleta_id WHERE cv.jogo_id=? ORDER BY a.nome ASC",(jogo_id,))
    lista = cur.fetchall()
    cur.execute("SELECT nome,cargo FROM tecnicos WHERE mod=? AND cat=? AND naipe=? ORDER BY nome ASC",(mod,cat,naipe))
    tecs = cur.fetchall(); conn.close()
    logo = get_logo_tempfile(); pdf = make_pdf(logo); pdf.add_page(); pdf.add_watermark()
    pdf.set_font("Arial","B",14); pdf.set_text_color(34,28,137)
    pdf.cell(190,10,safe("LISTA DE CONVOCACAO"),ln=True,align="C")
    pdf.set_font("Arial","",11); pdf.set_text_color(74,74,106)
    pdf.cell(190,8,safe(f"{mod} {cat} {naipe}  .  C7S x {jogo['adv']}  .  {jogo['data']} as {jogo['hora']}"),ln=True,align="C")
    pdf.cell(190,7,safe(f"Local: {jogo['local']}"),ln=True,align="C"); pdf.ln(6)
    if tecs:
        pdf.set_font("Arial","B",10); pdf.set_fill_color(242,179,26); pdf.set_text_color(34,28,137)
        pdf.cell(190,8,safe("COMISSAO TECNICA"),1,1,"C",True)
        pdf.set_font("Arial","",10); pdf.set_text_color(0,0,0)
        for tnome,tcargo in tecs: pdf.cell(130,9,safe(f"  {tnome}"),1,0,"L"); pdf.cell(60,9,safe(f"  {tcargo}"),1,1,"L")
        pdf.ln(4)
    pdf.set_font("Arial","B",10); pdf.set_fill_color(34,28,137); pdf.set_text_color(255,255,255)
    pdf.cell(15,10,"#",1,0,"C",True); pdf.cell(120,10,safe("NOME"),1,0,"L",True); pdf.cell(55,10,safe("STATUS"),1,1,"C",True)
    pdf.set_font("Arial","",10); z = False
    for i,(nome,status) in enumerate(lista,1):
        pdf.set_fill_color(244,247,246) if z else pdf.set_fill_color(255,255,255)
        pdf.set_text_color(0,0,0)
        pdf.cell(15,10,str(i),1,0,"C",True); pdf.cell(120,10,safe(f"  {nome}"),1,0,"L",True)
        if status=="Presente": pdf.set_text_color(39,174,96)
        elif status=="Ausente": pdf.set_text_color(231,76,60)
        else: pdf.set_text_color(34,28,137)
        pdf.cell(55,10,safe(status),1,1,"C",True); z = not z
    fn = f"Convocacao_{mod}_{cat}_{jogo['data'].replace('/','_')}.pdf"
    cleanup(logo); return pdf_to_response(pdf, fn)

# ─────────────────────────────────────────────────────────────
# EXCEL — VIAGENS
# ─────────────────────────────────────────────────────────────
@app.route("/api/excel/viagens")
@login_required
def excel_viagens():
    if not XLS_OK: return "Instale openpyxl", 500
    jogos_fora = [j for j in get_jogos_futuros() if j["tipo"]=="FORA"]
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Viagens"
    azul, aln_c = xls_header(ws,"RELATÓRIO DE VIAGENS — JOGOS FORA DE CASA",f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",11)
    borda = xls_add_headers(ws,["#","Data","Hora Jogo","Modalidade","Categoria","Adversario","Competicao","Local (Destino)","Saida (C7S)","Retorno","Placar"],azul,aln_c)
    for idx,j in enumerate(jogos_fora,1):
        hs = calc_hora_saida(j["hora"],j["saida_ant"])
        ret = j["retorno"] if j["retorno"] and j["retorno"] not in ("A definir","") else "A confirmar"
        xls_add_row(ws,[idx,j["data"],j["hora"],j["mod"],j["cat"],j["adv"],j["comp"],j["local"],hs,ret,j.get("placar","")],borda,aln_c,idx)
    for col,w in zip("ABCDEFGHIJK",[5,12,11,12,10,22,22,28,12,12,10]): ws.column_dimensions[col].width = w
    ws.append([]); ws.append([f"Total de viagens: {len(jogos_fora)}"])
    ws.cell(ws.max_row,1).font = Font(name="Arial",bold=True,size=10,color="221C89")
    fn = f"Viagens_C7S_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fn)

# ─────────────────────────────────────────────────────────────
# FRONTEND
# ─────────────────────────────────────────────────────────────
@app.route("/")
@app.route("/<path:path>")
def index(path=""):
    return render_template("index.html")

# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 5000))
    print(f"SGE MasterPro Web v2  http://0.0.0.0:{port}")
    try:
        from waitress import serve
        serve(app, host="0.0.0.0", port=port)
    except ImportError:
        app.run(host="0.0.0.0", port=port, debug=not IS_RENDER)