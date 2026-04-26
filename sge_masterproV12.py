import tkinter as tk
from tkinter import messagebox
import sys

try:
    from tkinter import ttk
    import sqlite3
    import os
    from fpdf import FPDF
    from PIL import Image, ImageTk
    from datetime import datetime, timedelta
    import traceback
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    root = tk.Tk(); root.withdraw()
    messagebox.showerror("Falta de Biblioteca", f"Erro: {e}")
    sys.exit()

pasta_atual = os.path.dirname(os.path.abspath(__file__))
db_path     = os.path.join(pasta_atual, "SGE_MasterPro_V12.db")
PASTA_DIST  = os.path.join(pasta_atual, "dist")
os.makedirs(PASTA_DIST, exist_ok=True)

NOME_ESCOLA  = "COLÉGIO 7 DE SETEMBRO - NGS"
NOME_SISTEMA = "SISTEMA DE GESTÃO ESPORTIVA - C7S"
NOME_LOGO    = os.path.join(pasta_atual, "c7s-rebranding-2026-logo-variacoes-01.png")
NOME_BG      = os.path.join(pasta_atual, "c7s-rebranding-2026-background-03.jpg")

C_PRIMARIO   = "#221C89"
C_ACENTO     = "#F2B31A"
C_SECUNDARIO = "#4A4A6A"
C_FUNDO      = "#f4f7f6"
C_SUCESSO    = "#27AE60"
C_PERIGO     = "#E74C3C"

# ─────────────────────────────────────────────────────────────
# BANCO DE DADOS
# ─────────────────────────────────────────────────────────────
def init_db():
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    # Tabelas principais
    c.execute("CREATE TABLE IF NOT EXISTS usuarios (user TEXT PRIMARY KEY, senha TEXT)")
    c.execute("""CREATE TABLE IF NOT EXISTS atletas
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT, mod TEXT, cat TEXT, naipe TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS tecnicos
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  nome TEXT, mod TEXT, cat TEXT, naipe TEXT, cargo TEXT DEFAULT 'Técnico')""")
    c.execute("""CREATE TABLE IF NOT EXISTS jogos
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  data TEXT, hora TEXT, mod TEXT, cat TEXT, adv TEXT, comp TEXT,
                  tipo TEXT, local TEXT, saida TEXT,
                  retorno TEXT DEFAULT '', placar_c7s TEXT DEFAULT '', placar_adv TEXT DEFAULT '')""")
    c.execute("""CREATE TABLE IF NOT EXISTS convocacoes
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  jogo_id INTEGER, atleta_id INTEGER, status TEXT DEFAULT 'Convocado',
                  UNIQUE(jogo_id, atleta_id))""")

    # Listas de filtro global
    c.execute("CREATE TABLE IF NOT EXISTS modalidades (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT UNIQUE)")
    for mod in ["Futsal","Vôlei","Basquete","Handebol"]:
        c.execute("INSERT OR IGNORE INTO modalidades (nome) VALUES (?)",(mod,))

    c.execute("CREATE TABLE IF NOT EXISTS categorias (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT UNIQUE)")
    for cat in ["Sub-12","Sub-14","Sub-15","Sub-17","Livre"]:
        c.execute("INSERT OR IGNORE INTO categorias (nome) VALUES (?)",(cat,))

    c.execute("CREATE TABLE IF NOT EXISTS naipes (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT UNIQUE)")
    for nai in ["Masculino","Feminino"]:
        c.execute("INSERT OR IGNORE INTO naipes (nome) VALUES (?)",(nai,))

    # Migrações seguras de colunas de jogos
    for col in ["ALTER TABLE jogos ADD COLUMN retorno TEXT DEFAULT ''",
                "ALTER TABLE jogos ADD COLUMN placar_c7s TEXT DEFAULT ''",
                "ALTER TABLE jogos ADD COLUMN placar_adv TEXT DEFAULT ''"]:
        try: c.execute(col)
        except: pass

    # Estoque com migração segura
    c.execute("PRAGMA table_info(estoque)")
    cols_est = [r[1] for r in c.fetchall()]
    if "tamanho" in cols_est or "quantidade" in cols_est:
        c.execute("ALTER TABLE estoque RENAME TO estoque_legado")
        c.execute("""CREATE TABLE estoque
                     (id INTEGER PRIMARY KEY AUTOINCREMENT, tipo TEXT, nome TEXT, estoque_minimo INTEGER DEFAULT 5)""")
        c.execute("INSERT OR IGNORE INTO estoque (tipo,nome,estoque_minimo) SELECT DISTINCT tipo,nome,5 FROM estoque_legado")
        conn.commit()

    c.execute("PRAGMA table_info(estoque)")
    cols_est = [r[1] for r in c.fetchall()]
    if not cols_est:
        c.execute("""CREATE TABLE IF NOT EXISTS estoque
                     (id INTEGER PRIMARY KEY AUTOINCREMENT, tipo TEXT, nome TEXT, estoque_minimo INTEGER DEFAULT 5)""")
    if "estoque_minimo" not in cols_est:
        try: c.execute("ALTER TABLE estoque ADD COLUMN estoque_minimo INTEGER DEFAULT 5")
        except: pass

    c.execute("""CREATE TABLE IF NOT EXISTS estoque_variantes
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  estoque_id INTEGER, tamanho TEXT DEFAULT '', numero TEXT DEFAULT '', quantidade INTEGER DEFAULT 0)""")

    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='estoque_legado'")
    if c.fetchone():
        c.execute("""INSERT INTO estoque_variantes (estoque_id,tamanho,numero,quantidade)
                     SELECT e.id,l.tamanho,l.numero,l.quantidade FROM estoque_legado l
                     JOIN estoque e ON e.tipo=l.tipo AND e.nome=l.nome WHERE l.tamanho!='' OR l.quantidade>0""")
        c.execute("DROP TABLE estoque_legado")
        conn.commit()

    c.execute("CREATE TABLE IF NOT EXISTS config (chave TEXT PRIMARY KEY, valor TEXT)")
    c.execute("INSERT OR IGNORE INTO config VALUES ('assinatura_direcao','Atenciosamente,\nCoordenacao de Cursos Livres\nCOLEGIO 7 DE SETEMBRO - NGS')")
    c.execute("INSERT OR IGNORE INTO config VALUES ('assinatura_trans','Atenciosamente,\nCoordenacao de Cursos Livres\nCOLEGIO 7 DE SETEMBRO - NGS')")
    c.execute("INSERT OR IGNORE INTO usuarios VALUES ('admin','c7s2026')")
    c.execute("INSERT OR IGNORE INTO usuarios VALUES ('italo','esporte')")

    # Atividades esportivas
    c.execute("""CREATE TABLE IF NOT EXISTS atividades
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  tipo_atividade TEXT, nome_atividade TEXT, dia_semana TEXT,
                  horario_inicio TEXT, horario_fim TEXT, local TEXT,
                  professor TEXT, estagiario TEXT DEFAULT '', observacoes TEXT DEFAULT '')""")

    # ── LISTAS CADASTRÁVEIS para a aba Atividades ─────────────
    c.execute("CREATE TABLE IF NOT EXISTS atv_nomes    (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT UNIQUE)")
    c.execute("CREATE TABLE IF NOT EXISTS atv_profs    (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT UNIQUE)")
    c.execute("CREATE TABLE IF NOT EXISTS atv_estags   (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT UNIQUE)")
    c.execute("CREATE TABLE IF NOT EXISTS atv_quadras  (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT UNIQUE)")
    c.execute("CREATE TABLE IF NOT EXISTS atv_horarios (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT UNIQUE)")

    # Valores padrão
    for n in ["Futsal Escolinha","Vôlei Escolinha","Basquete Escolinha",
              "Futsal Seleção","Vôlei Seleção","Ed. Física Turma A"]:
        c.execute("INSERT OR IGNORE INTO atv_nomes (nome) VALUES (?)",(n,))
    for q in ["Quadra Coberta","Quadra Externa","Ginásio","Piscina","Campo de Futebol"]:
        c.execute("INSERT OR IGNORE INTO atv_quadras (nome) VALUES (?)",(q,))
    for h in ["07:00","07:30","08:00","08:30","09:00","10:00","11:00",
              "13:00","13:30","14:00","15:00","16:00","17:00","18:00","19:00"]:
        c.execute("INSERT OR IGNORE INTO atv_horarios (nome) VALUES (?)",(h,))

    conn.commit(); conn.close()

# ─────────────────────────────────────────────────────────────
# PDF
# ─────────────────────────────────────────────────────────────
class PDF_Elite(FPDF):
    def header(self):
        tx = 52 if os.path.exists(NOME_LOGO) else 10
        if os.path.exists(NOME_LOGO):
            self.image(NOME_LOGO, 10, 6, 38)
        self.set_xy(tx, 9)
        self.set_font("Arial","B",16); self.set_text_color(34,28,137)
        self.cell(200-tx, 9, NOME_ESCOLA, 0, 2, "L")
        self.set_x(tx); self.set_font("Arial","",9); self.set_text_color(100,100,100)
        self.cell(200-tx, 5, "Sistema de Gestao Esportiva - Coordenacao de Cursos Livres", 0, 1, "L")
        self.set_draw_color(242,179,26); self.set_line_width(0.8)
        self.line(10,38,200,38); self.ln(12)

    def add_watermark(self):
        if os.path.exists(NOME_LOGO):
            self.image(NOME_LOGO, x=45, y=100, w=120)

    def footer(self):
        self.set_y(-15); self.set_font("Arial","I",8); self.set_text_color(150,150,150)
        self.cell(0,10,f'Doc. Oficial C7S - {datetime.now().strftime("%d/%m/%Y %H:%M")} - Pag {self.page_no()}/{{nb}}',0,0,"C")

# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def make_btn(parent, text, command, bg=None, fg="white", font_size=10,
             height=2, padx=10, cursor="hand2", width=None, wraplength=None):
    bg = bg or C_PRIMARIO
    kw = dict(text=text, command=command, bg=bg, fg=fg,
              font=("Segoe UI", font_size, "bold"), bd=0,
              height=height, padx=padx, cursor=cursor,
              relief="flat", activebackground=bg)
    if width: kw["width"] = width
    if wraplength: kw["wraplength"] = wraplength
    return tk.Button(parent, **kw)

def tooltip(widget, text):
    tip = None
    def enter(e):
        nonlocal tip
        tip = tk.Toplevel(widget); tip.wm_overrideredirect(True)
        tip.wm_geometry(f"+{widget.winfo_rootx()+25}+{widget.winfo_rooty()+25}")
        tk.Label(tip, text=text, background="#ffffe0", relief="solid",
                 borderwidth=1, font=("Segoe UI",8)).pack()
    def leave(e):
        nonlocal tip
        if tip: tip.destroy(); tip = None
    widget.bind("<Enter>", enter); widget.bind("<Leave>", leave)

def parse_data(data_str):
    try: return datetime.strptime(data_str, "%d/%m/%Y").date()
    except: return datetime.max.date()

def db_lista(tabela):
    """Retorna lista de nomes de uma tabela de lista cadastrável."""
    conn = sqlite3.connect(db_path); c = conn.cursor()
    c.execute(f"SELECT nome FROM {tabela} ORDER BY nome")
    r = [x[0] for x in c.fetchall()]; conn.close(); return r

def janela_gerenciar_lista(parent, tabela, titulo):
    """Janela genérica para gerenciar qualquer lista cadastrável (nomes, profs, etc.)."""
    w = tk.Toplevel(parent); w.title(f"Gerenciar — {titulo}")
    w.geometry("420x460"); w.configure(bg=C_FUNDO)
    w.transient(parent); w.grab_set()

    tk.Label(w, text=f"📋  {titulo.upper()}", bg=C_FUNDO, fg=C_PRIMARIO,
             font=("Segoe UI",13,"bold")).pack(pady=(14,6))
    tk.Label(w, text="Adicione os itens que aparecerão nos menus de seleção.",
             bg=C_FUNDO, fg=C_SECUNDARIO, font=("Segoe UI",9)).pack(pady=(0,8))

    add_fr = tk.Frame(w, bg=C_FUNDO); add_fr.pack(fill="x", padx=16, pady=(0,6))
    e = tk.Entry(add_fr, font=("Segoe UI",11), bd=1, relief="solid")
    e.pack(side="left", fill="x", expand=True, padx=(0,8))

    tree = ttk.Treeview(w, columns=("nome",), show="headings", height=14)
    tree.heading("nome", text=titulo); tree.column("nome", width=360, anchor="w")
    sc = ttk.Scrollbar(w, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=sc.set)

    def carregar():
        for i in tree.get_children(): tree.delete(i)
        for n in db_lista(tabela): tree.insert("","end",values=(n,))

    def adicionar(ev=None):
        nome = e.get().strip()
        if not nome: return
        conn = sqlite3.connect(db_path); cur = conn.cursor()
        try:
            cur.execute(f"INSERT INTO {tabela} (nome) VALUES (?)",(nome,))
            conn.commit(); e.delete(0,"end"); carregar()
        except sqlite3.IntegrityError:
            messagebox.showwarning("Aviso","Já existe!",parent=w)
        finally: conn.close()

    def excluir():
        sel = tree.selection()
        if not sel: return
        nome = tree.item(sel[0])["values"][0]
        if messagebox.askyesno("Confirmar",f"Excluir '{nome}'?",parent=w):
            conn = sqlite3.connect(db_path); cur = conn.cursor()
            cur.execute(f"DELETE FROM {tabela} WHERE nome=?",(nome,))
            conn.commit(); conn.close(); carregar()

    def renomear():
        sel = tree.selection()
        if not sel: return
        nome_atual = tree.item(sel[0])["values"][0]
        d = tk.Toplevel(w); d.title("Renomear"); d.geometry("340x140")
        d.configure(bg="white"); d.transient(w); d.grab_set()
        tk.Label(d,text="Novo nome:",bg="white",font=("Segoe UI",10)).pack(pady=(18,4))
        en = tk.Entry(d,font=("Segoe UI",11),bd=1,relief="solid",width=28)
        en.insert(0,nome_atual); en.pack(pady=(0,12))
        def ok(ev=None):
            nn = en.get().strip()
            if not nn: return
            conn2 = sqlite3.connect(db_path); cur2 = conn2.cursor()
            cur2.execute(f"UPDATE {tabela} SET nome=? WHERE nome=?",(nn,nome_atual))
            conn2.commit(); conn2.close(); d.destroy(); carregar()
        en.bind("<Return>",ok)
        make_btn(d,"SALVAR",ok,bg=C_PRIMARIO,font_size=10,height=1).pack(fill="x",padx=40)
        en.focus(); en.select_range(0,"end")

    make_btn(add_fr,"+ ADICIONAR",adicionar,bg=C_PRIMARIO,font_size=9,height=1,padx=12).pack(side="left")
    e.bind("<Return>",adicionar)

    btn_fr = tk.Frame(w, bg=C_FUNDO); btn_fr.pack(fill="x", padx=16, pady=(0,4))
    make_btn(btn_fr,"✏️ RENOMEAR",renomear,bg=C_SECUNDARIO,font_size=9,height=1,padx=10).pack(side="left",padx=(0,6))
    make_btn(btn_fr,"🗑️ EXCLUIR", excluir, bg=C_PERIGO,   font_size=9,height=1,padx=10).pack(side="left")

    tree.pack(side="left", fill="both", expand=True, padx=(16,0), pady=(0,14))
    sc.pack(side="right", fill="y", pady=(0,14), padx=(0,16))
    carregar()
    return w

# ─────────────────────────────────────────────────────────────
# CLASSE PRINCIPAL
# ─────────────────────────────────────────────────────────────
class SGEMasterPro:
    def __init__(self, root):
        self.root = root
        self.root.title(f"{NOME_SISTEMA} - {NOME_ESCOLA}")
        self.root.geometry("1300x840"); self.root.minsize(1150,740)
        self.bg_image_ref = None; self.conv_jogos_map = {}
        self.mod_sel = tk.StringVar(value="Futsal")
        self.cat_sel = tk.StringVar(value="Sub-15")
        self.nai_sel = tk.StringVar(value="Masculino")
        self.ver_todos_jogos = tk.BooleanVar(value=False)
        # estados de ordenação por coluna em cada aba
        self._sort_state = {}
        self.setup_styles(); self.login_view()

    # ── Helpers de banco ──────────────────────────────────────
    def _get_modalidades(self): return db_lista("modalidades") or ["Futsal"]
    def _get_categorias(self):  return db_lista("categorias")  or ["Sub-15"]
    def _get_naipes(self):      return db_lista("naipes")      or ["Masculino"]

    def _refresh_filtros(self):
        mods = self._get_modalidades(); cats = self._get_categorias(); nais = self._get_naipes()
        if hasattr(self,'_cb_mod_hdr'):
            self._cb_mod_hdr["values"] = mods; self._cb_cat_hdr["values"] = cats; self._cb_nai_hdr["values"] = nais
            if self.mod_sel.get() not in mods: self.mod_sel.set(mods[0])
            if self.cat_sel.get() not in cats: self.cat_sel.set(cats[0])
            if self.nai_sel.get() not in nais: self.nai_sel.set(nais[0])

    def setup_styles(self):
        s = ttk.Style(); s.theme_use("clam")
        s.configure("TNotebook", background=C_FUNDO, borderwidth=0)
        s.configure("TNotebook.Tab", font=("Segoe UI",10,"bold"), padding=[20,7], background="#dde0e8")
        s.map("TNotebook.Tab", background=[("selected",C_PRIMARIO)], foreground=[("selected","white")])
        s.configure("Treeview", font=("Segoe UI",10), rowheight=30, background="white", fieldbackground="white")
        s.configure("Treeview.Heading", font=("Segoe UI",10,"bold"), background="#eef0f5", foreground=C_PRIMARIO)
        s.map("Treeview", background=[("selected","#dde4ff")], foreground=[("selected",C_PRIMARIO)])

    def limpar_tela(self):
        for w in self.root.winfo_children(): w.destroy()

    def aplicar_fundo(self):
        if os.path.exists(NOME_BG):
            img = Image.open(NOME_BG).resize((1300,840), getattr(Image,"LANCZOS",1))
            self.bg_image_ref = ImageTk.PhotoImage(img)
            lbl = tk.Label(self.root, image=self.bg_image_ref)
            lbl.place(x=0,y=0,relwidth=1,relheight=1); lbl.lower()
        else:
            self.root.configure(bg=C_FUNDO)

    # ── Ordenação genérica de Treeview ────────────────────────
    def _sort_tree(self, tree, col, tree_id):
        key = (tree_id, col)
        reverse = self._sort_state.get(key, False)
        data = [(tree.set(iid, col), iid) for iid in tree.get_children("")]
        data.sort(key=lambda x: x[0].lower() if isinstance(x[0],str) else x[0], reverse=reverse)
        for idx,(_, iid) in enumerate(data): tree.move(iid,"",idx)
        self._sort_state[key] = not reverse
        # Atualiza seta no cabeçalho
        for c2 in tree["columns"]:
            txt = tree.heading(c2,"text").replace(" ↑","").replace(" ↓","")
            tree.heading(c2, text=txt)
        arrow = " ↓" if reverse else " ↑"
        tree.heading(col, text=tree.heading(col,"text").replace(" ↑","").replace(" ↓","")+arrow)

    def _make_sortable(self, tree, tree_id):
        for col in tree["columns"]:
            tree.heading(col, text=tree.heading(col,"text"),
                         command=lambda c=col: self._sort_tree(tree, c, tree_id))

    # ═══════════════════════════════════════════════════════════
    # LOGIN
    # ═══════════════════════════════════════════════════════════
    def login_view(self):
        self.limpar_tela(); self.aplicar_fundo()
        f = tk.Frame(self.root, bg="white", highlightbackground=C_ACENTO, highlightthickness=2)
        f.place(relx=.5, rely=.5, anchor="center", width=440, height=580)
        hdr = tk.Frame(f, bg="white", height=150); hdr.pack(fill="x", pady=20)
        if os.path.exists(NOME_LOGO):
            img = Image.open(NOME_LOGO).resize((185,135), getattr(Image,"LANCZOS",1))
            self.logo_img = ImageTk.PhotoImage(img)
            tk.Label(hdr, image=self.logo_img, bg="white").pack(pady=10)
        else:
            tk.Label(hdr, text="[ LOGO C7S ]", fg=C_PRIMARIO, bg="white",
                     font=("Segoe UI",16,"bold")).pack(pady=30)
        body = tk.Frame(f, bg="white", padx=45, pady=10); body.pack(fill="both", expand=True)
        tk.Label(body,text="Usuário:",bg="white",fg=C_PRIMARIO,font=("Segoe UI",10,"bold")).pack(anchor="w")
        self.u_ent = tk.Entry(body, font=("Segoe UI",12), bd=1, relief="solid"); self.u_ent.pack(fill="x",pady=(4,14))
        tk.Label(body,text="Senha:",bg="white",fg=C_PRIMARIO,font=("Segoe UI",10,"bold")).pack(anchor="w")
        self.p_ent = tk.Entry(body, font=("Segoe UI",12), bd=1, relief="solid", show="*"); self.p_ent.pack(fill="x",pady=(4,24))
        self.u_ent.bind("<Return>", lambda e: self.p_ent.focus())
        self.p_ent.bind("<Return>", lambda e: self.autenticar())
        make_btn(body,"ACESSAR SISTEMA",self.autenticar,bg=C_PRIMARIO,font_size=12).pack(fill="x")
        make_btn(body,"CRIAR NOVO CADASTRO",self.tela_cadastro,bg="white",fg=C_SECUNDARIO,
                 font_size=10,height=1).pack(pady=(12,0))
        self.u_ent.focus()

    def tela_cadastro(self):
        w = tk.Toplevel(self.root); w.title("Novo Cadastro"); w.geometry("360x300")
        w.configure(bg="white"); w.transient(self.root); w.grab_set()
        tk.Label(w,text="CRIAR CONTA",bg="white",fg=C_PRIMARIO,font=("Segoe UI",14,"bold")).pack(pady=(20,10))
        tk.Label(w,text="Novo Usuário:",bg="white",fg=C_SECUNDARIO,font=("Segoe UI",10,"bold")).pack(anchor="w",padx=40)
        u = tk.Entry(w,font=("Segoe UI",12),bd=1,relief="solid"); u.pack(fill="x",padx=40,pady=(0,15))
        tk.Label(w,text="Nova Senha:",bg="white",fg=C_SECUNDARIO,font=("Segoe UI",10,"bold")).pack(anchor="w",padx=40)
        p = tk.Entry(w,font=("Segoe UI",12),bd=1,relief="solid",show="*"); p.pack(fill="x",padx=40,pady=(0,20))
        def salvar(e=None):
            un,se = u.get().strip(),p.get().strip()
            if not un or not se: messagebox.showwarning("Aviso","Preencha tudo!",parent=w); return
            conn = sqlite3.connect(db_path); c2 = conn.cursor()
            try:
                c2.execute("INSERT INTO usuarios VALUES (?,?)",(un,se)); conn.commit()
                messagebox.showinfo("Sucesso","Cadastrado!",parent=w); w.destroy()
            except sqlite3.IntegrityError: messagebox.showerror("Erro","Usuário já existe.",parent=w)
            finally: conn.close()
        u.bind("<Return>",lambda e: p.focus()); p.bind("<Return>",salvar)
        make_btn(w,"SALVAR",salvar,bg=C_ACENTO,fg=C_PRIMARIO,font_size=11).pack(fill="x",padx=40)
        u.focus()

    def autenticar(self):
        conn = sqlite3.connect(db_path); c = conn.cursor()
        c.execute("SELECT * FROM usuarios WHERE user=? AND senha=?",(self.u_ent.get(),self.p_ent.get()))
        ok = c.fetchone(); conn.close()
        if ok: self.main_dashboard()
        else: messagebox.showerror("Acesso Negado","Usuário ou senha inválidos.")

    # ═══════════════════════════════════════════════════════════
    # DASHBOARD
    # ═══════════════════════════════════════════════════════════
    def main_dashboard(self):
        self.limpar_tela(); self.aplicar_fundo()
        hdr = tk.Frame(self.root, bg="white", highlightbackground=C_ACENTO, highlightthickness=2)
        hdr.pack(fill="x", padx=10, pady=(10,0))
        tf = tk.Frame(hdr,bg="white"); tf.pack(side="left",padx=20,pady=10)
        if os.path.exists(NOME_LOGO):
            img = Image.open(NOME_LOGO).resize((90,65), getattr(Image,"LANCZOS",1))
            self.logo_dash = ImageTk.PhotoImage(img)
            tk.Label(tf,image=self.logo_dash,bg="white").grid(row=0,column=0,rowspan=2,padx=(0,15))
        tk.Label(tf,text=NOME_ESCOLA,fg=C_PRIMARIO,bg="white",font=("Segoe UI",14,"bold")).grid(row=0,column=1,sticky="w")
        tk.Label(tf,text=f"Painel de Controle Esportivo  ·  {NOME_SISTEMA}",
                 fg=C_SECUNDARIO,bg="white",font=("Segoe UI",10)).grid(row=1,column=1,sticky="w")

        ff = tk.LabelFrame(hdr,text=" Seleção Ativa / Filtro GERAL ",bg="white",
                           font=("Segoe UI",8,"bold"),fg=C_PRIMARIO)
        ff.pack(side="right",padx=20,pady=10)

        def cf(p,t,v,col):
            tk.Label(p,text=t,bg="white",font=("Segoe UI",8,"bold")).grid(row=0,column=col,padx=6,sticky="w")
            cb = ttk.Combobox(p,textvariable=v,width=16,state="readonly")
            cb.grid(row=1,column=col,padx=6,pady=(0,6)); return cb

        self._cb_mod_hdr = cf(ff,"MODALIDADE:",self.mod_sel,0)
        self._cb_cat_hdr = cf(ff,"CATEGORIA:", self.cat_sel,1)
        self._cb_nai_hdr = cf(ff,"NAIPE:",     self.nai_sel,2)
        self._cb_mod_hdr["values"] = self._get_modalidades()
        self._cb_cat_hdr["values"] = self._get_categorias()
        self._cb_nai_hdr["values"] = self._get_naipes()

        make_btn(ff,"🔄 ATUALIZAR",self.refresh_all,bg=C_ACENTO,fg=C_PRIMARIO,
                 font_size=9,height=1,padx=18).grid(row=0,column=3,rowspan=2,padx=4,pady=6,sticky="ns")
        make_btn(ff,"⚙️ LISTAS",self.tela_gerenciar_listas,bg=C_SECUNDARIO,fg="white",
                 font_size=8,height=1,padx=10).grid(row=0,column=4,rowspan=2,padx=4,pady=6,sticky="ns")

        self.tabs = ttk.Notebook(self.root)
        self.tabs.pack(expand=1,fill="both",padx=10,pady=(6,10))
        self.setup_tab_atletas()
        self.setup_tab_jogos()
        self.setup_tab_convocacao()
        self.setup_tab_estatisticas()
        self.setup_tab_estoque()
        self.setup_tab_atividades()
        self.setup_tab_comunicacao()
        self.setup_tab_configuracoes()
        self.refresh_all()

    # ── Gerenciar listas globais (modalidades/categorias/naipes)
    def tela_gerenciar_listas(self):
        w = tk.Toplevel(self.root); w.title("Gerenciar Listas do Sistema")
        w.geometry("720x520"); w.configure(bg=C_FUNDO)
        w.transient(self.root); w.grab_set()
        tk.Label(w,text="⚙️  LISTAS DO SISTEMA",bg=C_FUNDO,fg=C_PRIMARIO,
                 font=("Segoe UI",14,"bold")).pack(pady=(14,4))
        nb = ttk.Notebook(w); nb.pack(fill="both",expand=True,padx=14,pady=(0,10))

        def aba_lista(tabela,titulo):
            fr = ttk.Frame(nb); nb.add(fr,text=f"  {titulo}  ")
            top = tk.Frame(fr,bg="white"); top.pack(fill="x",padx=8,pady=8)
            e = tk.Entry(top,font=("Segoe UI",11),bd=1,relief="solid",width=28); e.pack(side="left",padx=(0,8))
            tree = ttk.Treeview(fr,columns=("id","nome"),show="headings",height=12)
            tree.heading("id",text=""); tree.heading("nome",text=titulo)
            tree.column("id",width=0,stretch=False); tree.column("nome",width=320,anchor="w")
            sc = ttk.Scrollbar(fr,orient="vertical",command=tree.yview)
            tree.configure(yscrollcommand=sc.set)

            def carregar():
                for i in tree.get_children(): tree.delete(i)
                conn = sqlite3.connect(db_path); cur = conn.cursor()
                cur.execute(f"SELECT id,nome FROM {tabela} ORDER BY nome")
                for r in cur.fetchall(): tree.insert("","end",values=r)
                conn.close()

            def adicionar(ev=None):
                nome = e.get().strip()
                if not nome: return
                conn = sqlite3.connect(db_path); cur = conn.cursor()
                try:
                    cur.execute(f"INSERT INTO {tabela} (nome) VALUES (?)",(nome,))
                    conn.commit(); e.delete(0,"end"); carregar(); self._refresh_filtros()
                except sqlite3.IntegrityError: messagebox.showwarning("Aviso","Já existe!",parent=w)
                finally: conn.close()

            def excluir():
                sel = tree.selection()
                if not sel: return
                vals = tree.item(sel[0])["values"]
                if messagebox.askyesno("Confirmar",f"Excluir '{vals[1]}'?",parent=w):
                    conn = sqlite3.connect(db_path); cur = conn.cursor()
                    cur.execute(f"DELETE FROM {tabela} WHERE id=?",(vals[0],))
                    conn.commit(); conn.close(); carregar(); self._refresh_filtros()

            def renomear():
                sel = tree.selection()
                if not sel: return
                vals = tree.item(sel[0])["values"]
                d = tk.Toplevel(w); d.title("Renomear"); d.geometry("320x150")
                d.configure(bg="white"); d.transient(w); d.grab_set()
                tk.Label(d,text="Novo nome:",bg="white",font=("Segoe UI",10)).pack(pady=(20,6))
                en = tk.Entry(d,font=("Segoe UI",11),bd=1,relief="solid",width=22)
                en.insert(0,vals[1]); en.pack(pady=(0,14))
                def ok(ev=None):
                    nn = en.get().strip()
                    if not nn: return
                    conn2 = sqlite3.connect(db_path); cur2 = conn2.cursor()
                    cur2.execute(f"UPDATE {tabela} SET nome=? WHERE id=?",(nn,vals[0]))
                    conn2.commit(); conn2.close(); d.destroy(); carregar(); self._refresh_filtros()
                en.bind("<Return>",ok)
                make_btn(d,"SALVAR",ok,bg=C_PRIMARIO,font_size=10,height=1).pack(fill="x",padx=40)
                en.focus(); en.select_range(0,"end")

            make_btn(top,"+ ADICIONAR",adicionar,bg=C_PRIMARIO,font_size=9,height=1).pack(side="left",padx=2)
            make_btn(top,"✏️ RENOMEAR",renomear,bg=C_SECUNDARIO,font_size=9,height=1).pack(side="left",padx=2)
            make_btn(top,"🗑️ EXCLUIR", excluir, bg=C_PERIGO,   font_size=9,height=1).pack(side="left",padx=2)
            e.bind("<Return>",adicionar)
            tree.pack(side="left",fill="both",expand=True,padx=(8,0),pady=(0,8))
            sc.pack(side="right",fill="y",pady=(0,8),padx=(0,8))
            carregar()

        aba_lista("modalidades","Modalidades")
        aba_lista("categorias","Categorias")
        aba_lista("naipes","Naipes")

    # ═══════════════════════════════════════════════════════════
    # ABA ATLETAS
    # ═══════════════════════════════════════════════════════════
    def setup_tab_atletas(self):
        tab = ttk.Frame(self.tabs); self.tabs.add(tab,text="  👥 ATLETAS  ")

        # ── Barra de busca / filtro ───────────────────────────
        fb = tk.Frame(tab,bg=C_FUNDO); fb.pack(fill="x",padx=10,pady=(8,2))
        tk.Label(fb,text="🔍 Buscar:",bg=C_FUNDO,font=("Segoe UI",9,"bold"),fg=C_PRIMARIO).pack(side="left",padx=(0,4))
        self.at_busca = tk.Entry(fb,font=("Segoe UI",10),bd=1,relief="solid",width=30)
        self.at_busca.pack(side="left",padx=(0,8))
        self.at_busca.bind("<KeyRelease>",lambda e: self._filtrar_atletas())
        make_btn(fb,"✕",lambda: (self.at_busca.delete(0,"end"), self._filtrar_atletas()),
                 bg="#aaa",font_size=8,height=1,padx=6).pack(side="left")
        tk.Label(fb,text="  Clique no cabeçalho de qualquer coluna para ordenar.",
                 bg=C_FUNDO,fg="#888",font=("Segoe UI",8)).pack(side="left",padx=14)

        lf = tk.Frame(tab,bg="white"); lf.pack(side="left",expand=True,fill="both",padx=10,pady=(0,10))
        sub_nb = ttk.Notebook(lf); sub_nb.pack(fill="both",expand=True)

        # Sub-aba atletas
        aba_at = ttk.Frame(sub_nb); sub_nb.add(aba_at,text="  Atletas  ")
        cols = ("ID","Nome","Modalidade","Categoria","Naipe")
        self.tree_at = ttk.Treeview(aba_at,columns=cols,show="headings",height=20)
        ws = {"ID":40,"Nome":240,"Modalidade":110,"Categoria":100,"Naipe":110}
        for col in cols:
            self.tree_at.heading(col,text=col)
            self.tree_at.column(col,width=ws[col],anchor="w" if col=="Nome" else "center")
        sc = ttk.Scrollbar(aba_at,orient="vertical",command=self.tree_at.yview)
        self.tree_at.configure(yscrollcommand=sc.set)
        self.tree_at.pack(side="left",expand=True,fill="both"); sc.pack(side="right",fill="y")
        self._make_sortable(self.tree_at,"tree_at")

        # Sub-aba técnicos
        aba_tec = ttk.Frame(sub_nb); sub_nb.add(aba_tec,text="  Técnicos / Comissão  ")
        cols_t = ("ID","Nome","Cargo","Modalidade","Categoria","Naipe")
        self.tree_tec = ttk.Treeview(aba_tec,columns=cols_t,show="headings",height=20)
        ws_t = {"ID":40,"Nome":200,"Cargo":130,"Modalidade":110,"Categoria":100,"Naipe":110}
        for col in cols_t:
            self.tree_tec.heading(col,text=col)
            self.tree_tec.column(col,width=ws_t[col],anchor="w" if col=="Nome" else "center")
        sc_t = ttk.Scrollbar(aba_tec,orient="vertical",command=self.tree_tec.yview)
        self.tree_tec.configure(yscrollcommand=sc_t.set)
        self.tree_tec.pack(side="left",expand=True,fill="both"); sc_t.pack(side="right",fill="y")
        self._make_sortable(self.tree_tec,"tree_tec")

        # ── Painel direito ────────────────────────────────────
        af_outer = tk.Frame(tab,bg=C_FUNDO,width=250)
        af_outer.pack(side="right",fill="y",pady=(0,10)); af_outer.pack_propagate(False)
        canvas_af = tk.Canvas(af_outer,bg=C_FUNDO,highlightthickness=0)
        scroll_af = ttk.Scrollbar(af_outer,orient="vertical",command=canvas_af.yview)
        canvas_af.configure(yscrollcommand=scroll_af.set)
        scroll_af.pack(side="right",fill="y"); canvas_af.pack(side="left",fill="both",expand=True)
        af = tk.Frame(canvas_af,bg=C_FUNDO,padx=6)
        af_win = canvas_af.create_window((0,0),window=af,anchor="nw")

        def _on_af(e):
            canvas_af.configure(scrollregion=canvas_af.bbox("all"))
            canvas_af.itemconfig(af_win,width=canvas_af.winfo_width())
        af.bind("<Configure>",_on_af)

        def sec(parent, titulo):
            return tk.LabelFrame(parent,text=f" {titulo} ",bg=C_FUNDO,
                                 font=("Segoe UI",9,"bold"),fg=C_PRIMARIO)

        # Adicionar atleta
        cf2 = sec(af,"➕ Adicionar Atleta"); cf2.pack(fill="x",pady=(4,6))
        tk.Label(cf2,text="Nome Completo:",bg=C_FUNDO,font=("Segoe UI",9)).pack(anchor="w",padx=5,pady=(5,0))
        self.at_nome = tk.Entry(cf2,font=("Segoe UI",10),bd=1,relief="solid")
        self.at_nome.pack(fill="x",padx=5,pady=4)
        self.at_nome.bind("<Return>",lambda e: self.add_atleta())
        make_btn(cf2,"➕ ADICIONAR ATLETA",self.add_atleta,bg=C_PRIMARIO,font_size=9,height=1
                 ).pack(fill="x",padx=5,pady=(0,6))

        ef = sec(af,"Editar / Excluir Atleta"); ef.pack(fill="x",pady=(0,6))
        make_btn(ef,"✏️  EDITAR ATLETA",   self.editar_atleta,  bg=C_SECUNDARIO,font_size=9,height=1
                 ).pack(fill="x",padx=5,pady=(6,3))
        make_btn(ef,"🗑️  EXCLUIR ATLETA",  self.excluir_atleta, bg=C_PERIGO,    font_size=9,height=1
                 ).pack(fill="x",padx=5,pady=(0,6))

        # Adicionar técnico
        tf2 = sec(af,"🧑‍💼 Adicionar Técnico"); tf2.pack(fill="x",pady=(0,6))
        tk.Label(tf2,text="Nome:",bg=C_FUNDO,font=("Segoe UI",9)).pack(anchor="w",padx=5,pady=(5,0))
        self.tec_nome = tk.Entry(tf2,font=("Segoe UI",10),bd=1,relief="solid")
        self.tec_nome.pack(fill="x",padx=5,pady=(0,4))
        tk.Label(tf2,text="Cargo:",bg=C_FUNDO,font=("Segoe UI",9)).pack(anchor="w",padx=5)
        self.tec_cargo = ttk.Combobox(tf2,
            values=["Técnico","Auxiliar Técnico","Preparador Físico","Fisioterapeuta","Coordenador","Outros"],
            font=("Segoe UI",9),state="readonly",width=22)
        self.tec_cargo.current(0); self.tec_cargo.pack(fill="x",padx=5,pady=(0,4))
        make_btn(tf2,"➕ ADD TÉCNICO",self.add_tecnico,bg=C_ACENTO,fg=C_PRIMARIO,font_size=9,height=1
                 ).pack(fill="x",padx=5,pady=(0,6))

        tef = sec(af,"Editar / Excluir Técnico"); tef.pack(fill="x",pady=(0,6))
        make_btn(tef,"✏️  EDITAR TÉCNICO",  self.editar_tecnico,  bg=C_SECUNDARIO,font_size=9,height=1
                 ).pack(fill="x",padx=5,pady=(6,3))
        make_btn(tef,"🗑️  EXCLUIR TÉCNICO", self.excluir_tecnico, bg=C_PERIGO,    font_size=9,height=1
                 ).pack(fill="x",padx=5,pady=(0,6))

        pf = sec(af,"Documentos PDF"); pf.pack(fill="x",pady=(0,6))
        make_btn(pf,"📄  LISTA VISITANTE",    lambda:self.gerar_pdf_elite("visitante"),
                 bg=C_PRIMARIO,font_size=9,height=1).pack(fill="x",padx=5,pady=(6,3))
        make_btn(pf,"📋  DISPENSA ED. FÍSICA",lambda:self.gerar_pdf_elite("dispensa"),
                 bg=C_ACENTO,fg=C_PRIMARIO,font_size=9,height=1).pack(fill="x",padx=5,pady=(0,6))

    def _filtrar_atletas(self):
        termo = self.at_busca.get().lower().strip()
        m,c,n = self.mod_sel.get(),self.cat_sel.get(),self.nai_sel.get()
        for i in self.tree_at.get_children(): self.tree_at.delete(i)
        conn = sqlite3.connect(db_path); cur = conn.cursor()
        cur.execute("SELECT id,nome,mod,cat,naipe FROM atletas WHERE mod=? AND cat=? AND naipe=?",(m,c,n))
        for r in cur.fetchall():
            if not termo or termo in str(r[1]).lower():
                self.tree_at.insert("","end",values=r)
        conn.close()

    def add_atleta(self):
        nome = self.at_nome.get().strip()
        if not nome: return
        conn = sqlite3.connect(db_path); c = conn.cursor()
        c.execute("INSERT INTO atletas (nome,mod,cat,naipe) VALUES (?,?,?,?)",
                  (nome,self.mod_sel.get(),self.cat_sel.get(),self.nai_sel.get()))
        conn.commit(); conn.close(); self.at_nome.delete(0,"end"); self.refresh_all()

    def editar_atleta(self):
        sel = self.tree_at.selection()
        if not sel: messagebox.showwarning("Aviso","Selecione um atleta."); return
        vals = self.tree_at.item(sel[0])["values"]; at_id = vals[0]
        conn = sqlite3.connect(db_path); cur = conn.cursor()
        cur.execute("SELECT nome,mod,cat,naipe FROM atletas WHERE id=?",(at_id,))
        row = cur.fetchone(); conn.close()
        if not row: return
        w = tk.Toplevel(self.root); w.title("Editar Atleta"); w.geometry("420x330")
        w.configure(bg="white"); w.transient(self.root); w.grab_set()
        tk.Label(w,text="EDITAR ATLETA",bg="white",fg=C_PRIMARIO,font=("Segoe UI",13,"bold")).pack(pady=(16,8))
        fm = tk.Frame(w,bg="white",padx=30); fm.pack(fill="both")
        tk.Label(fm,text="Nome Completo:",bg="white",font=("Segoe UI",10)).pack(anchor="w")
        e_nome = tk.Entry(fm,font=("Segoe UI",11),bd=1,relief="solid")
        e_nome.insert(0,row[0]); e_nome.pack(fill="x",pady=(0,10))
        tk.Label(fm,text="Modalidade:",bg="white",font=("Segoe UI",10)).pack(anchor="w")
        v_mod = tk.StringVar(value=row[1])
        ttk.Combobox(fm,textvariable=v_mod,values=self._get_modalidades(),state="readonly").pack(fill="x",pady=(0,10))
        tk.Label(fm,text="Categoria:",bg="white",font=("Segoe UI",10)).pack(anchor="w")
        v_cat = tk.StringVar(value=row[2])
        ttk.Combobox(fm,textvariable=v_cat,values=self._get_categorias(),state="readonly").pack(fill="x",pady=(0,10))
        tk.Label(fm,text="Naipe:",bg="white",font=("Segoe UI",10)).pack(anchor="w")
        v_nai = tk.StringVar(value=row[3])
        ttk.Combobox(fm,textvariable=v_nai,values=self._get_naipes(),state="readonly").pack(fill="x",pady=(0,14))
        def salvar(ev=None):
            n = e_nome.get().strip()
            if not n: return
            conn2 = sqlite3.connect(db_path); c2 = conn2.cursor()
            c2.execute("UPDATE atletas SET nome=?,mod=?,cat=?,naipe=? WHERE id=?",(n,v_mod.get(),v_cat.get(),v_nai.get(),at_id))
            conn2.commit(); conn2.close(); w.destroy(); self.refresh_all()
        e_nome.bind("<Return>",salvar)
        make_btn(w,"SALVAR",salvar,bg=C_PRIMARIO,font_size=11,height=1).pack(fill="x",padx=30,pady=10)
        e_nome.focus(); e_nome.select_range(0,"end")

    def excluir_atleta(self):
        sel = self.tree_at.selection()
        if not sel: messagebox.showwarning("Aviso","Selecione um atleta."); return
        vals = self.tree_at.item(sel[0])["values"]
        if messagebox.askyesno("Confirmar",f"Excluir '{vals[1]}'?"):
            conn = sqlite3.connect(db_path); c = conn.cursor()
            c.execute("DELETE FROM atletas WHERE id=?",(vals[0],))
            conn.commit(); conn.close(); self.refresh_all()

    def add_tecnico(self):
        nome = self.tec_nome.get().strip()
        if not nome: messagebox.showwarning("Aviso","Informe o nome."); return
        conn = sqlite3.connect(db_path); c = conn.cursor()
        c.execute("INSERT INTO tecnicos (nome,mod,cat,naipe,cargo) VALUES (?,?,?,?,?)",
                  (nome,self.mod_sel.get(),self.cat_sel.get(),self.nai_sel.get(),self.tec_cargo.get()))
        conn.commit(); conn.close(); self.tec_nome.delete(0,"end"); self.refresh_all()

    def editar_tecnico(self):
        sel = self.tree_tec.selection()
        if not sel: messagebox.showwarning("Aviso","Selecione um técnico."); return
        vals = self.tree_tec.item(sel[0])["values"]; tec_id = vals[0]
        conn = sqlite3.connect(db_path); cur = conn.cursor()
        cur.execute("SELECT nome,cargo,mod,cat,naipe FROM tecnicos WHERE id=?",(tec_id,))
        row = cur.fetchone(); conn.close()
        if not row: return
        w = tk.Toplevel(self.root); w.title("Editar Técnico"); w.geometry("420x380")
        w.configure(bg="white"); w.transient(self.root); w.grab_set()
        tk.Label(w,text="EDITAR TÉCNICO",bg="white",fg=C_PRIMARIO,font=("Segoe UI",13,"bold")).pack(pady=(16,8))
        fm = tk.Frame(w,bg="white",padx=30); fm.pack(fill="both")
        tk.Label(fm,text="Nome:",bg="white",font=("Segoe UI",10)).pack(anchor="w")
        e_nome = tk.Entry(fm,font=("Segoe UI",11),bd=1,relief="solid")
        e_nome.insert(0,row[0]); e_nome.pack(fill="x",pady=(0,10))
        tk.Label(fm,text="Cargo:",bg="white",font=("Segoe UI",10)).pack(anchor="w")
        e_cargo = ttk.Combobox(fm,values=["Técnico","Auxiliar Técnico","Preparador Físico","Fisioterapeuta","Coordenador","Outros"],
                               font=("Segoe UI",10),state="readonly")
        e_cargo.set(row[1]); e_cargo.pack(fill="x",pady=(0,10))
        tk.Label(fm,text="Modalidade:",bg="white",font=("Segoe UI",10)).pack(anchor="w")
        v_mod = tk.StringVar(value=row[2])
        ttk.Combobox(fm,textvariable=v_mod,values=self._get_modalidades(),state="readonly").pack(fill="x",pady=(0,10))
        tk.Label(fm,text="Categoria:",bg="white",font=("Segoe UI",10)).pack(anchor="w")
        v_cat = tk.StringVar(value=row[3])
        ttk.Combobox(fm,textvariable=v_cat,values=self._get_categorias(),state="readonly").pack(fill="x",pady=(0,10))
        tk.Label(fm,text="Naipe:",bg="white",font=("Segoe UI",10)).pack(anchor="w")
        v_nai = tk.StringVar(value=row[4])
        ttk.Combobox(fm,textvariable=v_nai,values=self._get_naipes(),state="readonly").pack(fill="x",pady=(0,14))
        def salvar(ev=None):
            n = e_nome.get().strip()
            if not n: return
            conn2 = sqlite3.connect(db_path); c2 = conn2.cursor()
            c2.execute("UPDATE tecnicos SET nome=?,cargo=?,mod=?,cat=?,naipe=? WHERE id=?",
                       (n,e_cargo.get(),v_mod.get(),v_cat.get(),v_nai.get(),tec_id))
            conn2.commit(); conn2.close(); w.destroy(); self.refresh_all()
        e_nome.bind("<Return>",salvar)
        make_btn(w,"SALVAR",salvar,bg=C_PRIMARIO,font_size=11,height=1).pack(fill="x",padx=30,pady=10)
        e_nome.focus(); e_nome.select_range(0,"end")

    def excluir_tecnico(self):
        sel = self.tree_tec.selection()
        if not sel: messagebox.showwarning("Aviso","Selecione um técnico."); return
        vals = self.tree_tec.item(sel[0])["values"]
        if messagebox.askyesno("Confirmar",f"Excluir '{vals[1]}'?"):
            conn = sqlite3.connect(db_path); c = conn.cursor()
            c.execute("DELETE FROM tecnicos WHERE id=?",(vals[0],))
            conn.commit(); conn.close(); self.refresh_all()

    # ═══════════════════════════════════════════════════════════
    # ABA JOGOS
    # ═══════════════════════════════════════════════════════════
    def setup_tab_jogos(self):
        tab = ttk.Frame(self.tabs); self.tabs.add(tab,text="  📅 AGENDA DE JOGOS  ")

        f = tk.LabelFrame(tab,text=" Lançar Novo Jogo ",font=("Segoe UI",9,"bold"),fg=C_PRIMARIO)
        f.pack(fill="x",padx=10,pady=10)

        def lbl_ent(parent,text,row,col,w=10,tt=""):
            tk.Label(parent,text=text).grid(row=row,column=col,padx=6,sticky="w")
            e = tk.Entry(parent,font=("Segoe UI",10),bd=1,relief="solid",width=w)
            e.grid(row=row+1,column=col,padx=6,pady=(0,10))
            if tt: tooltip(e,tt)
            return e

        self.jg_d   = lbl_ent(f,"Data (DDMMAA):",0,0,tt="Ex: 250726 → 25/07/26")
        self.jg_h   = lbl_ent(f,"Hora Jogo (HHMM):",0,1,9,"Ex: 1430 → 14:30")
        self.jg_ret = lbl_ent(f,"Retorno (HHMM):",0,2,9,"Ex: 1800 → 18:00")
        self.jg_c   = lbl_ent(f,"Competição:",0,3,20)
        self.jg_a   = lbl_ent(f,"Adversário:",0,4,20)
        tk.Label(f,text="Mando:").grid(row=2,column=0,padx=6,sticky="w")
        self.jg_t = ttk.Combobox(f,values=["CASA","FORA"],width=9,state="readonly")
        self.jg_t.grid(row=3,column=0,padx=6,pady=(0,10)); self.jg_t.current(0)
        tk.Label(f,text="Local / Quadra / Destino:").grid(row=2,column=1,columnspan=2,padx=6,sticky="w")
        self.jg_local = tk.Entry(f,font=("Segoe UI",10),bd=1,relief="solid",width=34)
        self.jg_local.grid(row=3,column=1,columnspan=2,padx=6,pady=(0,10))
        tk.Label(f,text="Saída (antecedência):").grid(row=2,column=3,padx=6,sticky="w")
        self.jg_ant = ttk.Combobox(f,values=["N/A","1h antes","2h antes"],width=14,state="readonly")
        self.jg_ant.grid(row=3,column=3,padx=6,pady=(0,10)); self.jg_ant.current(0)
        entries = [self.jg_d,self.jg_h,self.jg_ret,self.jg_c,self.jg_a,self.jg_local]
        for i,e in enumerate(entries):
            nxt = entries[i+1] if i<len(entries)-1 else None
            e.bind("<Return>",(lambda n: lambda ev: n.focus())(nxt) if nxt else lambda ev: self.add_jogo())
        make_btn(f,"  ➕  AGENDAR  ",self.add_jogo,bg=C_PRIMARIO,font_size=10,height=2
                 ).grid(row=2,column=4,rowspan=2,padx=14,pady=(0,10),sticky="ns")

        # Barra de filtro / busca
        fb = tk.Frame(tab,bg=C_FUNDO); fb.pack(fill="x",padx=10,pady=(0,4))
        self.lbl_casa  = tk.Label(fb,text="🏠 Em Casa: 0",bg=C_PRIMARIO,fg="white",font=("Segoe UI",10,"bold"),padx=12,pady=4)
        self.lbl_casa.pack(side="left",padx=(0,4))
        self.lbl_fora  = tk.Label(fb,text="✈️ Fora: 0",bg=C_SECUNDARIO,fg="white",font=("Segoe UI",10,"bold"),padx=12,pady=4)
        self.lbl_fora.pack(side="left",padx=(0,4))
        self.lbl_total = tk.Label(fb,text="📋 Total: 0",bg="#555",fg="white",font=("Segoe UI",10,"bold"),padx=12,pady=4)
        self.lbl_total.pack(side="left")
        chk = tk.Checkbutton(fb,text="👁 Ver TODOS",variable=self.ver_todos_jogos,command=self.refresh_all,
                             bg=C_FUNDO,fg=C_PRIMARIO,font=("Segoe UI",9,"bold"),
                             activebackground=C_FUNDO,selectcolor=C_FUNDO)
        chk.pack(side="left",padx=12)
        tk.Label(fb,text="🔍",bg=C_FUNDO,font=("Segoe UI",10)).pack(side="left",padx=(10,2))
        self.jg_busca = tk.Entry(fb,font=("Segoe UI",10),bd=1,relief="solid",width=22)
        self.jg_busca.pack(side="left"); self.jg_busca.bind("<KeyRelease>",lambda e: self._filtrar_jogos())
        make_btn(fb,"✕",lambda:(self.jg_busca.delete(0,"end"),self._filtrar_jogos()),
                 bg="#aaa",font_size=8,height=1,padx=6).pack(side="left",padx=4)
        tk.Label(fb,text="Coluna:",bg=C_FUNDO,font=("Segoe UI",9)).pack(side="left",padx=(10,2))
        self.jg_busca_col = ttk.Combobox(fb,values=["Adversário","Competição","Local","Data"],
                                          width=13,state="readonly")
        self.jg_busca_col.current(0); self.jg_busca_col.pack(side="left")
        self.jg_busca_col.bind("<<ComboboxSelected>>",lambda e: self._filtrar_jogos())

        cols = ("ID","Data","Hora","Modalidade","Categoria","Adversário","Competição","Mando","Local","Saída","Retorno","Placar")
        self.tree_jg = ttk.Treeview(tab,columns=cols,show="headings",height=12)
        ws2 = {"ID":0,"Data":85,"Hora":62,"Modalidade":75,"Categoria":70,"Adversário":130,
               "Competição":130,"Mando":65,"Local":130,"Saída":68,"Retorno":68,"Placar":80}
        for col in cols:
            self.tree_jg.heading(col,text=col)
            self.tree_jg.column(col,width=ws2[col],anchor="center")
        self.tree_jg.column("ID",width=0,stretch=False)
        self.tree_jg.column("Adversário",anchor="w"); self.tree_jg.column("Local",anchor="w")
        sc2 = ttk.Scrollbar(tab,orient="vertical",command=self.tree_jg.yview)
        self.tree_jg.configure(yscrollcommand=sc2.set)
        self._make_sortable(self.tree_jg,"tree_jg")

        # ── Botões com espaço adequado ────────────────────────
        btn_row = tk.Frame(tab,bg=C_FUNDO); btn_row.pack(fill="x",padx=10,pady=(4,4))
        btns_jg = [
            ("✏️  EDITAR JOGO",    self.editar_jogo,        C_SECUNDARIO),
            ("🗑️  EXCLUIR JOGO",   self.excluir_jogo,       C_PERIGO),
            ("⚽  REGISTRAR PLACAR",self.registrar_placar,   C_SUCESSO),
            ("📊  EXCEL VIAGENS",  self.gerar_excel_viagens, "#1a7a3c"),
            ("📄  PDF AGENDA",     self.gerar_pdf_agenda,    C_PRIMARIO),
        ]
        for i,(txt,cmd,cor) in enumerate(btns_jg):
            make_btn(btn_row,txt,cmd,bg=cor,font_size=9,height=1,padx=12
                     ).grid(row=0,column=i,padx=5,pady=4,sticky="ew")
            btn_row.grid_columnconfigure(i,weight=1)

        self.tree_jg.pack(fill="both",expand=True,padx=(10,0),pady=(0,4),side="left")
        sc2.pack(side="right",fill="y",padx=(0,10),pady=(0,4))

    def _filtrar_jogos(self):
        termo = self.jg_busca.get().lower().strip()
        col_map = {"Adversário":"Adversário","Competição":"Competição","Local":"Local","Data":"Data"}
        col = col_map.get(self.jg_busca_col.get(),"Adversário")
        m,c = self.mod_sel.get(),self.cat_sel.get()
        for i in self.tree_jg.get_children(): self.tree_jg.delete(i)
        conn = sqlite3.connect(db_path); cur = conn.cursor()
        if self.ver_todos_jogos.get():
            cur.execute("""SELECT id,data,hora,mod,cat,adv,comp,tipo,local,saida,retorno,
                                  placar_c7s||CASE WHEN placar_c7s!='' THEN ' x ' ELSE '' END||placar_adv FROM jogos""")
        else:
            cur.execute("""SELECT id,data,hora,mod,cat,adv,comp,tipo,local,saida,retorno,
                                  placar_c7s||CASE WHEN placar_c7s!='' THEN ' x ' ELSE '' END||placar_adv
                           FROM jogos WHERE mod=? AND cat=?""",(m,c))
        rows = sorted(cur.fetchall(),key=lambda r: parse_data(str(r[1]))); conn.close()
        col_idx = {"Data":1,"Adversário":5,"Competição":6,"Local":8}
        ci = col_idx.get(col,5)
        for r in rows:
            if not termo or termo in str(r[ci]).lower():
                self.tree_jg.insert("","end",values=r)

    def add_jogo(self):
        d,h,ret = self.jg_d.get().strip(),self.jg_h.get().strip(),self.jg_ret.get().strip()
        if len(d)<6 or len(h)<4: return messagebox.showwarning("Erro","Data ou Hora incompleta!")
        data_f = f"{d[:2]}/{d[2:4]}/20{d[4:]}"; hora_f = f"{h[:2]}:{h[2:]}"
        ret_f  = f"{ret[:2]}:{ret[2:]}" if len(ret)>=4 else "A definir"
        conn = sqlite3.connect(db_path); c = conn.cursor()
        c.execute("INSERT INTO jogos (data,hora,mod,cat,adv,comp,tipo,local,saida,retorno) VALUES (?,?,?,?,?,?,?,?,?,?)",
                  (data_f,hora_f,self.mod_sel.get(),self.cat_sel.get(),
                   self.jg_a.get(),self.jg_c.get(),self.jg_t.get(),self.jg_local.get(),self.jg_ant.get(),ret_f))
        conn.commit(); conn.close()
        for we in [self.jg_d,self.jg_h,self.jg_ret,self.jg_a,self.jg_c,self.jg_local]: we.delete(0,"end")
        self.jg_t.current(0); self.jg_ant.current(0); self.refresh_all()

    def editar_jogo(self):
        sel = self.tree_jg.selection()
        if not sel: messagebox.showwarning("Aviso","Selecione um jogo."); return
        vals = self.tree_jg.item(sel[0])["values"]; jid = vals[0]
        conn = sqlite3.connect(db_path); cur = conn.cursor()
        cur.execute("SELECT data,hora,mod,cat,adv,comp,tipo,local,saida,retorno FROM jogos WHERE id=?",(jid,))
        row = cur.fetchone(); conn.close()
        if not row: return
        win = tk.Toplevel(self.root); win.title("Editar Jogo"); win.geometry("540x560")
        win.configure(bg="white"); win.transient(self.root); win.grab_set()
        tk.Label(win,text="EDITAR JOGO",bg="white",fg=C_PRIMARIO,font=("Segoe UI",13,"bold")).pack(pady=(14,6))
        form = tk.Frame(win,bg="white",padx=30); form.pack(fill="both",expand=True)
        form.grid_columnconfigure(0,weight=1)

        def rle(lbl_txt,val,r):
            tk.Label(form,text=lbl_txt,bg="white",font=("Segoe UI",9,"bold"),fg=C_SECUNDARIO).grid(row=r,column=0,sticky="w",pady=(8,0))
            e = tk.Entry(form,font=("Segoe UI",10),bd=1,relief="solid")
            e.insert(0,str(val)); e.grid(row=r+1,column=0,sticky="ew",pady=(2,0)); return e

        def rlcb(lbl_txt,val,vals_list,r):
            tk.Label(form,text=lbl_txt,bg="white",font=("Segoe UI",9,"bold"),fg=C_SECUNDARIO).grid(row=r,column=0,sticky="w",pady=(8,0))
            v = tk.StringVar(value=val)
            ttk.Combobox(form,textvariable=v,values=vals_list,state="readonly").grid(row=r+1,column=0,sticky="ew",pady=(2,0)); return v

        e_data  = rle("Data (DD/MM/AAAA):",   row[0],0)
        e_hora  = rle("Hora do Jogo (HH:MM):",row[1],2)
        e_ret   = rle("Retorno (HH:MM):",      row[9] or "",4)
        e_adv   = rle("Adversário:",           row[4],6)
        e_comp  = rle("Competição:",           row[5],8)
        e_local = rle("Local / Quadra:",       row[7],10)
        v_mod   = rlcb("Modalidade:",row[2],self._get_modalidades(),12)
        v_cat   = rlcb("Categoria:", row[3],self._get_categorias(), 14)
        v_tipo  = rlcb("Mando:",     row[6],["CASA","FORA"],        16)

        def salvar(ev=None):
            conn2 = sqlite3.connect(db_path); c2 = conn2.cursor()
            c2.execute("UPDATE jogos SET data=?,hora=?,retorno=?,adv=?,comp=?,local=?,mod=?,cat=?,tipo=? WHERE id=?",
                       (e_data.get(),e_hora.get(),e_ret.get(),e_adv.get(),e_comp.get(),
                        e_local.get(),v_mod.get(),v_cat.get(),v_tipo.get(),jid))
            conn2.commit(); conn2.close(); win.destroy(); self.refresh_all()

        make_btn(win,"SALVAR ALTERAÇÕES",salvar,bg=C_PRIMARIO,font_size=11,height=1).pack(fill="x",padx=30,pady=14)
        e_data.focus()

    def excluir_jogo(self):
        sel = self.tree_jg.selection()
        if not sel: messagebox.showwarning("Aviso","Selecione um jogo."); return
        vals = self.tree_jg.item(sel[0])["values"]
        if messagebox.askyesno("Confirmar",f"Excluir C7S x {vals[5]} em {vals[1]}?"):
            conn = sqlite3.connect(db_path); c = conn.cursor()
            c.execute("DELETE FROM jogos WHERE id=?",(vals[0],))
            conn.commit(); conn.close(); self.refresh_all()

    def registrar_placar(self):
        sel = self.tree_jg.selection()
        if not sel: messagebox.showwarning("Aviso","Selecione um jogo."); return
        vals = self.tree_jg.item(sel[0])["values"]; jid = vals[0]
        win = tk.Toplevel(self.root); win.title("Placar"); win.geometry("380x230")
        win.configure(bg="white"); win.transient(self.root); win.grab_set()
        tk.Label(win,text="REGISTRAR PLACAR",bg="white",fg=C_PRIMARIO,font=("Segoe UI",13,"bold")).pack(pady=(16,4))
        tk.Label(win,text=f"C7S x {vals[5]}  —  {vals[1]}",bg="white",fg=C_SECUNDARIO,font=("Segoe UI",10)).pack(pady=(0,12))
        rf = tk.Frame(win,bg="white"); rf.pack()
        tk.Label(rf,text="C7S:",bg="white",font=("Segoe UI",11,"bold"),fg=C_PRIMARIO).grid(row=0,column=0,padx=10)
        p1 = tk.Entry(rf,font=("Segoe UI",18,"bold"),width=4,justify="center",bd=1,relief="solid"); p1.grid(row=0,column=1,padx=6)
        tk.Label(rf,text="x",bg="white",font=("Segoe UI",16)).grid(row=0,column=2,padx=6)
        tk.Label(rf,text=f"{vals[5]}:",bg="white",font=("Segoe UI",11,"bold"),fg=C_SECUNDARIO).grid(row=0,column=3,padx=10)
        p2 = tk.Entry(rf,font=("Segoe UI",18,"bold"),width=4,justify="center",bd=1,relief="solid"); p2.grid(row=0,column=4,padx=6)
        pl = str(vals[11])
        if "x" in pl:
            pts = pl.split("x"); p1.insert(0,pts[0].strip()); p2.insert(0,pts[1].strip())
        def salvar(ev=None):
            conn = sqlite3.connect(db_path); c2 = conn.cursor()
            c2.execute("UPDATE jogos SET placar_c7s=?,placar_adv=? WHERE id=?",(p1.get().strip(),p2.get().strip(),jid))
            conn.commit(); conn.close(); win.destroy(); self.refresh_all()
        p1.bind("<Return>",lambda e: p2.focus()); p2.bind("<Return>",salvar)
        make_btn(win,"SALVAR PLACAR",salvar,bg=C_SUCESSO,font_size=11,height=1).pack(fill="x",padx=50,pady=16)
        p1.focus()

    def gerar_excel_viagens(self):
        jogos = self._get_todos_jogos_futuros()
        jogos_fora = [j for j in jogos if j["tipo"]=="FORA"]
        if not jogos_fora: messagebox.showwarning("Aviso","Nenhum jogo FORA futuro."); return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Viagens"
        azul=PatternFill("solid",fgColor="221C89"); cinza=PatternFill("solid",fgColor="EEF0F5")
        branco=PatternFill("solid",fgColor="FFFFFF")
        borda=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
        aln_c=Alignment(horizontal="center",vertical="center",wrap_text=True)
        aln_l=Alignment(horizontal="left",vertical="center",wrap_text=True)
        ws.merge_cells("A1:K1"); ws["A1"]=NOME_ESCOLA
        ws["A1"].font=Font(name="Arial",bold=True,size=14,color="221C89"); ws["A1"].alignment=aln_c
        ws.merge_cells("A2:K2"); ws["A2"]="RELATÓRIO DE VIAGENS — JOGOS FORA DE CASA"
        ws["A2"].font=Font(name="Arial",bold=True,size=11,color="4A4A6A"); ws["A2"].alignment=aln_c
        ws.merge_cells("A3:K3"); ws["A3"]=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A3"].font=Font(name="Arial",italic=True,size=9,color="888888"); ws["A3"].alignment=aln_c
        ws.row_dimensions[1].height=22; ws.row_dimensions[2].height=18; ws.append([])
        headers=["#","Data","Hora Jogo","Modalidade","Categoria","Adversário","Competição","Local (Destino)","Saída (C7S)","Retorno","Placar"]
        ws.append(headers); hr=ws.max_row
        for i,h in enumerate(headers,1):
            cel=ws.cell(hr,i); cel.fill=azul; cel.font=Font(name="Arial",bold=True,size=10,color="FFFFFF")
            cel.alignment=aln_c; cel.border=borda
        ws.row_dimensions[hr].height=24
        for idx,j in enumerate(jogos_fora,1):
            hs=self._calc_hora_saida(j["hora"],j["saida_ant"])
            ret=j["retorno"] if j["retorno"] and j["retorno"] not in ("A definir","") else "A confirmar"
            ws.append([idx,j["data"],j["hora"],j["mod"],j["cat"],j["adv"],j["comp"],j["local"],hs,ret,j.get("placar","")])
            r=ws.max_row; fill=cinza if idx%2==0 else branco
            for ci in range(1,12):
                cel=ws.cell(r,ci); cel.fill=fill; cel.font=Font(name="Arial",size=10)
                cel.alignment=aln_c if ci!=6 else aln_l; cel.border=borda
            ws.row_dimensions[r].height=20
        for col,w3 in zip("ABCDEFGHIJK",[5,12,11,12,10,22,22,28,12,12,10]):
            ws.column_dimensions[col].width=w3
        ws.append([]); ws.append([f"Total de viagens: {len(jogos_fora)}"])
        ws.cell(ws.max_row,1).font=Font(name="Arial",bold=True,size=10,color="221C89")
        fn=f"Viagens_C7S_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(os.path.join(PASTA_DIST,fn))
        messagebox.showinfo("Sucesso ✅",f"Excel salvo em:\ndist/{fn}")

    # ═══════════════════════════════════════════════════════════
    # ABA CONVOCAÇÃO
    # ═══════════════════════════════════════════════════════════
    def setup_tab_convocacao(self):
        tab = ttk.Frame(self.tabs); self.tabs.add(tab,text="  📋 CONVOCAÇÃO  ")
        top = tk.Frame(tab,bg=C_FUNDO); top.pack(fill="x",padx=10,pady=10)
        tk.Label(top,text="Selecione o Jogo:",bg=C_FUNDO,font=("Segoe UI",10,"bold"),fg=C_PRIMARIO).pack(side="left",padx=(0,8))
        self.conv_jogo_var = tk.StringVar()
        self.conv_jogo_cb  = ttk.Combobox(top,textvariable=self.conv_jogo_var,width=55,state="readonly")
        self.conv_jogo_cb.pack(side="left",padx=(0,10))
        self.conv_jogo_cb.bind("<<ComboboxSelected>>",lambda e: self.carregar_convocacao())
        make_btn(top,"📋 CARREGAR",self.carregar_convocacao,bg=C_PRIMARIO,font_size=9,height=1,padx=14).pack(side="left",padx=4)
        make_btn(top,"🔢 NÚMEROS",self.lista_numeros_disponiveis,bg=C_SECUNDARIO,font_size=9,height=1,padx=14).pack(side="right",padx=4)
        make_btn(top,"📄 PDF CONVOCAÇÃO",self.gerar_pdf_convocacao,bg=C_ACENTO,fg=C_PRIMARIO,font_size=9,height=1,padx=14).pack(side="right",padx=4)

        mid = tk.Frame(tab,bg=C_FUNDO); mid.pack(fill="both",expand=True,padx=10)
        lc = tk.Frame(mid,bg="white",bd=1,relief="solid"); lc.pack(side="left",fill="both",expand=True)
        tk.Label(lc,text="ATLETAS DA SELEÇÃO",bg=C_PRIMARIO,fg="white",font=("Segoe UI",10,"bold"),pady=6).pack(fill="x")

        # Busca rápida na convocação
        fb_conv = tk.Frame(lc,bg="#eef0f5",pady=4); fb_conv.pack(fill="x",padx=6)
        tk.Label(fb_conv,text="🔍",bg="#eef0f5",font=("Segoe UI",10)).pack(side="left",padx=(0,2))
        self.conv_busca = tk.Entry(fb_conv,font=("Segoe UI",10),bd=1,relief="solid",width=24)
        self.conv_busca.pack(side="left"); self.conv_busca.bind("<KeyRelease>",lambda e: self._filtrar_convocacao())
        make_btn(fb_conv,"✕",lambda:(self.conv_busca.delete(0,"end"),self._filtrar_convocacao()),
                 bg="#aaa",font_size=8,height=1,padx=6).pack(side="left",padx=4)

        self.tree_conv = ttk.Treeview(lc,columns=("Nome","Status"),show="headings",height=18)
        self.tree_conv.heading("Nome",text="Nome do Atleta")
        self.tree_conv.heading("Status",text="Status")
        self.tree_conv.column("Nome",width=240); self.tree_conv.column("Status",width=130,anchor="center")
        self.tree_conv.tag_configure("Convocado",foreground=C_PRIMARIO)
        self.tree_conv.tag_configure("Presente",foreground=C_SUCESSO)
        self.tree_conv.tag_configure("Ausente",foreground=C_PERIGO)
        self._make_sortable(self.tree_conv,"tree_conv")
        self.tree_conv.pack(fill="both",expand=True,padx=5,pady=5)

        rc = tk.Frame(mid,bg=C_FUNDO,padx=12); rc.pack(side="right",fill="y")
        tk.Label(rc,text="Alterar Status:",bg=C_FUNDO,font=("Segoe UI",10,"bold"),fg=C_PRIMARIO).pack(pady=(16,8))
        for status,cor in [("Convocado",C_PRIMARIO),("Presente",C_SUCESSO),("Ausente",C_PERIGO)]:
            make_btn(rc,f"● {status}",lambda s=status: self.set_status_conv(s),
                     bg=cor,font_size=10,height=2,width=16).pack(pady=5)

    def _filtrar_convocacao(self):
        termo = self.conv_busca.get().lower()
        for iid in self.tree_conv.get_children():
            nome = str(self.tree_conv.item(iid,"values")[0]).lower()
            self.tree_conv.detach(iid) if termo and termo not in nome else self.tree_conv.reattach(iid,"",tk.END)

    def lista_numeros_disponiveis(self):
        conn = sqlite3.connect(db_path); cur = conn.cursor()
        cur.execute("""SELECT e.nome,ev.numero,ev.tamanho,ev.quantidade FROM estoque e
                       JOIN estoque_variantes ev ON ev.estoque_id=e.id
                       WHERE e.tipo='Selecao' AND ev.numero!='' AND ev.quantidade>0
                       ORDER BY e.nome,CAST(ev.numero AS INTEGER)""")
        rows = cur.fetchall(); conn.close()
        win = tk.Toplevel(self.root); win.title("Números Disponíveis"); win.geometry("640x520")
        win.configure(bg=C_FUNDO); win.transient(self.root); win.grab_set()
        tk.Label(win,text="🔢  NÚMEROS DISPONÍVEIS",bg=C_FUNDO,fg=C_PRIMARIO,font=("Segoe UI",13,"bold")).pack(pady=(14,2))
        fr = tk.Frame(win,bg="white",bd=1,relief="solid"); fr.pack(fill="both",expand=True,padx=12,pady=(0,12))
        cols = ("Uniforme","Número","Tamanho","Qtd")
        tree = ttk.Treeview(fr,columns=cols,show="headings",height=18)
        for col in cols: tree.heading(col,text=col); tree.column(col,width=140,anchor="center")
        tree.column("Uniforme",width=210,anchor="w")
        sc = ttk.Scrollbar(fr,orient="vertical",command=tree.yview); tree.configure(yscrollcommand=sc.set)
        if rows:
            for r in rows: tree.insert("","end",values=r)
        else: tree.insert("","end",values=("Nenhum número cadastrado","—","—","—"))
        tree.pack(side="left",fill="both",expand=True,padx=(4,0),pady=4)
        sc.pack(side="right",fill="y",pady=4,padx=(0,4))
        make_btn(win,"📊 EXPORTAR PDF",lambda: self._pdf_numeros_disponiveis(rows),
                 bg=C_PRIMARIO,font_size=10,height=1).pack(fill="x",padx=12,pady=(0,10))

    def _pdf_numeros_disponiveis(self,rows):
        if not rows: messagebox.showwarning("Aviso","Nenhum dado."); return
        pdf = PDF_Elite(); pdf.alias_nb_pages(); pdf.add_page()
        pdf.set_font("Arial","B",14); pdf.set_text_color(34,28,137)
        pdf.cell(190,10,"NÚMEROS DISPONÍVEIS — UNIFORMES DE SELEÇÃO",ln=True,align="C"); pdf.ln(4)
        pdf.set_font("Arial","B",10); pdf.set_fill_color(34,28,137); pdf.set_text_color(255,255,255)
        for h,w in [("Uniforme",80),("Número",30),("Tamanho",40),("Qtd",30)]: pdf.cell(w,10,h,1,0,"C",True)
        pdf.ln(); pdf.set_font("Arial","",10); pdf.set_text_color(0,0,0); z=False
        for unif,num,tam,qty in rows:
            pdf.set_fill_color(244,247,246) if z else pdf.set_fill_color(255,255,255)
            pdf.cell(80,9,str(unif)[:28],1,0,"L",True); pdf.cell(30,9,str(num),1,0,"C",True)
            pdf.cell(40,9,str(tam),1,0,"C",True); pdf.cell(30,9,str(qty),1,1,"C",True); z=not z
        fn=f"Numeros_Disponiveis_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        try: pdf.output(os.path.join(PASTA_DIST,fn)); messagebox.showinfo("Sucesso",f"dist/{fn}")
        except: messagebox.showerror("Erro","Feche o PDF antes.")

    def carregar_convocacao(self):
        m,c = self.mod_sel.get(),self.cat_sel.get()
        conn = sqlite3.connect(db_path); cur = conn.cursor()
        cur.execute("SELECT id,data,hora,adv FROM jogos WHERE mod=? AND cat=? ORDER BY data ASC",(m,c))
        jogos = cur.fetchall()
        self.conv_jogos_map = {f"{j[1]} {j[2]} — C7S x {j[3]}":j[0] for j in jogos}
        self.conv_jogo_cb["values"] = list(self.conv_jogos_map.keys())
        jk = self.conv_jogo_var.get()
        if not jk or jk not in self.conv_jogos_map: conn.close(); return
        jid = self.conv_jogos_map[jk]
        cur.execute("SELECT id,nome FROM atletas WHERE mod=? AND cat=? AND naipe=? ORDER BY nome ASC",
                    (m,c,self.nai_sel.get()))
        atletas = cur.fetchall()
        for i in self.tree_conv.get_children(): self.tree_conv.delete(i)
        for aid,nome in atletas:
            cur.execute("SELECT status FROM convocacoes WHERE jogo_id=? AND atleta_id=?",(jid,aid))
            row = cur.fetchone(); status = row[0] if row else "Convocado"
            if not row:
                cur.execute("INSERT OR IGNORE INTO convocacoes (jogo_id,atleta_id,status) VALUES (?,?,?)",(jid,aid,"Convocado"))
            self.tree_conv.insert("","end",iid=str(aid),values=(nome,status),tags=(status,))
        conn.commit(); conn.close()

    def set_status_conv(self,status):
        sel = self.tree_conv.selection()
        if not sel: messagebox.showwarning("Aviso","Selecione um atleta."); return
        jk = self.conv_jogo_var.get()
        if not jk or jk not in self.conv_jogos_map: return
        jid = self.conv_jogos_map[jk]; aid = int(sel[0])
        conn = sqlite3.connect(db_path); c = conn.cursor()
        c.execute("UPDATE convocacoes SET status=? WHERE jogo_id=? AND atleta_id=?",(status,jid,aid))
        conn.commit(); conn.close(); self.carregar_convocacao()

    def gerar_pdf_convocacao(self):
        jk = self.conv_jogo_var.get()
        if not jk or jk not in self.conv_jogos_map: messagebox.showwarning("Aviso","Selecione um jogo."); return
        jid = self.conv_jogos_map[jk]; m,c,n = self.mod_sel.get(),self.cat_sel.get(),self.nai_sel.get()
        conn = sqlite3.connect(db_path); cur = conn.cursor()
        cur.execute("SELECT data,hora,adv,local FROM jogos WHERE id=?",(jid,)); jogo = cur.fetchone()
        cur.execute("SELECT a.nome,cv.status FROM atletas a JOIN convocacoes cv ON a.id=cv.atleta_id WHERE cv.jogo_id=? ORDER BY a.nome ASC",(jid,))
        lista = cur.fetchall()
        cur.execute("SELECT nome,cargo FROM tecnicos WHERE mod=? AND cat=? AND naipe=? ORDER BY nome ASC",(m,c,n))
        tecs = cur.fetchall(); conn.close()
        pdf = PDF_Elite(); pdf.alias_nb_pages(); pdf.add_page(); pdf.add_watermark()
        pdf.set_font("Arial","B",14); pdf.set_text_color(34,28,137)
        pdf.cell(190,10,"LISTA DE CONVOCACAO",ln=True,align="C")
        pdf.set_font("Arial","",11); pdf.set_text_color(74,74,106)
        pdf.cell(190,8,f"{m} {c} {n}  .  C7S x {jogo[2]}  .  {jogo[0]} as {jogo[1]}",ln=True,align="C")
        pdf.cell(190,7,f"Local: {jogo[3]}",ln=True,align="C"); pdf.ln(6)
        if tecs:
            pdf.set_font("Arial","B",10); pdf.set_fill_color(242,179,26); pdf.set_text_color(34,28,137)
            pdf.cell(190,8,"COMISSAO TECNICA",1,1,"C",True)
            pdf.set_font("Arial","",10); pdf.set_text_color(0,0,0)
            for tnome,tcargo in tecs:
                pdf.cell(130,9,f"  {tnome}",1,0,"L"); pdf.cell(60,9,f"  {tcargo}",1,1,"L")
            pdf.ln(4)
        pdf.set_font("Arial","B",10); pdf.set_fill_color(34,28,137); pdf.set_text_color(255,255,255)
        pdf.cell(15,10,"#",1,0,"C",True); pdf.cell(120,10,"NOME",1,0,"L",True); pdf.cell(55,10,"STATUS",1,1,"C",True)
        pdf.set_font("Arial","",10); z=False
        for i,(nome,status) in enumerate(lista,1):
            pdf.set_fill_color(244,247,246) if z else pdf.set_fill_color(255,255,255)
            pdf.set_text_color(0,0,0); pdf.cell(15,10,str(i),1,0,"C",True); pdf.cell(120,10,f"  {nome}",1,0,"L",True)
            if status=="Presente": pdf.set_text_color(39,174,96)
            elif status=="Ausente": pdf.set_text_color(231,76,60)
            else: pdf.set_text_color(34,28,137)
            pdf.cell(55,10,status,1,1,"C",True); z=not z
        fn=f"Convocacao_{m}_{c}_{jogo[0].replace('/','_')}.pdf"
        try: pdf.output(os.path.join(PASTA_DIST,fn)); messagebox.showinfo("Sucesso",f"dist/{fn}")
        except: messagebox.showerror("Erro","Feche o PDF antes.")

    # ═══════════════════════════════════════════════════════════
    # ABA ESTATÍSTICAS
    # ═══════════════════════════════════════════════════════════
    def setup_tab_estatisticas(self):
        tab = ttk.Frame(self.tabs); self.tabs.add(tab,text="  📊 ESTATÍSTICAS  ")
        self.stats_frame = tk.Frame(tab,bg=C_FUNDO)
        self.stats_frame.pack(fill="both",expand=True,padx=20,pady=20)
        make_btn(tab,"🔄  ATUALIZAR ESTATÍSTICAS",self.atualizar_stats,bg=C_PRIMARIO,font_size=10,height=1
                 ).pack(side="bottom",pady=10)

    def atualizar_stats(self):
        for w in self.stats_frame.winfo_children(): w.destroy()
        m,c = self.mod_sel.get(),self.cat_sel.get()
        conn = sqlite3.connect(db_path); cur = conn.cursor()
        cur.execute("SELECT data,adv,tipo,placar_c7s,placar_adv FROM jogos WHERE mod=? AND cat=? AND placar_c7s!='' ORDER BY data ASC",(m,c))
        jogos = cur.fetchall(); conn.close()
        vit=emp=der=gp=gc=0; rows=[]
        for data,adv,tipo,pc,pa in jogos:
            try:
                pi,ai=int(pc),int(pa); gp+=pi; gc+=ai
                if pi>ai: res="Vitoria"; vit+=1
                elif pi==ai: res="Empate"; emp+=1
                else: res="Derrota"; der+=1
                rows.append((data,adv,tipo,f"{pc} x {pa}",res))
            except: pass
        total=vit+emp+der
        cf2=tk.Frame(self.stats_frame,bg=C_FUNDO); cf2.pack(fill="x",pady=(0,20))
        def card(p,t,v,col,color):
            c2=tk.Frame(p,bg=color,padx=20,pady=14); c2.grid(row=0,column=col,padx=8,sticky="ew")
            tk.Label(c2,text=str(v),bg=color,fg="white",font=("Segoe UI",24,"bold")).pack()
            tk.Label(c2,text=t,bg=color,fg="white",font=("Segoe UI",9)).pack()
            p.grid_columnconfigure(col,weight=1)
        card(cf2,"JOGOS",total,0,C_PRIMARIO); card(cf2,"VITÓRIAS",vit,1,C_SUCESSO)
        card(cf2,"EMPATES",emp,2,"#E67E22"); card(cf2,"DERROTAS",der,3,C_PERIGO)
        card(cf2,"GOLS PRÓ",gp,4,C_SECUNDARIO); card(cf2,"GOLS CONT",gc,5,"#7F8C8D")
        if total>0:
            pct=round((vit/total)*100,1); s=gp-gc
            af2=tk.Frame(self.stats_frame,bg="white",padx=14,pady=8); af2.pack(fill="x",pady=(0,12))
            tk.Label(af2,text=f"Aproveitamento: {pct}%   ·   Saldo: {'+' if s>=0 else ''}{s}",
                     bg="white",fg=C_PRIMARIO,font=("Segoe UI",12,"bold")).pack()
        tr=ttk.Treeview(self.stats_frame,columns=("Data","Adversário","Mando","Placar","Resultado"),show="headings",height=10)
        for col,w2 in [("Data",90),("Adversário",170),("Mando",70),("Placar",90),("Resultado",120)]:
            tr.heading(col,text=col); tr.column(col,width=w2,anchor="center")
        tr.tag_configure("Vitoria",foreground=C_SUCESSO); tr.tag_configure("Derrota",foreground=C_PERIGO)
        tr.tag_configure("Empate",foreground="#E67E22")
        for row in rows: tr.insert("","end",values=row,tags=(row[4],))
        self._make_sortable(tr,"tree_stats")
        tr.pack(fill="both",expand=True)

    # ═══════════════════════════════════════════════════════════
    # ABA ESTOQUE
    # ═══════════════════════════════════════════════════════════
    def setup_tab_estoque(self):
        tab = ttk.Frame(self.tabs); self.tabs.add(tab,text="  👕 ESTOQUE  ")
        top = tk.Frame(tab,bg=C_PRIMARIO,pady=8); top.pack(fill="x")
        tk.Label(top,text="👕  GESTÃO DE ESTOQUE DE UNIFORMES",bg=C_PRIMARIO,fg="white",
                 font=("Segoe UI",12,"bold")).pack(side="left",padx=16)
        tf3=tk.Frame(top,bg=C_PRIMARIO); tf3.pack(side="right",padx=16)
        tk.Label(tf3,text="Tipo:",bg=C_PRIMARIO,fg=C_ACENTO,font=("Segoe UI",10,"bold")).pack(side="left",padx=(0,6))
        self.est_tipo_var=tk.StringVar(value="Escolinha")
        for t in ["Escolinha","Selecao"]:
            tk.Radiobutton(tf3,text=t,variable=self.est_tipo_var,value=t,bg=C_PRIMARIO,fg="white",
                           selectcolor=C_SECUNDARIO,activebackground=C_PRIMARIO,activeforeground=C_ACENTO,
                           font=("Segoe UI",10,"bold"),command=self._on_tipo_change).pack(side="left",padx=6)

        body=tk.Frame(tab,bg=C_FUNDO); body.pack(fill="both",expand=True,padx=10,pady=8)
        left=tk.Frame(body,bg="white",bd=1,relief="solid",width=290)
        left.pack(side="left",fill="y",padx=(0,8)); left.pack_propagate(False)
        tk.Label(left,text="UNIFORMES CADASTRADOS",bg=C_PRIMARIO,fg="white",
                 font=("Segoe UI",10,"bold"),pady=7).pack(fill="x")

        # Busca estoque
        fb_est=tk.Frame(left,bg="#eef0f5",pady=3); fb_est.pack(fill="x",padx=4)
        tk.Label(fb_est,text="🔍",bg="#eef0f5",font=("Segoe UI",9)).pack(side="left")
        self.est_busca=tk.Entry(fb_est,font=("Segoe UI",9),bd=1,relief="solid",width=18)
        self.est_busca.pack(side="left",padx=2)
        self.est_busca.bind("<KeyRelease>",lambda e: self.refresh_estoque())
        make_btn(fb_est,"✕",lambda:(self.est_busca.delete(0,"end"),self.refresh_estoque()),
                 bg="#aaa",font_size=7,height=1,padx=4).pack(side="left",padx=2)

        make_btn(left,"➕  NOVO UNIFORME",self._novo_uniforme,bg=C_ACENTO,fg=C_PRIMARIO,
                 font_size=10,height=2).pack(fill="x",padx=6,pady=(6,4))
        bm=tk.Frame(left,bg="white"); bm.pack(fill="x",padx=6,pady=(0,6))
        make_btn(bm,"✏️ EDITAR",  self._editar_uniforme, bg=C_SECUNDARIO,font_size=9,height=1,padx=8
                 ).pack(side="left",expand=True,fill="x",padx=(0,3))
        make_btn(bm,"🗑️ EXCLUIR",self._excluir_uniforme,bg=C_PERIGO,    font_size=9,height=1,padx=8
                 ).pack(side="left",expand=True,fill="x")
        tk.Frame(left,bg="#dde0e8",height=1).pack(fill="x",padx=6,pady=(0,4))
        self.tree_unif=ttk.Treeview(left,columns=("id","nome","total"),show="headings",height=14)
        self.tree_unif.heading("id",text=""); self.tree_unif.heading("nome",text="Uniforme")
        self.tree_unif.heading("total",text="Total")
        self.tree_unif.column("id",width=0,stretch=False)
        self.tree_unif.column("nome",width=200,anchor="w"); self.tree_unif.column("total",width=55,anchor="center")
        sc_unif=ttk.Scrollbar(left,orient="vertical",command=self.tree_unif.yview)
        self.tree_unif.configure(yscrollcommand=sc_unif.set)
        self.tree_unif.pack(side="left",fill="both",expand=True,padx=(4,0),pady=(0,4))
        sc_unif.pack(side="right",fill="y",pady=(0,4),padx=(0,4))
        self.tree_unif.bind("<<TreeviewSelect>>",lambda e: self._carregar_variantes())
        self.tree_unif.tag_configure("ok",     foreground="#1a7a3c",background="#f0fff4")
        self.tree_unif.tag_configure("baixo",  foreground="#8B4000",background="#fff8e1")
        self.tree_unif.tag_configure("critico",foreground="#c0392b",background="#fff0f0")
        self.tree_unif.tag_configure("zero",   foreground="white",  background="#c0392b")

        right=tk.Frame(body,bg="white",bd=1,relief="solid"); right.pack(side="left",fill="both",expand=True)
        self.det_header=tk.Label(right,text="Selecione um uniforme para ver os tamanhos",
                                  bg=C_SECUNDARIO,fg="white",font=("Segoe UI",10,"bold"),pady=7)
        self.det_header.pack(fill="x")
        leg=tk.Frame(right,bg="#f9f9fb",pady=4); leg.pack(fill="x",padx=8)
        for cor,txt in [("#1a7a3c","OK"),("#E67E22","Baixo"),("#c0392b","Crítico"),("#888","Zerado")]:
            tk.Label(leg,text=f"  ●  {txt}",bg="#f9f9fb",fg=cor,font=("Segoe UI",8,"bold")).pack(side="left",padx=6)
        self.tree_var=ttk.Treeview(right,columns=("vid","Tamanho","Numero","Quantidade","Status"),
                                    show="headings",height=16)
        self.tree_var.heading("vid",text=""); self.tree_var.heading("Tamanho",text="Tamanho")
        self.tree_var.heading("Numero",text="Número"); self.tree_var.heading("Quantidade",text="Qtd")
        self.tree_var.heading("Status",text="Status")
        self.tree_var.column("vid",width=0,stretch=False)
        self.tree_var.column("Tamanho",width=120,anchor="center"); self.tree_var.column("Numero",width=100,anchor="center")
        self.tree_var.column("Quantidade",width=80,anchor="center"); self.tree_var.column("Status",width=140,anchor="center")
        self.tree_var.tag_configure("ok",     foreground="#1a7a3c",background="#f0fff4",font=("Segoe UI",10,"bold"))
        self.tree_var.tag_configure("baixo",  foreground="#8B4000",background="#fff8e1",font=("Segoe UI",10,"bold"))
        self.tree_var.tag_configure("critico",foreground="#c0392b",background="#fff5f5",font=("Segoe UI",10,"bold"))
        self.tree_var.tag_configure("zero",   foreground="white",  background="#c0392b",font=("Segoe UI",10,"bold"))
        sv=ttk.Scrollbar(right,orient="vertical",command=self.tree_var.yview)
        self.tree_var.configure(yscrollcommand=sv.set)
        self._make_sortable(self.tree_var,"tree_var")
        self.tree_var.pack(side="left",fill="both",expand=True,padx=(8,0),pady=6)
        sv.pack(side="right",fill="y",pady=6,padx=(0,4))

        bv=tk.Frame(right,bg="#eef0f5"); bv.pack(fill="x",side="bottom",pady=(0,6),padx=8)
        for i in range(5): bv.grid_columnconfigure(i,weight=1)
        btns_est=[
            ("➕ ADD TAM.",    self._nova_variante,    C_PRIMARIO),
            ("✏️ EDITAR TAM.", self._editar_variante,  C_SECUNDARIO),
            ("🛍️ VENDA",      self._vender_variante,  C_ACENTO),
            ("➕ ENTRADA",    self._entrada_variante, C_SUCESSO),
            ("🗑️ EXCLUIR TAM.",self._excluir_variante, C_PERIGO),
        ]
        for i,(txt,cmd,cor) in enumerate(btns_est):
            fg2 = C_PRIMARIO if cor==C_ACENTO else "white"
            make_btn(bv,txt,cmd,bg=cor,fg=fg2,font_size=8,height=1,padx=6
                     ).grid(row=0,column=i,padx=3,pady=4,sticky="ew")

        self.est_resumo=tk.Label(right,text="",bg="white",fg=C_SECUNDARIO,font=("Segoe UI",9),anchor="e")
        self.est_resumo.pack(fill="x",padx=10,pady=(0,4))
        self.refresh_estoque()

    def _on_tipo_change(self):
        self.det_header.config(text="Selecione um uniforme para ver os tamanhos")
        for i in self.tree_var.get_children(): self.tree_var.delete(i)
        self.est_resumo.config(text=""); self.refresh_estoque()

    def _status_cor(self,qty,minimo):
        if qty==0: return "zero","ZERADO"
        pct=qty/max(minimo,1)
        if pct>1.0:  return "ok",    f"OK  ({qty})"
        if pct>0.5:  return "baixo", f"BAIXO  ({qty})"
        return "critico",f"CRÍTICO  ({qty})"

    def refresh_estoque(self):
        for i in self.tree_unif.get_children(): self.tree_unif.delete(i)
        tipo=self.est_tipo_var.get()
        termo=self.est_busca.get().lower().strip() if hasattr(self,"est_busca") else ""
        try:
            conn=sqlite3.connect(db_path); cur=conn.cursor()
            cur.execute("""SELECT e.id,e.nome,e.estoque_minimo,COALESCE(SUM(v.quantidade),0)
                           FROM estoque e LEFT JOIN estoque_variantes v ON v.estoque_id=e.id
                           WHERE e.tipo=? GROUP BY e.id ORDER BY e.nome ASC""",(tipo,))
            for uid,nome,minimo,total in cur.fetchall():
                if termo and termo not in nome.lower(): continue
                tag,_=self._status_cor(total,minimo)
                self.tree_unif.insert("","end",iid=str(uid),values=(uid,nome,total),tags=(tag,))
            conn.close()
        except Exception as err: messagebox.showerror("Erro no Estoque",f"{err}")

    def _carregar_variantes(self):
        sel=self.tree_unif.selection()
        if not sel: return
        uid=int(sel[0])
        conn=sqlite3.connect(db_path); cur=conn.cursor()
        cur.execute("SELECT nome,estoque_minimo FROM estoque WHERE id=?",(uid,))
        row=cur.fetchone()
        if not row: conn.close(); return
        nome_unif,minimo=row
        self.det_header.config(text=f"  👕  {nome_unif}   |   Estoque mínimo: {minimo}")
        tipo=self.est_tipo_var.get()
        cur.execute("SELECT id,tamanho,numero,quantidade FROM estoque_variantes WHERE estoque_id=? ORDER BY tamanho,numero",(uid,))
        variantes=cur.fetchall(); conn.close()
        total_qtd=sum(v[3] for v in variantes); zeros=sum(1 for v in variantes if v[3]==0)
        criticos=sum(1 for v in variantes if 0<v[3]<=max(minimo//2,1))
        self.est_resumo.config(text=f"Total: {total_qtd}  |  Tamanhos: {len(variantes)}  |  Zerados: {zeros}  |  Críticos: {criticos}")
        if tipo=="Selecao": self.tree_var.column("Numero",width=100); self.tree_var.heading("Numero",text="Número")
        else: self.tree_var.column("Numero",width=0,stretch=False); self.tree_var.heading("Numero",text="")
        for i in self.tree_var.get_children(): self.tree_var.delete(i)
        for vid,tam,num,qty in variantes:
            tag,status=self._status_cor(qty,minimo)
            self.tree_var.insert("","end",iid=str(vid),values=(vid,tam,num,qty,status),tags=(tag,))

    def _get_sel_unif(self):
        sel=self.tree_unif.selection()
        if not sel: messagebox.showwarning("Aviso","Selecione um uniforme."); return None
        return int(sel[0])

    def _get_sel_var(self):
        sel=self.tree_var.selection()
        if not sel: messagebox.showwarning("Aviso","Selecione um tamanho."); return None
        return int(sel[0])

    def _novo_uniforme(self):
        tipo=self.est_tipo_var.get()
        w=tk.Toplevel(self.root); w.title("Novo Uniforme"); w.geometry("400x230")
        w.configure(bg="white"); w.transient(self.root); w.grab_set()
        tk.Label(w,text=f"NOVO UNIFORME — {tipo}",bg=C_PRIMARIO,fg="white",
                 font=("Segoe UI",12,"bold"),pady=10).pack(fill="x")
        fm=tk.Frame(w,bg="white",padx=30,pady=10); fm.pack(fill="both")
        tk.Label(fm,text="Nome do Uniforme:",bg="white",font=("Segoe UI",10)).grid(row=0,column=0,sticky="w",pady=6)
        e_nome=tk.Entry(fm,font=("Segoe UI",11),bd=1,relief="solid",width=24); e_nome.grid(row=0,column=1,padx=10,pady=6,sticky="ew")
        tk.Label(fm,text="Estoque Mínimo:",bg="white",font=("Segoe UI",10)).grid(row=1,column=0,sticky="w",pady=6)
        e_min=tk.Entry(fm,font=("Segoe UI",11),bd=1,relief="solid",width=8); e_min.insert(0,"5")
        e_min.grid(row=1,column=1,padx=10,pady=6,sticky="w"); fm.grid_columnconfigure(1,weight=1)
        def salvar(ev=None):
            nome=e_nome.get().strip()
            if not nome: messagebox.showwarning("Aviso","Informe o nome.",parent=w); return
            try: minimo=int(e_min.get().strip())
            except: minimo=5
            try:
                conn=sqlite3.connect(db_path); c=conn.cursor()
                c.execute("INSERT INTO estoque (tipo,nome,estoque_minimo) VALUES (?,?,?)",(tipo,nome,minimo))
                conn.commit(); conn.close(); w.destroy(); self.refresh_estoque()
            except Exception as err: messagebox.showerror("Erro",f"{err}",parent=w)
        e_nome.bind("<Return>",lambda e: e_min.focus()); e_min.bind("<Return>",salvar)
        make_btn(w,"SALVAR",salvar,bg=C_PRIMARIO,font_size=11,height=1).pack(fill="x",padx=30,pady=10)
        e_nome.focus()

    def _editar_uniforme(self):
        uid=self._get_sel_unif()
        if uid is None: return
        conn=sqlite3.connect(db_path); cur=conn.cursor()
        cur.execute("SELECT nome,estoque_minimo FROM estoque WHERE id=?",(uid,))
        row=cur.fetchone(); conn.close()
        if not row: return
        w=tk.Toplevel(self.root); w.title("Editar Uniforme"); w.geometry("400x220")
        w.configure(bg="white"); w.transient(self.root); w.grab_set()
        tk.Label(w,text="EDITAR UNIFORME",bg=C_PRIMARIO,fg="white",
                 font=("Segoe UI",12,"bold"),pady=10).pack(fill="x")
        fm=tk.Frame(w,bg="white",padx=30,pady=10); fm.pack(fill="both")
        tk.Label(fm,text="Nome:",bg="white",font=("Segoe UI",10)).grid(row=0,column=0,sticky="w",pady=6)
        e_nome=tk.Entry(fm,font=("Segoe UI",11),bd=1,relief="solid",width=24); e_nome.insert(0,row[0])
        e_nome.grid(row=0,column=1,padx=10,pady=6,sticky="ew")
        tk.Label(fm,text="Estoque Mínimo:",bg="white",font=("Segoe UI",10)).grid(row=1,column=0,sticky="w",pady=6)
        e_min=tk.Entry(fm,font=("Segoe UI",11),bd=1,relief="solid",width=8); e_min.insert(0,str(row[1]))
        e_min.grid(row=1,column=1,padx=10,pady=6,sticky="w"); fm.grid_columnconfigure(1,weight=1)
        def salvar(ev=None):
            nome=e_nome.get().strip()
            if not nome: return
            try: minimo=int(e_min.get().strip())
            except: minimo=5
            conn2=sqlite3.connect(db_path); c2=conn2.cursor()
            c2.execute("UPDATE estoque SET nome=?,estoque_minimo=? WHERE id=?",(nome,minimo,uid))
            conn2.commit(); conn2.close(); w.destroy(); self.refresh_estoque(); self._carregar_variantes()
        e_nome.bind("<Return>",lambda e: e_min.focus()); e_min.bind("<Return>",salvar)
        make_btn(w,"SALVAR",salvar,bg=C_PRIMARIO,font_size=11,height=1).pack(fill="x",padx=30,pady=10)
        e_nome.focus(); e_nome.select_range(0,"end")

    def _excluir_uniforme(self):
        uid=self._get_sel_unif()
        if uid is None: return
        conn=sqlite3.connect(db_path); cur=conn.cursor()
        cur.execute("SELECT nome FROM estoque WHERE id=?",(uid,)); row=cur.fetchone(); conn.close()
        if not row: return
        if messagebox.askyesno("Confirmar",f"Excluir '{row[0]}' e TODOS os tamanhos?"):
            conn2=sqlite3.connect(db_path); c2=conn2.cursor()
            c2.execute("DELETE FROM estoque_variantes WHERE estoque_id=?",(uid,))
            c2.execute("DELETE FROM estoque WHERE id=?",(uid,))
            conn2.commit(); conn2.close()
            for i in self.tree_var.get_children(): self.tree_var.delete(i)
            self.det_header.config(text="Selecione um uniforme para ver os tamanhos")
            self.est_resumo.config(text=""); self.refresh_estoque()

    def _nova_variante(self):
        uid=self._get_sel_unif()
        if uid is None: return
        tipo=self.est_tipo_var.get()
        w=tk.Toplevel(self.root); w.title("Adicionar Tamanho")
        w.geometry("380x270" if tipo=="Selecao" else "380x215")
        w.configure(bg="white"); w.transient(self.root); w.grab_set()
        tk.Label(w,text="NOVO TAMANHO / VARIANTE",bg=C_PRIMARIO,fg="white",
                 font=("Segoe UI",12,"bold"),pady=10).pack(fill="x")
        fm=tk.Frame(w,bg="white",padx=30,pady=10); fm.pack(fill="both")
        tk.Label(fm,text="Tamanho:",bg="white",font=("Segoe UI",10)).grid(row=0,column=0,columnspan=2,sticky="w")
        e_tam=tk.Entry(fm,font=("Segoe UI",12),bd=1,relief="solid",width=12); e_tam.grid(row=1,column=0,pady=(2,10),sticky="w")
        e_num=None
        if tipo=="Selecao":
            tk.Label(fm,text="Número da Camisa:",bg="white",font=("Segoe UI",10)).grid(row=2,column=0,columnspan=2,sticky="w")
            e_num=tk.Entry(fm,font=("Segoe UI",12),bd=1,relief="solid",width=8); e_num.grid(row=3,column=0,pady=(2,10),sticky="w")
        r_qty=4 if tipo=="Selecao" else 2
        tk.Label(fm,text="Quantidade inicial:",bg="white",font=("Segoe UI",10)).grid(row=r_qty,column=0,columnspan=2,sticky="w")
        e_qty=tk.Entry(fm,font=("Segoe UI",14,"bold"),bd=1,relief="solid",width=8,justify="center")
        e_qty.insert(0,"0"); e_qty.grid(row=r_qty+1,column=0,pady=(2,6),sticky="w")
        def salvar(ev=None):
            tam=e_tam.get().strip().upper()
            if not tam: messagebox.showwarning("Aviso","Informe o tamanho.",parent=w); return
            try: qty=int(e_qty.get().strip())
            except: qty=0
            num=e_num.get().strip() if e_num else ""
            conn=sqlite3.connect(db_path); c=conn.cursor()
            c.execute("INSERT INTO estoque_variantes (estoque_id,tamanho,numero,quantidade) VALUES (?,?,?,?)",(uid,tam,num,qty))
            conn.commit(); conn.close(); w.destroy(); self.refresh_estoque(); self._carregar_variantes()
        e_tam.bind("<Return>",lambda e: (e_num.focus() if e_num else e_qty.focus()))
        if e_num: e_num.bind("<Return>",lambda e: e_qty.focus())
        e_qty.bind("<Return>",salvar)
        make_btn(w,"SALVAR TAMANHO",salvar,bg=C_PRIMARIO,font_size=11,height=1).pack(fill="x",padx=30,pady=8)
        e_tam.focus()

    def _editar_variante(self):
        vid=self._get_sel_var()
        if vid is None: return
        tipo=self.est_tipo_var.get()
        conn=sqlite3.connect(db_path); cur=conn.cursor()
        cur.execute("SELECT tamanho,numero,quantidade FROM estoque_variantes WHERE id=?",(vid,))
        row=cur.fetchone(); conn.close()
        if not row: return
        w=tk.Toplevel(self.root); w.title("Editar Tamanho")
        w.geometry("380x270" if tipo=="Selecao" else "380x215")
        w.configure(bg="white"); w.transient(self.root); w.grab_set()
        tk.Label(w,text="EDITAR TAMANHO",bg=C_PRIMARIO,fg="white",
                 font=("Segoe UI",12,"bold"),pady=10).pack(fill="x")
        fm=tk.Frame(w,bg="white",padx=30,pady=10); fm.pack(fill="both")
        tk.Label(fm,text="Tamanho:",bg="white",font=("Segoe UI",10)).grid(row=0,column=0,columnspan=2,sticky="w")
        e_tam=tk.Entry(fm,font=("Segoe UI",12),bd=1,relief="solid",width=12); e_tam.insert(0,row[0])
        e_tam.grid(row=1,column=0,pady=(2,10),sticky="w")
        e_num=None
        if tipo=="Selecao":
            tk.Label(fm,text="Número:",bg="white",font=("Segoe UI",10)).grid(row=2,column=0,columnspan=2,sticky="w")
            e_num=tk.Entry(fm,font=("Segoe UI",12),bd=1,relief="solid",width=8); e_num.insert(0,row[1])
            e_num.grid(row=3,column=0,pady=(2,10),sticky="w")
        r_qty=4 if tipo=="Selecao" else 2
        tk.Label(fm,text="Quantidade:",bg="white",font=("Segoe UI",10)).grid(row=r_qty,column=0,columnspan=2,sticky="w")
        e_qty=tk.Entry(fm,font=("Segoe UI",14,"bold"),bd=1,relief="solid",width=8,justify="center")
        e_qty.insert(0,str(row[2])); e_qty.grid(row=r_qty+1,column=0,pady=(2,6),sticky="w")
        def salvar(ev=None):
            tam=e_tam.get().strip().upper()
            if not tam: return
            try: qty=int(e_qty.get().strip())
            except: qty=0
            num=e_num.get().strip() if e_num else ""
            conn2=sqlite3.connect(db_path); c2=conn2.cursor()
            c2.execute("UPDATE estoque_variantes SET tamanho=?,numero=?,quantidade=? WHERE id=?",(tam,num,qty,vid))
            conn2.commit(); conn2.close(); w.destroy(); self.refresh_estoque(); self._carregar_variantes()
        e_tam.bind("<Return>",lambda e: (e_num.focus() if e_num else e_qty.focus()))
        if e_num: e_num.bind("<Return>",lambda e: e_qty.focus())
        e_qty.bind("<Return>",salvar)
        make_btn(w,"SALVAR",salvar,bg=C_PRIMARIO,font_size=11,height=1).pack(fill="x",padx=30,pady=8)
        e_tam.focus(); e_tam.select_range(0,"end")

    def _ajustar_variante(self,label,delta_fn):
        uid=self._get_sel_unif(); vid=self._get_sel_var()
        if uid is None or vid is None: return
        conn=sqlite3.connect(db_path); cur=conn.cursor()
        cur.execute("SELECT tamanho,numero,quantidade FROM estoque_variantes WHERE id=?",(vid,)); vrow=cur.fetchone()
        cur.execute("SELECT nome FROM estoque WHERE id=?",(uid,)); nrow=cur.fetchone(); conn.close()
        if not vrow or not nrow: return
        desc=f"{nrow[0]} — {vrow[0]}"+(f"  #{vrow[1]}" if vrow[1] else "")
        w=tk.Toplevel(self.root); w.title(label); w.geometry("340x200")
        w.configure(bg="white"); w.transient(self.root); w.grab_set()
        tk.Label(w,text=label,bg=C_PRIMARIO,fg="white",font=("Segoe UI",12,"bold"),pady=10).pack(fill="x")
        tk.Label(w,text=desc,bg="white",fg=C_SECUNDARIO,font=("Segoe UI",10,"bold")).pack(pady=(10,2))
        tk.Label(w,text=f"Estoque atual: {vrow[2]}",bg="white",fg="#555",font=("Segoe UI",9)).pack()
        rf=tk.Frame(w,bg="white"); rf.pack(pady=12)
        tk.Label(rf,text="Quantidade:",bg="white",font=("Segoe UI",10)).pack(side="left",padx=8)
        e_qty=tk.Entry(rf,font=("Segoe UI",16,"bold"),width=6,justify="center",bd=1,relief="solid")
        e_qty.insert(0,"1"); e_qty.pack(side="left")
        def confirmar(ev=None):
            try: qtd=int(e_qty.get().strip()); assert qtd>0
            except: messagebox.showwarning("Aviso","Quantidade inválida.",parent=w); return
            nova=delta_fn(vrow[2],qtd)
            if nova<0: messagebox.showwarning("Aviso",f"Estoque insuficiente! Disponível: {vrow[2]}",parent=w); return
            conn2=sqlite3.connect(db_path); c2=conn2.cursor()
            c2.execute("UPDATE estoque_variantes SET quantidade=? WHERE id=?",(nova,vid))
            conn2.commit(); conn2.close(); w.destroy(); self.refresh_estoque(); self._carregar_variantes()
        e_qty.bind("<Return>",confirmar)
        make_btn(w,"CONFIRMAR",confirmar,bg=C_PRIMARIO,font_size=11,height=1).pack(fill="x",padx=40,pady=4)
        e_qty.focus(); e_qty.select_range(0,"end")

    def _vender_variante(self):  self._ajustar_variante("🛍️  Registrar Venda",    lambda a,q: a-q)
    def _entrada_variante(self): self._ajustar_variante("➕  Entrada de Estoque", lambda a,q: a+q)

    def _excluir_variante(self):
        vid=self._get_sel_var()
        if vid is None: return
        conn=sqlite3.connect(db_path); cur=conn.cursor()
        cur.execute("SELECT tamanho,numero FROM estoque_variantes WHERE id=?",(vid,))
        row=cur.fetchone(); conn.close()
        if not row: return
        desc=row[0]+(f"  #{row[1]}" if row[1] else "")
        if messagebox.askyesno("Confirmar",f"Excluir variante '{desc}'?"):
            conn2=sqlite3.connect(db_path); c2=conn2.cursor()
            c2.execute("DELETE FROM estoque_variantes WHERE id=?",(vid,))
            conn2.commit(); conn2.close(); self.refresh_estoque(); self._carregar_variantes()

  # ═══════════════════════════════════════════════════════════
    # ABA ATIVIDADES — FORMULÁRIO NO TOPO (LAYOUT CORRIGIDO)
    # ═══════════════════════════════════════════════════════════
    def setup_tab_atividades(self):
        tab = ttk.Frame(self.tabs); self.tabs.add(tab, text="  🏃 ATIVIDADES  ")

        # ── Cabeçalho ─────────────────────────────────────────
        hdr = tk.Frame(tab, bg=C_PRIMARIO, pady=8); hdr.pack(fill="x")
        tk.Label(hdr, text="🏃  GRADE DE ATIVIDADES ESPORTIVAS", bg=C_PRIMARIO, fg="white",
                 font=("Segoe UI", 12, "bold")).pack(side="left", padx=16)
        tk.Label(hdr, text="Escolinhas · Seleções · Educação Física",
                 bg=C_PRIMARIO, fg=C_ACENTO, font=("Segoe UI", 9)).pack(side="left", padx=4)
        make_btn(hdr, "⚙️ GERENCIAR LISTAS", self._gerenciar_listas_atividades,
                 bg=C_ACENTO, fg=C_PRIMARIO, font_size=9, height=1, padx=14).pack(side="right", padx=14)

        # ── Formulário de nova atividade (grade horizontal no topo) ──
        sec = tk.LabelFrame(tab, text=" ➕ Nova Atividade ", bg=C_FUNDO,
                            font=("Segoe UI", 9, "bold"), fg=C_PRIMARIO)
        sec.pack(fill="x", padx=10, pady=(6, 2))
        for col in range(6):
            sec.grid_columnconfigure(col, weight=1)

        def lbl_sec(col, row, txt):
            tk.Label(sec, text=txt, bg=C_FUNDO, font=("Segoe UI", 8, "bold"), fg=C_PRIMARIO
                     ).grid(row=row, column=col, sticky="w", padx=(8, 2), pady=(6, 0))

        def make_combo_btn(tabela, attr_var, attr_cb, col, row):
            """Combobox editável com botão ⚙️ para gerenciar a lista."""
            fr = tk.Frame(sec, bg=C_FUNDO)
            fr.grid(row=row, column=col, sticky="ew", padx=(6, 8), pady=(2, 6))
            var = tk.StringVar()
            cb = ttk.Combobox(fr, textvariable=var, values=db_lista(tabela), font=("Segoe UI", 9))
            cb.pack(side="left", fill="x", expand=True)
            make_btn(fr, "⚙", lambda t=tabela, c=cb: self._abrir_lista_e_atualizar(t, [c]),
                     bg=C_SECUNDARIO, font_size=7, height=1, padx=3).pack(side="left", padx=(2, 0))
            setattr(self, attr_var, var)
            setattr(self, attr_cb, cb)

        # ── Linha 0 — Labels ──────────────────────────────────
        lbl_sec(0, 0, "Tipo:")
        lbl_sec(1, 0, "Nome / Turma:")
        lbl_sec(2, 0, "Dia da Semana:")
        lbl_sec(3, 0, "Início:")
        lbl_sec(4, 0, "Fim:")
        lbl_sec(5, 0, "Local / Quadra:")

        # ── Linha 1 — Widgets Linha 1 ─────────────────────────
        self.atv_tipo = ttk.Combobox(sec,
            values=["Escolinha", "Seleção", "Educação Física", "Outro"],
            font=("Segoe UI", 9), state="readonly")
        self.atv_tipo.current(0)
        self.atv_tipo.grid(row=1, column=0, sticky="ew", padx=(6, 8), pady=(2, 6))

        make_combo_btn("atv_nomes",   "atv_nome_var",   "atv_nome_var_cb",   1, 1)

        self.atv_dia = ttk.Combobox(sec,
            values=["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira",
                    "Sexta-feira", "Sábado", "Domingo",
                    "Seg e Qua", "Ter e Qui", "Seg, Qua e Sex"],
            font=("Segoe UI", 9), state="readonly")
        self.atv_dia.current(0)
        self.atv_dia.grid(row=1, column=2, sticky="ew", padx=(6, 8), pady=(2, 6))

        # Início com botão gerenciar horários
        fr_ini = tk.Frame(sec, bg=C_FUNDO)
        fr_ini.grid(row=1, column=3, sticky="ew", padx=(6, 8), pady=(2, 6))
        self.atv_inicio_var = tk.StringVar()
        self.atv_inicio_cb = ttk.Combobox(fr_ini, textvariable=self.atv_inicio_var,
                                           values=db_lista("atv_horarios"), font=("Segoe UI", 9))
        self.atv_inicio_cb.pack(side="left", fill="x", expand=True)
        make_btn(fr_ini, "⚙", lambda: self._abrir_lista_e_atualizar(
                 "atv_horarios", [self.atv_inicio_cb, self.atv_fim_cb]),
                 bg=C_SECUNDARIO, font_size=7, height=1, padx=3).pack(side="left", padx=(2, 0))

        # Fim com botão gerenciar horários
        fr_fim = tk.Frame(sec, bg=C_FUNDO)
        fr_fim.grid(row=1, column=4, sticky="ew", padx=(6, 8), pady=(2, 6))
        self.atv_fim_var = tk.StringVar()
        self.atv_fim_cb = ttk.Combobox(fr_fim, textvariable=self.atv_fim_var,
                                        values=db_lista("atv_horarios"), font=("Segoe UI", 9))
        self.atv_fim_cb.pack(side="left", fill="x", expand=True)
        make_btn(fr_fim, "⚙", lambda: self._abrir_lista_e_atualizar(
                 "atv_horarios", [self.atv_inicio_cb, self.atv_fim_cb]),
                 bg=C_SECUNDARIO, font_size=7, height=1, padx=3).pack(side="left", padx=(2, 0))

        make_combo_btn("atv_quadras", "atv_local_var",  "atv_local_var_cb",  5, 1)

        # ── Linha 2 — Labels ──────────────────────────────────
        lbl_sec(0, 2, "Professor:")
        lbl_sec(1, 2, "Estagiário (opcional):")
        lbl_sec(2, 2, "Observações:")

        # ── Linha 3 — Widgets Linha 2 ─────────────────────────
        make_combo_btn("atv_profs",   "atv_prof_var",   "atv_prof_var_cb",   0, 3)
        make_combo_btn("atv_estags",  "atv_estag_var",  "atv_estag_var_cb",  1, 3)

        self.atv_obs = tk.Entry(sec, font=("Segoe UI", 9), bd=1, relief="solid")
        self.atv_obs.grid(row=3, column=2, columnspan=3, sticky="ew", padx=(6, 8), pady=(2, 6))

        make_btn(sec, "  ➕ ADICIONAR  ", self._add_atividade,
                 bg=C_PRIMARIO, font_size=9, height=1
                 ).grid(row=3, column=5, sticky="ew", padx=(6, 8), pady=(2, 6))

        # ── Tabela de atividades (largura total) ──────────────
        tbl_outer = tk.Frame(tab, bg=C_FUNDO); tbl_outer.pack(fill="both", expand=True, padx=10, pady=(4, 8))
        tbl_frame = tk.Frame(tbl_outer, bg="white", bd=1, relief="solid")
        tbl_frame.pack(fill="both", expand=True)

        tk.Label(tbl_frame, text="ATIVIDADES CADASTRADAS", bg=C_PRIMARIO, fg="white",
                 font=("Segoe UI", 10, "bold"), pady=6).pack(fill="x")

        # Barra de filtro
        flt = tk.Frame(tbl_frame, bg="#eef0f5", pady=5); flt.pack(fill="x", padx=6)
        tk.Label(flt, text="Tipo:", bg="#eef0f5", font=("Segoe UI", 9, "bold"), fg=C_PRIMARIO).pack(side="left", padx=(0, 4))
        self.atv_filtro_var = tk.StringVar(value="Todos")
        self.atv_filtro_cb  = ttk.Combobox(flt, textvariable=self.atv_filtro_var,
                                            values=["Todos", "Escolinha", "Seleção", "Educação Física", "Outro"],
                                            width=16, state="readonly")
        self.atv_filtro_cb.pack(side="left", padx=(0, 10))
        self.atv_filtro_cb.bind("<<ComboboxSelected>>", lambda e: self.refresh_atividades())
        tk.Label(flt, text="🔍", bg="#eef0f5", font=("Segoe UI", 10)).pack(side="left", padx=(0, 2))
        self.atv_busca = tk.Entry(flt, font=("Segoe UI", 10), bd=1, relief="solid", width=24)
        self.atv_busca.pack(side="left"); self.atv_busca.bind("<KeyRelease>", lambda e: self.refresh_atividades())
        make_btn(flt, "✕", lambda: (self.atv_busca.delete(0, "end"), self.refresh_atividades()),
                 bg="#aaa", font_size=8, height=1, padx=6).pack(side="left", padx=4)
        tk.Label(flt, text="Coluna:", bg="#eef0f5", font=("Segoe UI", 9)).pack(side="left", padx=(8, 2))
        self.atv_busca_col = ttk.Combobox(flt,
            values=["Nome/Turma", "Professor", "Estagiário", "Local", "Dia"],
            width=13, state="readonly"); self.atv_busca_col.current(0)
        self.atv_busca_col.pack(side="left")
        self.atv_busca_col.bind("<<ComboboxSelected>>", lambda e: self.refresh_atividades())
        make_btn(flt, "🔄", self.refresh_atividades, bg=C_ACENTO, fg=C_PRIMARIO,
                 font_size=8, height=1, padx=8).pack(side="left", padx=6)

        # Botões de ação (empacotados ANTES da tabela para ficarem visíveis)
        btn_bar = tk.Frame(tbl_frame, bg="#eef0f5"); btn_bar.pack(fill="x", padx=6, pady=(0, 2))
        for i in range(3): btn_bar.grid_columnconfigure(i, weight=1)
        make_btn(btn_bar, "✏️  EDITAR ATIVIDADE",   self._editar_atividade,
                 bg=C_SECUNDARIO, font_size=9, height=1, padx=14
                 ).grid(row=0, column=0, padx=4, pady=4, sticky="ew")
        make_btn(btn_bar, "🗑️  EXCLUIR ATIVIDADE",  self._excluir_atividade,
                 bg=C_PERIGO, font_size=9, height=1, padx=14
                 ).grid(row=0, column=1, padx=4, pady=4, sticky="ew")
        make_btn(btn_bar, "📊  EXPORTAR EXCEL",     self._gerar_excel_atividades,
                 bg=C_SUCESSO, font_size=9, height=1, padx=14
                 ).grid(row=0, column=2, padx=4, pady=4, sticky="ew")

        # Legenda de cores
        leg_bar = tk.Frame(tbl_frame, bg="#f9f9fb", pady=3); leg_bar.pack(fill="x", padx=8)
        for fg_cor, bg_cor, txt in [
            ("#1a5f7a", "#e8f4f8", "🏫 Escolinha"),
            ("#1a7a3c", "#e8f5e9", "🏆 Seleção"),
            ("#6a1a7a", "#f3e8f8", "📚 Ed. Física"),
            ("#7a4a1a", "#fdf3e3", "🔹 Outro"),
        ]:
            tk.Label(leg_bar, text=f"  {txt}  ", bg=bg_cor, fg=fg_cor,
                     font=("Segoe UI", 8, "bold"), padx=6, pady=2).pack(side="left", padx=4)

        # Treeview ocupa o espaço restante
        cols = ("ID", "Tipo", "Nome / Turma", "Dia", "Início", "Fim", "Local", "Professor", "Estagiário")
        self.tree_atv = ttk.Treeview(tbl_frame, columns=cols, show="headings", height=13)
        ws_a = {"ID": 0, "Tipo": 95, "Nome / Turma": 180, "Dia": 110, "Início": 68, "Fim": 68,
                "Local": 150, "Professor": 170, "Estagiário": 150}
        for col in cols:
            self.tree_atv.heading(col, text=col)
            self.tree_atv.column(col, width=ws_a[col],
                                  anchor="w" if col in ("Nome / Turma", "Local", "Professor", "Estagiário") else "center",
                                  stretch=(col != "ID"))
        self.tree_atv.column("ID", width=0, stretch=False)
        self.tree_atv.tag_configure("Escolinha",       foreground="#1a5f7a", background="#e8f4f8")
        self.tree_atv.tag_configure("Seleção",         foreground="#1a7a3c", background="#e8f5e9")
        self.tree_atv.tag_configure("Educação Física", foreground="#6a1a7a", background="#f3e8f8")
        self.tree_atv.tag_configure("Outro",           foreground="#7a4a1a", background="#fdf3e3")
        sc_atv = ttk.Scrollbar(tbl_frame, orient="vertical", command=self.tree_atv.yview)
        self.tree_atv.configure(yscrollcommand=sc_atv.set)
        self._make_sortable(self.tree_atv, "tree_atv")
        self.tree_atv.bind("<Double-1>", lambda e: self._editar_atividade())
        self.tree_atv.pack(side="left", fill="both", expand=True, padx=(4, 0), pady=(0, 4))
        sc_atv.pack(side="right", fill="y", pady=(0, 4), padx=(0, 4))

        self.refresh_atividades()

    def _abrir_lista_e_atualizar(self, tabela, combos_para_atualizar):
        nomes_tabelas = {
            "atv_nomes":    "Nomes de Atividades / Turmas",
            "atv_profs":    "Professores",
            "atv_estags":   "Estagiários",
            "atv_quadras":  "Locais / Quadras",
            "atv_horarios": "Horários",
        }
        titulo = nomes_tabelas.get(tabela, tabela)
        w = janela_gerenciar_lista(self.root, tabela, titulo)
        def ao_fechar():
            novos = db_lista(tabela)
            for cb in combos_para_atualizar:
                if cb and cb.winfo_exists():
                    cb["values"] = novos
            w.destroy()
        w.protocol("WM_DELETE_WINDOW", ao_fechar)

    def _gerenciar_listas_atividades(self):
        w = tk.Toplevel(self.root); w.title("Gerenciar Listas de Atividades")
        w.geometry("760x520"); w.configure(bg=C_FUNDO)
        w.transient(self.root); w.grab_set()
        tk.Label(w, text="⚙️  LISTAS DE ATIVIDADES", bg=C_FUNDO, fg=C_PRIMARIO,
                 font=("Segoe UI", 14, "bold")).pack(pady=(14, 4))
        tk.Label(w, text="Gerencie os menus de seleção que aparecem no formulário de Atividades.",
                 bg=C_FUNDO, fg=C_SECUNDARIO, font=("Segoe UI", 9)).pack(pady=(0, 8))
        nb = ttk.Notebook(w); nb.pack(fill="both", expand=True, padx=14, pady=(0, 12))
        tabelas = [("atv_nomes", "Nomes / Turmas"), ("atv_profs", "Professores"),
                   ("atv_estags", "Estagiários"), ("atv_quadras", "Locais / Quadras"), ("atv_horarios", "Horários")]
        for tabela, titulo in tabelas:
            fr = ttk.Frame(nb); nb.add(fr, text=f"  {titulo}  ")
            top = tk.Frame(fr, bg="white"); top.pack(fill="x", padx=8, pady=8)
            e = tk.Entry(top, font=("Segoe UI", 11), bd=1, relief="solid", width=26); e.pack(side="left", padx=(0, 8))
            tree = ttk.Treeview(fr, columns=("nome",), show="headings", height=13)
            tree.heading("nome", text=titulo); tree.column("nome", width=340, anchor="w")
            sc = ttk.Scrollbar(fr, orient="vertical", command=tree.yview); tree.configure(yscrollcommand=sc.set)

            def carregar(t=tabela, tr=tree):
                for i in tr.get_children(): tr.delete(i)
                for n in db_lista(t): tr.insert("", "end", values=(n,))

            def adicionar(ev=None, t=tabela, entry=e, tr=tree):
                nome = entry.get().strip()
                if not nome: return
                conn = sqlite3.connect(db_path); cur = conn.cursor()
                try:
                    cur.execute(f"INSERT INTO {t} (nome) VALUES (?)", (nome,))
                    conn.commit(); entry.delete(0, "end"); carregar(t, tr)
                except sqlite3.IntegrityError: messagebox.showwarning("Aviso", "Já existe!", parent=w)
                finally: conn.close()

            def excluir(t=tabela, tr=tree):
                sel = tr.selection()
                if not sel: return
                nome = tr.item(sel[0])["values"][0]
                if messagebox.askyesno("Confirmar", f"Excluir '{nome}'?", parent=w):
                    conn = sqlite3.connect(db_path); cur = conn.cursor()
                    cur.execute(f"DELETE FROM {t} WHERE nome=?", (nome,))
                    conn.commit(); conn.close(); carregar(t, tr)

            make_btn(top, "+ ADICIONAR", lambda ev=None, a=adicionar: a(ev),
                     bg=C_PRIMARIO, font_size=9, height=1, padx=12).pack(side="left", padx=2)
            make_btn(top, "🗑️ EXCLUIR", lambda t=tabela, tr=tree: excluir(t, tr),
                     bg=C_PERIGO, font_size=9, height=1, padx=12).pack(side="left", padx=2)
            e.bind("<Return>", adicionar)
            tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=(0, 8))
            sc.pack(side="right", fill="y", pady=(0, 8), padx=(0, 8))
            carregar()

    def _add_atividade(self):
        nome   = self.atv_nome_var.get().strip()
        prof   = self.atv_prof_var.get().strip()
        inicio = self.atv_inicio_var.get().strip()
        if not nome:   messagebox.showwarning("Aviso", "Informe o nome / turma."); return
        if not prof:   messagebox.showwarning("Aviso", "Informe o professor."); return
        if not inicio: messagebox.showwarning("Aviso", "Informe o horário de início."); return
        conn = sqlite3.connect(db_path); c = conn.cursor()
        c.execute("""INSERT INTO atividades
                     (tipo_atividade,nome_atividade,dia_semana,horario_inicio,horario_fim,
                      local,professor,estagiario,observacoes) VALUES (?,?,?,?,?,?,?,?,?)""",
                  (self.atv_tipo.get(), nome, self.atv_dia.get(), inicio,
                   self.atv_fim_var.get().strip(), self.atv_local_var.get().strip(),
                   prof, self.atv_estag_var.get().strip(),
                   self.atv_obs.get().strip()))   # Entry em vez de Text
        conn.commit(); conn.close()
        for attr in ["atv_nome_var", "atv_inicio_var", "atv_fim_var",
                     "atv_local_var", "atv_prof_var", "atv_estag_var"]:
            getattr(self, attr).set("")
        self.atv_obs.delete(0, "end")            # Entry em vez de Text
        self.refresh_atividades()

    def refresh_atividades(self):
        if not hasattr(self, "tree_atv"): return
        for i in self.tree_atv.get_children(): self.tree_atv.delete(i)
        filtro = self.atv_filtro_var.get() if hasattr(self, "atv_filtro_var") else "Todos"
        termo  = self.atv_busca.get().lower().strip() if hasattr(self, "atv_busca") else ""
        col_map = {"Nome/Turma": 2, "Professor": 7, "Estagiário": 8, "Local": 6, "Dia": 3}
        ci = col_map.get(self.atv_busca_col.get() if hasattr(self, "atv_busca_col") else "Nome/Turma", 2)
        conn = sqlite3.connect(db_path); cur = conn.cursor()
        if filtro == "Todos":
            cur.execute("""SELECT id,tipo_atividade,nome_atividade,dia_semana,horario_inicio,
                                  horario_fim,local,professor,estagiario
                           FROM atividades ORDER BY tipo_atividade,dia_semana,horario_inicio""")
        else:
            cur.execute("""SELECT id,tipo_atividade,nome_atividade,dia_semana,horario_inicio,
                                  horario_fim,local,professor,estagiario
                           FROM atividades WHERE tipo_atividade=?
                           ORDER BY dia_semana,horario_inicio""", (filtro,))
        rows = cur.fetchall(); conn.close()
        for r in rows:
            if termo and termo not in str(r[ci]).lower(): continue
            tag = r[1] if r[1] in ("Escolinha", "Seleção", "Educação Física", "Outro") else "Outro"
            self.tree_atv.insert("", "end", values=r, tags=(tag,))

    def _editar_atividade(self):
        sel = self.tree_atv.selection()
        if not sel: messagebox.showwarning("Aviso", "Selecione uma atividade."); return
        atv_id = self.tree_atv.item(sel[0])["values"][0]
        conn = sqlite3.connect(db_path); cur = conn.cursor()
        cur.execute("""SELECT tipo_atividade,nome_atividade,dia_semana,horario_inicio,horario_fim,
                              local,professor,estagiario,observacoes FROM atividades WHERE id=?""", (atv_id,))
        row = cur.fetchone(); conn.close()
        if not row: return

        win = tk.Toplevel(self.root); win.title("Editar Atividade"); win.geometry("520x580")
        win.configure(bg="white"); win.transient(self.root); win.grab_set()
        tk.Label(win, text="✏️  EDITAR ATIVIDADE", bg=C_PRIMARIO, fg="white",
                 font=("Segoe UI", 13, "bold"), pady=10).pack(fill="x")
        fm = tk.Frame(win, bg="white", padx=28); fm.pack(fill="both", expand=True)
        fm.grid_columnconfigure(0, weight=1)

        def lbl_cb_edit(txt, val, tabela, r):
            tk.Label(fm, text=txt, bg="white", font=("Segoe UI", 9, "bold"), fg=C_SECUNDARIO
                     ).grid(row=r, column=0, sticky="w", pady=(8, 0))
            fr2 = tk.Frame(fm, bg="white"); fr2.grid(row=r+1, column=0, sticky="ew", pady=(2, 0))
            v = tk.StringVar(value=val or "")
            cb = ttk.Combobox(fr2, textvariable=v, values=db_lista(tabela), font=("Segoe UI", 10))
            cb.pack(side="left", fill="x", expand=True)
            make_btn(fr2, "⚙️", lambda t=tabela, c=cb: self._abrir_lista_e_atualizar(t, [c]),
                     bg=C_SECUNDARIO, font_size=8, height=1, padx=5).pack(side="left", padx=(4, 0))
            return v

        tk.Label(fm, text="Tipo:", bg="white", font=("Segoe UI", 9, "bold"), fg=C_SECUNDARIO
                 ).grid(row=0, column=0, sticky="w", pady=(8, 0))
        v_tipo = tk.StringVar(value=row[0])
        ttk.Combobox(fm, textvariable=v_tipo,
                     values=["Escolinha", "Seleção", "Educação Física", "Outro"],
                     state="readonly").grid(row=1, column=0, sticky="ew", pady=(2, 0))

        v_nome = lbl_cb_edit("Nome / Turma:", row[1], "atv_nomes",  2)

        tk.Label(fm, text="Dia da Semana:", bg="white", font=("Segoe UI", 9, "bold"), fg=C_SECUNDARIO
                 ).grid(row=4, column=0, sticky="w", pady=(8, 0))
        v_dia = tk.StringVar(value=row[2])
        ttk.Combobox(fm, textvariable=v_dia,
                     values=["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira",
                             "Sexta-feira", "Sábado", "Domingo",
                             "Seg e Qua", "Ter e Qui", "Seg, Qua e Sex"],
                     state="readonly").grid(row=5, column=0, sticky="ew", pady=(2, 0))

        tk.Label(fm, text="Horário Início:", bg="white", font=("Segoe UI", 9, "bold"), fg=C_SECUNDARIO
                 ).grid(row=6, column=0, sticky="w", pady=(8, 0))
        fr_ini2 = tk.Frame(fm, bg="white"); fr_ini2.grid(row=7, column=0, sticky="ew", pady=(2, 0))
        v_ini = tk.StringVar(value=row[3] or "")
        cb_ini = ttk.Combobox(fr_ini2, textvariable=v_ini, values=db_lista("atv_horarios"), font=("Segoe UI", 10))
        cb_ini.pack(side="left", fill="x", expand=True)
        make_btn(fr_ini2, "⚙️", lambda: self._abrir_lista_e_atualizar("atv_horarios", [cb_ini]),
                 bg=C_SECUNDARIO, font_size=8, height=1, padx=5).pack(side="left", padx=(4, 0))

        tk.Label(fm, text="Horário Fim:", bg="white", font=("Segoe UI", 9, "bold"), fg=C_SECUNDARIO
                 ).grid(row=8, column=0, sticky="w", pady=(8, 0))
        fr_fim2 = tk.Frame(fm, bg="white"); fr_fim2.grid(row=9, column=0, sticky="ew", pady=(2, 0))
        v_fim = tk.StringVar(value=row[4] or "")
        cb_fim = ttk.Combobox(fr_fim2, textvariable=v_fim, values=db_lista("atv_horarios"), font=("Segoe UI", 10))
        cb_fim.pack(side="left", fill="x", expand=True)
        make_btn(fr_fim2, "⚙️", lambda: self._abrir_lista_e_atualizar("atv_horarios", [cb_fim]),
                 bg=C_SECUNDARIO, font_size=8, height=1, padx=5).pack(side="left", padx=(4, 0))

        v_local = lbl_cb_edit("Local / Quadra:", row[5], "atv_quadras", 10)
        v_prof  = lbl_cb_edit("Professor:",      row[6], "atv_profs",   12)
        v_estag = lbl_cb_edit("Estagiário:",     row[7], "atv_estags",  14)

        tk.Label(fm, text="Observações:", bg="white", font=("Segoe UI", 9, "bold"), fg=C_SECUNDARIO
                 ).grid(row=16, column=0, sticky="w", pady=(8, 0))
        e_obs = tk.Text(fm, font=("Segoe UI", 9), bd=1, relief="solid", height=3, wrap="word")
        e_obs.insert("1.0", row[8] or ""); e_obs.grid(row=17, column=0, sticky="ew", pady=(2, 0))

        def salvar(ev=None):
            if not v_nome.get().strip() or not v_prof.get().strip():
                messagebox.showwarning("Aviso", "Nome e Professor são obrigatórios.", parent=win); return
            conn2 = sqlite3.connect(db_path); c2 = conn2.cursor()
            c2.execute("""UPDATE atividades SET tipo_atividade=?,nome_atividade=?,dia_semana=?,
                                 horario_inicio=?,horario_fim=?,local=?,professor=?,estagiario=?,observacoes=?
                          WHERE id=?""",
                       (v_tipo.get(), v_nome.get().strip(), v_dia.get(),
                        v_ini.get().strip(), v_fim.get().strip(),
                        v_local.get().strip(), v_prof.get().strip(),
                        v_estag.get().strip(), e_obs.get("1.0", "end-1c").strip(), atv_id))
            conn2.commit(); conn2.close(); win.destroy(); self.refresh_atividades()

        make_btn(win, "SALVAR ALTERAÇÕES", salvar, bg=C_PRIMARIO, font_size=11, height=1
                 ).pack(fill="x", padx=28, pady=14)

    def _excluir_atividade(self):
        sel = self.tree_atv.selection()
        if not sel: messagebox.showwarning("Aviso", "Selecione uma atividade."); return
        vals = self.tree_atv.item(sel[0])["values"]
        if messagebox.askyesno("Confirmar", f"Excluir '{vals[2]}'?"):
            conn = sqlite3.connect(db_path); c = conn.cursor()
            c.execute("DELETE FROM atividades WHERE id=?", (vals[0],))
            conn.commit(); conn.close(); self.refresh_atividades()

    def _gerar_excel_atividades(self):
        filtro = self.atv_filtro_var.get()
        conn = sqlite3.connect(db_path); cur = conn.cursor()
        if filtro == "Todos":
            cur.execute("""SELECT tipo_atividade,nome_atividade,dia_semana,horario_inicio,horario_fim,
                                  local,professor,estagiario,observacoes
                           FROM atividades ORDER BY tipo_atividade,dia_semana,horario_inicio""")
        else:
            cur.execute("""SELECT tipo_atividade,nome_atividade,dia_semana,horario_inicio,horario_fim,
                                  local,professor,estagiario,observacoes
                           FROM atividades WHERE tipo_atividade=?
                           ORDER BY dia_semana,horario_inicio""", (filtro,))
        rows = cur.fetchall(); conn.close()
        if not rows: messagebox.showwarning("Aviso", "Nenhuma atividade para exportar."); return

        wb = openpyxl.Workbook()
        cores_tipo = {"Escolinha": ("1A5F7A", "E8F4F8"), "Seleção": ("1A7A3C", "E8F5E9"),
                      "Educação Física": ("6A1A7A", "F3E8F8"), "Outro": ("7A4A1A", "FDF3E3")}
        azul = PatternFill("solid", fgColor="221C89")
        borda = Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"), bottom=Side(style="thin"))
        aln_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
        aln_l = Alignment(horizontal="left", vertical="center", wrap_text=True)
        headers = ["Tipo", "Nome / Turma", "Dia da Semana", "Início", "Fim",
                   "Local / Quadra", "Professor", "Estagiário", "Observações"]
        larguras = {"A": 15, "B": 28, "C": 18, "D": 10, "E": 10,
                    "F": 22, "G": 24, "H": 22, "I": 30}

        ws = wb.active; ws.title = "Todas as Atividades"
        ws.merge_cells("A1:I1"); ws["A1"] = NOME_ESCOLA
        ws["A1"].font = Font(name="Arial", bold=True, size=14, color="221C89"); ws["A1"].alignment = aln_c
        ws.merge_cells("A2:I2")
        ws["A2"] = f"GRADE DE ATIVIDADES{' — ' + filtro if filtro != 'Todos' else ''}"
        ws["A2"].font = Font(name="Arial", bold=True, size=11, color="4A4A6A"); ws["A2"].alignment = aln_c
        ws.merge_cells("A3:I3")
        ws["A3"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total: {len(rows)}"
        ws["A3"].font = Font(name="Arial", italic=True, size=9, color="888888"); ws["A3"].alignment = aln_c
        ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 18; ws.append([])
        ws.append(headers); hr = ws.max_row; ws.row_dimensions[hr].height = 26
        for i, h in enumerate(headers, 1):
            cel = ws.cell(hr, i); cel.value = h; cel.fill = azul; cel.border = borda; cel.alignment = aln_c
            cel.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        for r in rows:
            ws.append(list(r)); rn = ws.max_row; ws.row_dimensions[rn].height = 20
            fg_h, bg_h = cores_tipo.get(r[0], ("000000", "FFFFFF"))
            fill_r = PatternFill("solid", fgColor=bg_h); font_r = Font(name="Arial", size=10, color=fg_h)
            for ci in range(1, 10):
                cel = ws.cell(rn, ci); cel.fill = fill_r; cel.font = font_r; cel.border = borda
                cel.alignment = aln_l if ci in (2, 6, 7, 8, 9) else aln_c
        for col, w2 in larguras.items(): ws.column_dimensions[col].width = w2
        ws.append([]); ws.append([f"Total: {len(rows)}"])
        ws.cell(ws.max_row, 1).font = Font(name="Arial", bold=True, size=10, color="221C89")

        tipos_unicos = sorted(set(r[0] for r in rows))
        for tipo in tipos_unicos:
            ws2 = wb.create_sheet(title=tipo[:30])
            fg_h, bg_h = cores_tipo.get(tipo, ("000000", "FFFFFF"))
            ws2.merge_cells("A1:I1"); ws2["A1"] = f"{NOME_ESCOLA} — {tipo.upper()}"
            ws2["A1"].font = Font(name="Arial", bold=True, size=13, color=fg_h); ws2["A1"].alignment = aln_c
            ws2.append([]); ws2.append(headers); hr2 = ws2.max_row; ws2.row_dimensions[hr2].height = 26
            fill_hdr2 = PatternFill("solid", fgColor=fg_h)
            for i, h in enumerate(headers, 1):
                cel = ws2.cell(hr2, i); cel.value = h; cel.fill = fill_hdr2; cel.border = borda
                cel.alignment = aln_c; cel.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            cnt = 0
            for r in rows:
                if r[0] != tipo: continue
                ws2.append(list(r)); rn2 = ws2.max_row; ws2.row_dimensions[rn2].height = 20
                fill_r2 = PatternFill("solid", fgColor=bg_h); font_r2 = Font(name="Arial", size=10, color=fg_h)
                for ci in range(1, 10):
                    cel = ws2.cell(rn2, ci); cel.fill = fill_r2; cel.font = font_r2; cel.border = borda
                    cel.alignment = aln_l if ci in (2, 6, 7, 8, 9) else aln_c
                cnt += 1
            for col, w2 in larguras.items(): ws2.column_dimensions[col].width = w2
            ws2.append([]); ws2.append([f"Total de {tipo}: {cnt}"])
            ws2.cell(ws2.max_row, 1).font = Font(name="Arial", bold=True, size=10, color=fg_h)

        fn = f"Atividades_C7S_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(os.path.join(PASTA_DIST, fn))
        messagebox.showinfo("✅ Excel Gerado!",
                            f"Arquivo: dist/{fn}\nTotal: {len(rows)}  |  Tipos: {len(tipos_unicos)}")

    # ═══════════════════════════════════════════════════════════
    # ABA COMUNICAÇÃO
    # ═══════════════════════════════════════════════════════════
    def setup_tab_comunicacao(self):
        tab=ttk.Frame(self.tabs); self.tabs.add(tab,text="  📣 COMUNICAÇÃO  ")
        mf=tk.Frame(tab,bg="white",padx=20,pady=12); mf.pack(expand=True,fill="both")
        tk.Label(mf,text="📣  GERADOR DE E-MAILS / MENSAGENS",font=("Segoe UI",14,"bold"),bg="white",fg=C_PRIMARIO).pack(anchor="w")
        tk.Label(mf,text="Puxa TODOS os jogos futuros. Edite antes de enviar.",bg="white",fg=C_SECUNDARIO,font=("Segoe UI",10)).pack(anchor="w",pady=(2,10))

        cfg=tk.LabelFrame(mf,text=" ✏️  Assinaturas (editáveis) ",bg="white",font=("Segoe UI",9,"bold"),fg=C_PRIMARIO)
        cfg.pack(fill="x",pady=(0,10))
        cd=tk.Frame(cfg,bg="white"); cd.pack(side="left",expand=True,fill="both",padx=8,pady=6)
        tk.Label(cd,text="Assinatura — Direção:",bg="white",font=("Segoe UI",8,"bold"),fg=C_SECUNDARIO).pack(anchor="w")
        self.txt_ass_dir=tk.Text(cd,height=3,font=("Consolas",9),bd=1,relief="solid",wrap="word"); self.txt_ass_dir.pack(fill="x")
        ct=tk.Frame(cfg,bg="white"); ct.pack(side="left",expand=True,fill="both",padx=8,pady=6)
        tk.Label(ct,text="Assinatura — Transbaby:",bg="white",font=("Segoe UI",8,"bold"),fg=C_SECUNDARIO).pack(anchor="w")
        self.txt_ass_trn=tk.Text(ct,height=3,font=("Consolas",9),bd=1,relief="solid",wrap="word"); self.txt_ass_trn.pack(fill="x")
        make_btn(cfg,"💾  SALVAR",self.salvar_assinaturas,bg=C_SUCESSO,font_size=9,height=1).pack(side="right",padx=8,pady=6)
        conn=sqlite3.connect(db_path); cur=conn.cursor()
        for chave,widget in [("assinatura_direcao",self.txt_ass_dir),("assinatura_trans",self.txt_ass_trn)]:
            cur.execute("SELECT valor FROM config WHERE chave=?",(chave,)); r=cur.fetchone()
            if r: widget.insert("1.0",r[0])
        conn.close()

        txf=tk.Frame(mf,bg="white",highlightbackground=C_ACENTO,highlightthickness=1); txf.pack(fill="both",expand=True)
        self.txt_output=tk.Text(txf,height=15,font=("Consolas",10),bg="#fafbfc",bd=0,padx=12,pady=10,wrap="word",undo=True)
        stx=ttk.Scrollbar(txf,orient="vertical",command=self.txt_output.yview)
        self.txt_output.configure(yscrollcommand=stx.set)
        self.txt_output.pack(side="left",fill="both",expand=True); stx.pack(side="right",fill="y")
        self.txt_output.bind("<Control-a>",lambda e:(self.txt_output.tag_add("sel","1.0","end"),"break"))

        bf=tk.Frame(mf,bg="white",pady=10); bf.pack(fill="x")
        for i in range(4): bf.grid_columnconfigure(i,weight=1)
        btns_com=[
            ("📧 E-MAIL DIREÇÃO",   self.gerar_txt_direcao, C_PRIMARIO),
            ("🚌 E-MAIL TRANSBABY", self.gerar_txt_trans,   C_ACENTO),
            ("📋 COPIAR TEXTO",     self.copiar_txt,        C_SECUNDARIO),
            ("🗑️ LIMPAR",           lambda:self.txt_output.delete("1.0","end"), "#999"),
        ]
        for i,(txt,cmd,cor) in enumerate(btns_com):
            fg2=C_PRIMARIO if cor==C_ACENTO else "white"
            make_btn(bf,txt,cmd,bg=cor,fg=fg2,font_size=9,height=1,padx=12
                     ).grid(row=0,column=i,padx=8,sticky="ew")

    def salvar_assinaturas(self):
        conn=sqlite3.connect(db_path); c=conn.cursor()
        c.execute("INSERT OR REPLACE INTO config VALUES ('assinatura_direcao',?)",(self.txt_ass_dir.get("1.0","end-1c").strip(),))
        c.execute("INSERT OR REPLACE INTO config VALUES ('assinatura_trans',?)",  (self.txt_ass_trn.get("1.0","end-1c").strip(),))
        conn.commit(); conn.close(); messagebox.showinfo("Salvo","Assinaturas salvas!")

    def copiar_txt(self):
        self.root.clipboard_clear(); self.root.clipboard_append(self.txt_output.get("1.0","end-1c"))
        messagebox.showinfo("Copiado!","Texto copiado!")

    # ═══════════════════════════════════════════════════════════
    # ABA CONFIGURAÇÕES
    # ═══════════════════════════════════════════════════════════
    def setup_tab_configuracoes(self):
        tab=ttk.Frame(self.tabs); self.tabs.add(tab,text="  ⚙️ CONFIG  ")
        mf=tk.Frame(tab,bg=C_FUNDO,padx=30,pady=20); mf.pack(fill="both",expand=True)
        tk.Label(mf,text="⚙️  CONFIGURAÇÕES DO SISTEMA",bg=C_FUNDO,fg=C_PRIMARIO,font=("Segoe UI",14,"bold")).pack(anchor="w")
        tk.Label(mf,text="Gerencie usuários, listas e dados do sistema.",bg=C_FUNDO,fg=C_SECUNDARIO,font=("Segoe UI",10)).pack(anchor="w",pady=(2,16))

        uf=tk.LabelFrame(mf,text=" 👤 Gerenciar Usuários ",bg=C_FUNDO,font=("Segoe UI",10,"bold"),fg=C_PRIMARIO)
        uf.pack(fill="x",pady=(0,14))
        self.tree_users=ttk.Treeview(uf,columns=("Usuário",),show="headings",height=5)
        self.tree_users.heading("Usuário",text="Usuário cadastrado"); self.tree_users.column("Usuário",width=300,anchor="w")
        sc_u=ttk.Scrollbar(uf,orient="vertical",command=self.tree_users.yview)
        self.tree_users.configure(yscrollcommand=sc_u.set)
        self.tree_users.pack(side="left",fill="both",expand=True,padx=(8,0),pady=8)
        sc_u.pack(side="right",fill="y",pady=8,padx=(0,8))
        bf_u=tk.Frame(uf,bg=C_FUNDO); bf_u.pack(side="right",fill="y",padx=8,pady=8)
        make_btn(bf_u,"➕ NOVO USUÁRIO",     self.tela_cadastro,          bg=C_PRIMARIO,  font_size=9,height=1).pack(fill="x",pady=(0,4))
        make_btn(bf_u,"🔑 ALTERAR SENHA",   self._alterar_senha_usuario, bg=C_SECUNDARIO,font_size=9,height=1).pack(fill="x",pady=(0,4))
        make_btn(bf_u,"🗑️ EXCLUIR USUÁRIO", self._excluir_usuario,       bg=C_PERIGO,    font_size=9,height=1).pack(fill="x")
        self._carregar_usuarios()

        lf2=tk.LabelFrame(mf,text=" 📋 Listas Globais (Modalidades / Categorias / Naipes) ",
                          bg=C_FUNDO,font=("Segoe UI",10,"bold"),fg=C_PRIMARIO)
        lf2.pack(fill="x",pady=(0,14))
        tk.Label(lf2,text="Adicione, renomeie ou exclua as listas globais usadas em Atletas, Jogos e Convocações.",
                 bg=C_FUNDO,fg=C_SECUNDARIO,font=("Segoe UI",9)).pack(anchor="w",padx=10,pady=(6,4))
        make_btn(lf2,"⚙️  ABRIR GERENCIADOR DE LISTAS GLOBAIS",self.tela_gerenciar_listas,
                 bg=C_PRIMARIO,font_size=10,height=1).pack(fill="x",padx=10,pady=(0,10))

        lf3=tk.LabelFrame(mf,text=" 🏃 Listas de Atividades (Profs / Quadras / Horários) ",
                          bg=C_FUNDO,font=("Segoe UI",10,"bold"),fg=C_PRIMARIO)
        lf3.pack(fill="x",pady=(0,14))
        tk.Label(lf3,text="Gerencie os menus de seleção que aparecem no formulário de Atividades.",
                 bg=C_FUNDO,fg=C_SECUNDARIO,font=("Segoe UI",9)).pack(anchor="w",padx=10,pady=(6,4))
        make_btn(lf3,"⚙️  ABRIR LISTAS DE ATIVIDADES",self._gerenciar_listas_atividades,
                 bg=C_ACENTO,fg=C_PRIMARIO,font_size=10,height=1).pack(fill="x",padx=10,pady=(0,10))

        dbf=tk.LabelFrame(mf,text=" 🗄️ Banco de Dados ",bg=C_FUNDO,font=("Segoe UI",10,"bold"),fg=C_PRIMARIO)
        dbf.pack(fill="x",pady=(0,14))
        tk.Label(dbf,text=f"Arquivo: {db_path}",bg=C_FUNDO,fg=C_SECUNDARIO,font=("Segoe UI",8)).pack(anchor="w",padx=10,pady=(6,4))
        bf_db=tk.Frame(dbf,bg=C_FUNDO); bf_db.pack(fill="x",padx=10,pady=(0,10))
        make_btn(bf_db,"📂 ABRIR PASTA DIST",       self._abrir_pasta_dist,   bg=C_SECUNDARIO,font_size=9,height=1).pack(side="left",padx=(0,8))
        make_btn(bf_db,"⚠️ LIMPAR DADOS DE TESTE",  self._limpar_dados_teste, bg=C_PERIGO,    font_size=9,height=1).pack(side="left")

        inf=tk.LabelFrame(mf,text=" ℹ️ Sobre o Sistema ",bg=C_FUNDO,font=("Segoe UI",10,"bold"),fg=C_PRIMARIO)
        inf.pack(fill="x")
        for i,(k,v) in enumerate([("Sistema",NOME_SISTEMA),("Escola",NOME_ESCOLA),("Versão","V12 — Abril 2026"),("Banco","SGE_MasterPro_V12.db")]):
            tk.Label(inf,text=f"{k}:",bg=C_FUNDO,fg=C_SECUNDARIO,font=("Segoe UI",9,"bold")).grid(row=i,column=0,sticky="w",padx=14,pady=2)
            tk.Label(inf,text=v,bg=C_FUNDO,fg=C_PRIMARIO,font=("Segoe UI",9)).grid(row=i,column=1,sticky="w",padx=6,pady=2)

    def _carregar_usuarios(self):
        for i in self.tree_users.get_children(): self.tree_users.delete(i)
        conn=sqlite3.connect(db_path); c=conn.cursor()
        c.execute("SELECT user FROM usuarios ORDER BY user")
        for (u,) in c.fetchall(): self.tree_users.insert("","end",values=(u,))
        conn.close()

    def _alterar_senha_usuario(self):
        sel=self.tree_users.selection()
        if not sel: messagebox.showwarning("Aviso","Selecione um usuário."); return
        usuario=self.tree_users.item(sel[0])["values"][0]
        w=tk.Toplevel(self.root); w.title("Alterar Senha"); w.geometry("360x220")
        w.configure(bg="white"); w.transient(self.root); w.grab_set()
        tk.Label(w,text=f"ALTERAR SENHA — {usuario}",bg=C_PRIMARIO,fg="white",
                 font=("Segoe UI",12,"bold"),pady=10).pack(fill="x")
        fm=tk.Frame(w,bg="white",padx=30,pady=10); fm.pack()
        tk.Label(fm,text="Nova Senha:",bg="white",font=("Segoe UI",10)).pack(anchor="w")
        e1=tk.Entry(fm,font=("Segoe UI",11),bd=1,relief="solid",show="*",width=24); e1.pack(pady=(0,8))
        tk.Label(fm,text="Confirmar Senha:",bg="white",font=("Segoe UI",10)).pack(anchor="w")
        e2=tk.Entry(fm,font=("Segoe UI",11),bd=1,relief="solid",show="*",width=24); e2.pack(pady=(0,14))
        def salvar(ev=None):
            s1,s2=e1.get().strip(),e2.get().strip()
            if not s1: return
            if s1!=s2: messagebox.showerror("Erro","Senhas não conferem.",parent=w); return
            conn=sqlite3.connect(db_path); c=conn.cursor()
            c.execute("UPDATE usuarios SET senha=? WHERE user=?",(s1,usuario))
            conn.commit(); conn.close(); messagebox.showinfo("Sucesso","Senha alterada!",parent=w); w.destroy()
        e1.bind("<Return>",lambda e: e2.focus()); e2.bind("<Return>",salvar)
        make_btn(w,"SALVAR",salvar,bg=C_PRIMARIO,font_size=11,height=1).pack(fill="x",padx=30)
        e1.focus()

    def _excluir_usuario(self):
        sel=self.tree_users.selection()
        if not sel: messagebox.showwarning("Aviso","Selecione um usuário."); return
        usuario=self.tree_users.item(sel[0])["values"][0]
        conn=sqlite3.connect(db_path); c=conn.cursor()
        c.execute("SELECT COUNT(*) FROM usuarios"); total=c.fetchone()[0]; conn.close()
        if total<=1: messagebox.showerror("Erro","Não pode excluir o único usuário."); return
        if messagebox.askyesno("Confirmar",f"Excluir '{usuario}'?"):
            conn=sqlite3.connect(db_path); c=conn.cursor()
            c.execute("DELETE FROM usuarios WHERE user=?",(usuario,))
            conn.commit(); conn.close(); self._carregar_usuarios()

    def _abrir_pasta_dist(self):
        import subprocess,platform
        try:
            if platform.system()=="Windows": os.startfile(PASTA_DIST)
            elif platform.system()=="Darwin": subprocess.Popen(["open",PASTA_DIST])
            else: subprocess.Popen(["xdg-open",PASTA_DIST])
        except: messagebox.showinfo("Pasta",f"Arquivos em:\n{PASTA_DIST}")

    def _limpar_dados_teste(self):
        if not messagebox.askyesno("⚠️ ATENÇÃO","Excluir TODOS os dados? (atletas, jogos, atividades, estoque...)"):
            return
        if not messagebox.askyesno("⚠️ CONFIRMAR","ÚLTIMA CHANCE! Continuar?"): return
        conn=sqlite3.connect(db_path); c=conn.cursor()
        for t in ["atletas","tecnicos","jogos","convocacoes","estoque","estoque_variantes","atividades"]:
            c.execute(f"DELETE FROM {t}")
        conn.commit(); conn.close()
        messagebox.showinfo("Concluído","Dados limpos.")
        self.refresh_all()

    # ═══════════════════════════════════════════════════════════
    # REFRESH GERAL
    # ═══════════════════════════════════════════════════════════
    def refresh_all(self):
        m,c,n = self.mod_sel.get(),self.cat_sel.get(),self.nai_sel.get()
        ver_todos = self.ver_todos_jogos.get()

        # Atletas
        for i in self.tree_at.get_children(): self.tree_at.delete(i)
        conn=sqlite3.connect(db_path); cur=conn.cursor()
        cur.execute("SELECT id,nome,mod,cat,naipe FROM atletas WHERE mod=? AND cat=? AND naipe=?",(m,c,n))
        for r in cur.fetchall(): self.tree_at.insert("","end",values=r)

        # Técnicos
        for i in self.tree_tec.get_children(): self.tree_tec.delete(i)
        cur.execute("SELECT id,nome,cargo,mod,cat,naipe FROM tecnicos WHERE mod=? AND cat=? AND naipe=?",(m,c,n))
        for r in cur.fetchall(): self.tree_tec.insert("","end",values=r)

        # Jogos
        for i in self.tree_jg.get_children(): self.tree_jg.delete(i)
        if ver_todos:
            cur.execute("""SELECT id,data,hora,mod,cat,adv,comp,tipo,local,saida,retorno,
                                  placar_c7s||CASE WHEN placar_c7s!='' THEN ' x ' ELSE '' END||placar_adv FROM jogos""")
        else:
            cur.execute("""SELECT id,data,hora,mod,cat,adv,comp,tipo,local,saida,retorno,
                                  placar_c7s||CASE WHEN placar_c7s!='' THEN ' x ' ELSE '' END||placar_adv
                           FROM jogos WHERE mod=? AND cat=?""",(m,c))
        jogos_lst=sorted(cur.fetchall(),key=lambda r: parse_data(str(r[1]))); conn.close()
        for r in jogos_lst: self.tree_jg.insert("","end",values=r)
        n_casa=sum(1 for r in jogos_lst if r[7]=="CASA"); n_fora=sum(1 for r in jogos_lst if r[7]=="FORA")
        self.lbl_casa.config(text=f"🏠 Em Casa: {n_casa}")
        self.lbl_fora.config(text=f"✈️ Fora: {n_fora}")
        self.lbl_total.config(text=f"📋 Total: {len(jogos_lst)}")
        self.atualizar_combo_convocacao()
        if hasattr(self,"tree_users"): self._carregar_usuarios()
        if hasattr(self,"tree_atv"):   self.refresh_atividades()

    def atualizar_combo_convocacao(self):
        m,c=self.mod_sel.get(),self.cat_sel.get()
        conn=sqlite3.connect(db_path); cur=conn.cursor()
        cur.execute("SELECT id,data,hora,adv FROM jogos WHERE mod=? AND cat=? ORDER BY data ASC",(m,c))
        jogos=cur.fetchall(); conn.close()
        self.conv_jogos_map={f"{j[1]} {j[2]} — C7S x {j[3]}":j[0] for j in jogos}
        self.conv_jogo_cb["values"]=list(self.conv_jogos_map.keys())

    # ═══════════════════════════════════════════════════════════
    # GERAÇÃO DE TEXTOS
    # ═══════════════════════════════════════════════════════════
    def _get_todos_jogos_futuros(self):
        conn=sqlite3.connect(db_path); cur=conn.cursor()
        cur.execute("""SELECT id,data,hora,adv,local,tipo,comp,saida,retorno,mod,cat,
                              placar_c7s||CASE WHEN placar_c7s!='' THEN ' x ' ELSE '' END||placar_adv FROM jogos""")
        todos=cur.fetchall(); conn.close()
        futuros=[]; hoje=datetime.now().date()
        for r in todos:
            try:
                dt=datetime.strptime(f"{r[1]} {r[2]}","%d/%m/%Y %H:%M")
                if dt.date()>=hoje:
                    futuros.append({"id":r[0],"data":r[1],"hora":r[2],"adv":r[3],"local":r[4],
                                    "tipo":r[5],"comp":r[6],"saida_ant":r[7],"retorno":r[8],
                                    "mod":r[9],"cat":r[10],"placar":r[11],"dt_obj":dt})
            except: pass
        futuros.sort(key=lambda x: x["dt_obj"]); return futuros

    def _calc_hora_saida(self,hora_jogo,ant):
        if ant in ["1h antes","2h antes"]:
            try:
                h=datetime.strptime(hora_jogo,"%H:%M")-timedelta(hours=1 if ant=="1h antes" else 2)
                return h.strftime("%H:%M")
            except: pass
        return hora_jogo

    def gerar_txt_direcao(self):
        jogos=self._get_todos_jogos_futuros(); total=len(jogos)
        tem_casa=any(j["tipo"]=="CASA" for j in jogos); tem_fora=any(j["tipo"]=="FORA" for j in jogos)
        desc="dentro e fora dos nossos domínios" if tem_casa and tem_fora else ("dentro dos nossos domínios" if tem_casa else ("fora dos nossos domínios" if tem_fora else "previstos"))
        txt=f"Prezados,\n\nInformamos que teremos {total} jogo{'s' if total!=1 else ''} essa semana, {desc}.\n\n"
        if not jogos: txt+="Não há jogos futuros cadastrados no momento.\n\n"
        else:
            for secao,tipo_f in [("JOGOS EM CASA","CASA"),("JOGOS FORA DE CASA","FORA")]:
                bloco=[j for j in jogos if j["tipo"]==tipo_f]
                if not bloco: continue
                txt+=f"━━━  {secao}  ━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                for j in bloco:
                    hs=self._calc_hora_saida(j["hora"],j["saida_ant"])
                    ret=j["retorno"] if j["retorno"] and j["retorno"] not in ("A definir","") else "A confirmar"
                    txt+=f"  Data          : {j['data']}\n  Horário       : {j['hora']}h\n"
                    txt+=f"  Modalidade    : {j['mod']} {j['cat']}\n  Competição    : {j['comp']}\n"
                    txt+=f"  Confronto     : C7S x {j['adv']}\n  Local         : {j['local']}\n"
                    if tipo_f=="FORA": txt+=f"  Saída (C7S)   : {hs}h\n"
                    txt+=f"  Retorno       : {ret}\n\n"
        txt+=self.txt_ass_dir.get("1.0","end-1c").strip()
        self._exibir_e_copiar(txt)

    def gerar_txt_trans(self):
        jogos=[j for j in self._get_todos_jogos_futuros() if j["tipo"]=="FORA"]
        txt="Prezados,\n\nGostaríamos de solicitar os transportes da semana:\n\n"
        if not jogos: txt+="Não há jogos fora de casa previstos.\n\n"
        else:
            for i,j in enumerate(jogos,1):
                hs=self._calc_hora_saida(j["hora"],j["saida_ant"])
                ret=j["retorno"] if j["retorno"] and j["retorno"] not in ("A definir","") else "A confirmar"
                txt+=f"{'─'*44}\nVIAGEM {i}\n\n"
                txt+=f"  Jogo              : C7S x {j['adv']}\n  Modalidade        : {j['mod']} {j['cat']}\n"
                txt+=f"  Competição        : {j['comp']}\n  Data              : {j['data']}\n"
                txt+=f"  Saída (C7S)       : {hs}h\n  Destino           : {j['local']}\n"
                txt+=f"  Horário do Jogo   : {j['hora']}h\n  Retorno           : {ret}\n\n"
        txt+=f"{'─'*44}\n\n{self.txt_ass_trn.get('1.0','end-1c').strip()}"
        self._exibir_e_copiar(txt)

    def _exibir_e_copiar(self,txt):
        self.txt_output.delete("1.0","end"); self.txt_output.insert("1.0",txt)
        self.root.clipboard_clear(); self.root.clipboard_append(txt)
        messagebox.showinfo("✅ Pronto!","Texto gerado e copiado!\nEdite antes de enviar se quiser.")

    # ═══════════════════════════════════════════════════════════
    # PDFs
    # ═══════════════════════════════════════════════════════════
    def gerar_pdf_elite(self,tipo):
        m,c,n=self.mod_sel.get(),self.cat_sel.get(),self.nai_sel.get()
        conn=sqlite3.connect(db_path); cur=conn.cursor()
        cur.execute("SELECT nome FROM atletas WHERE mod=? AND cat=? AND naipe=? ORDER BY nome ASC",(m,c,n))
        atletas=cur.fetchall()
        cur.execute("SELECT nome,cargo FROM tecnicos WHERE mod=? AND cat=? AND naipe=? ORDER BY nome ASC",(m,c,n))
        tecs=cur.fetchall(); conn.close()
        if not atletas: return messagebox.showwarning("Aviso","Nenhum atleta na seleção ativa.")
        pdf=PDF_Elite(); pdf.alias_nb_pages(); pdf.add_page(); pdf.add_watermark()
        t_doc="LISTA OFICIAL VISITANTE" if tipo=="visitante" else "DISPENSA - EDUCACAO FISICA"
        pdf.set_font("Arial","B",14); pdf.set_text_color(34,28,137); pdf.cell(190,10,t_doc,ln=True,align="C")
        pdf.set_font("Arial","B",11); pdf.set_text_color(74,74,106)
        pdf.cell(190,8,f"Selecao: {m.upper()}  .  Categoria: {c}  .  Naipe: {n}",ln=True,align="C"); pdf.ln(8)
        if tecs:
            pdf.set_font("Arial","B",10); pdf.set_fill_color(242,179,26); pdf.set_text_color(34,28,137)
            pdf.cell(190,8,"COMISSAO TECNICA",1,1,"C",True); pdf.set_font("Arial","",10); pdf.set_text_color(0,0,0)
            for tnome,tcargo in tecs: pdf.cell(130,9,f"  {tnome}",1,0,"L"); pdf.cell(60,9,f"  {tcargo}",1,1,"L")
            pdf.ln(4)
        pdf.set_font("Arial","B",10); pdf.set_fill_color(34,28,137); pdf.set_text_color(255,255,255)
        pdf.cell(20,10,"No",1,0,"C",True); pdf.cell(170,10,"NOME COMPLETO DO ATLETA",1,1,"L",True)
        pdf.set_font("Arial","",10); pdf.set_text_color(0,0,0); z=False
        for i,(nome,) in enumerate(atletas,1):
            pdf.set_fill_color(244,247,246) if z else pdf.set_fill_color(255,255,255)
            pdf.cell(20,10,str(i),1,0,"C",True); pdf.cell(170,10,f"  {nome}",1,1,"L",True); z=not z
        fn=f"{'Visitante' if tipo=='visitante' else 'DispensaEF'}_{m}_{c}.pdf"
        try: pdf.output(os.path.join(PASTA_DIST,fn)); messagebox.showinfo("Sucesso",f"dist/{fn}")
        except: messagebox.showerror("Erro","Feche o PDF antes.")

    def gerar_pdf_agenda(self):
        m,c=self.mod_sel.get(),self.cat_sel.get(); ver_todos=self.ver_todos_jogos.get()
        conn=sqlite3.connect(db_path); cur=conn.cursor()
        if ver_todos: cur.execute("SELECT data,hora,adv,comp,tipo,local,saida,retorno,placar_c7s,placar_adv,mod,cat FROM jogos")
        else: cur.execute("SELECT data,hora,adv,comp,tipo,local,saida,retorno,placar_c7s,placar_adv,mod,cat FROM jogos WHERE mod=? AND cat=?",(m,c))
        jogos=sorted(cur.fetchall(),key=lambda r: parse_data(str(r[0]))); conn.close()
        if not jogos: return messagebox.showwarning("Aviso","Nenhum jogo cadastrado.")
        pdf=PDF_Elite(); pdf.alias_nb_pages(); pdf.add_page(); pdf.add_watermark()
        pdf.set_font("Arial","B",14); pdf.set_text_color(34,28,137); pdf.cell(190,10,"AGENDA DE JOGOS E COMPETICOES",ln=True,align="C")
        pdf.set_font("Arial","B",11); pdf.set_text_color(74,74,106)
        pdf.cell(190,8,"TODAS AS MODALIDADES" if ver_todos else f"Selecao: {m.upper()}  .  Categoria: {c}",ln=True,align="C"); pdf.ln(6)
        pdf.set_font("Arial","B",9); pdf.set_fill_color(34,28,137); pdf.set_text_color(255,255,255)
        for h,w2 in [("Data",24),("Hora",16),("Mod",18),("Cat",16),("Adversario",36),("Competicao",36),("Mando",16),("Local",26),("Retorno",14),("Placar",18)]:
            pdf.cell(w2,10,h,1,0,"C",True)
        pdf.ln(); pdf.set_font("Arial","",8); pdf.set_text_color(0,0,0); z=False
        for r in jogos:
            data,hora,adv,comp,tipo,local,saida,retorno,pc,pa,mod2,cat2=r
            pdf.set_fill_color(244,247,246) if z else pdf.set_fill_color(255,255,255)
            placar=f"{pc}x{pa}" if pc else "-"
            for val,w2 in [(data,24),(hora,16),(mod2,18),(cat2,16),(adv,36),(comp,36),(tipo,16),(local,26),(retorno or "-",14),(placar,18)]:
                pdf.cell(w2,9,str(val)[:16],1,0,"C",True)
            pdf.ln(); z=not z
        fn=f"Agenda_{'TODAS' if ver_todos else f'{m}_{c}'}.pdf"
        try: pdf.output(os.path.join(PASTA_DIST,fn)); messagebox.showinfo("Sucesso",f"dist/{fn}")
        except: messagebox.showerror("Erro","Feche o PDF antes.")


if __name__ == "__main__":
    try:
        init_db(); root=tk.Tk(); app=SGEMasterPro(root); root.mainloop()
    except Exception as e:
        with open("ERRO_SISTEMA.txt","w") as f: f.write(traceback.format_exc())
        try:
            er=tk.Tk(); er.withdraw()
            messagebox.showerror("Erro Fatal",f"Sistema fechou.\n{e}"); er.destroy()
        except: pass